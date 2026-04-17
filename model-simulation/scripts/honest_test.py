#!/usr/bin/env python3
"""
HONEST TESTING — Two real-world validations:

Test 1: Ball outcome prediction accuracy
  For each event (4, 6, W) in captured matches, predict the most likely
  outcome from our model. Score: correct/total.

Test 2: Win probability vs Polymarket
  For each event in the 2nd innings, compare our DP model's chase win
  probability against the actual Polymarket mid-price.
  Measure: RMSE, correlation, mean absolute error.

Uses the 3 latest captured matches with the most events.
"""
import json
import os
import re
import sys
import time
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
os.chdir(ROOT)

CAPTURES = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures")


def load_dp_and_model():
    """Load the trained DP table and LightGBM model."""
    from src.dp.solver import DPTable
    from src.dp.states import TransitionProbs
    import lightgbm as lgb

    # Load LightGBM
    model_path = ROOT / "models" / "lgbm_ball_outcome.txt"
    if model_path.exists():
        model = lgb.Booster(model_file=str(model_path))
        print(f"Loaded LightGBM from {model_path}")
    else:
        model = None
        print("No LightGBM model found — using base DP only")

    # Load deliveries for feature lookup
    del_path = ROOT / "data" / "deliveries.parquet"
    df = pd.read_parquet(del_path) if del_path.exists() else None

    # Build ML-DP table
    batting_stats = pd.read_parquet(ROOT / "data" / "features" / "batting_phase_stats.parquet")
    bowling_stats = pd.read_parquet(ROOT / "data" / "features" / "bowling_phase_stats.parquet")

    from src.transitions.outcome_model import OUTCOME_CLASSES, get_feature_columns, predict_transition_probs

    phase_transitions = {}
    if model:
        for phase in ["powerplay", "middle", "death"]:
            phase_num = {"powerplay": 0, "middle": 1, "death": 2}[phase]
            bat_phase = batting_stats[batting_stats["phase"] == phase]
            bowl_phase = bowling_stats[bowling_stats["phase"] == phase]
            context = {
                "phase_num": phase_num,
                "balls_remaining": {"powerplay": 90, "middle": 50, "death": 20}[phase],
                "cumulative_runs": {"powerplay": 40, "middle": 100, "death": 150}[phase],
                "cumulative_wickets": {"powerplay": 1, "middle": 3, "death": 5}[phase],
                "current_run_rate": {"powerplay": 7.5, "middle": 7.8, "death": 9.0}[phase],
                "is_second_innings": 1,
                "batter_sr": bat_phase["strike_rate"].median() if len(bat_phase) > 0 else 130,
                "batter_dot_pct": bat_phase["dot_pct"].median() if len(bat_phase) > 0 else 38,
                "batter_boundary_pct": bat_phase["boundary_pct"].median() if len(bat_phase) > 0 else 15,
                "batter_dismissal_rate": bat_phase["dismissal_rate"].median() if len(bat_phase) > 0 else 0.05,
                "batter_balls_career": 500,
                "batter_phase_sr": bat_phase["strike_rate"].median() if len(bat_phase) > 0 else 130,
                "batter_phase_boundary_pct": bat_phase["boundary_pct"].median() if len(bat_phase) > 0 else 15,
                "batter_phase_dismissal_rate": bat_phase["dismissal_rate"].median() if len(bat_phase) > 0 else 0.05,
                "bowler_economy": bowl_phase["economy"].median() if len(bowl_phase) > 0 else 8.0,
                "bowler_dot_pct": bowl_phase["dot_pct"].median() if len(bowl_phase) > 0 else 35,
                "bowler_boundary_rate": bowl_phase["boundary_concession_rate"].median() if len(bowl_phase) > 0 else 15,
                "bowler_wpm": bowl_phase["wickets_per_match"].median() if len(bowl_phase) > 0 else 1.0,
                "bowler_balls_career": 300,
                "bowler_phase_economy": bowl_phase["economy"].median() if len(bowl_phase) > 0 else 8.0,
                "bowler_phase_dot_pct": bowl_phase["dot_pct"].median() if len(bowl_phase) > 0 else 35,
            }
            probs = predict_transition_probs(model, context)
            phase_transitions[phase] = TransitionProbs(**probs).normalize()

    dp = DPTable()

    def get_ml_tp(b, w):
        overs_bowled = (120 - b) // 6
        if overs_bowled < 6:
            return phase_transitions.get("powerplay", TransitionProbs.from_phase_averages("powerplay"))
        elif overs_bowled < 15:
            return phase_transitions.get("middle", TransitionProbs.from_phase_averages("middle"))
        else:
            return phase_transitions.get("death", TransitionProbs.from_phase_averages("death"))

    dp.solve(get_transition_probs=get_ml_tp if phase_transitions else None)
    print(f"DP table solved. Sanity: target170={dp.lookup(120, 170, 10):.4f}")

    return dp, model, phase_transitions


def parse_events_from_workbook(wb_path):
    """Extract all events from an event book workbook."""
    wb = openpyxl.load_workbook(wb_path, read_only=True)
    filename = Path(wb_path).stem

    # Extract team names from first sheet's headers
    ws0 = wb[wb.sheetnames[0]]
    rows0 = list(ws0.iter_rows(values_only=True))
    headers = rows0[2] if len(rows0) > 2 else []
    team1_name = ""
    team2_name = ""
    for h in headers:
        if h and "_bid1_price" in str(h):
            name = str(h).replace("_bid1_price", "")
            if not team1_name:
                team1_name = name
            elif name != team1_name:
                team2_name = name

    events = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # Parse event info
        event_info = str(rows[0][0]) if rows[0][0] else ""
        score_info = str(rows[0][1]) if rows[0][1] else ""
        parts = sheet_name.split("_")
        event_type = parts[1] if len(parts) > 1 else "?"

        # Parse score
        m = re.search(r"(\d+)/(\d+)\s*\((\d+)\.(\d+)\)", score_info)
        if not m:
            continue
        runs = int(m.group(1))
        wickets = int(m.group(2))
        overs = int(m.group(3))
        balls_in_over = int(m.group(4))
        balls_bowled = overs * 6 + balls_in_over

        # Get market mid-price at t=0 for team1
        best_row = None
        best_ms = float("inf")
        for row in rows[3:]:
            if row[1] is not None and row[2] is not None:
                ms = abs(float(row[1]))
                if ms < best_ms:
                    best_ms = ms
                    best_row = row

        if best_row is None or best_row[2] is None:
            continue

        # Team1 mid
        bid1 = float(best_row[2]) if best_row[2] is not None else None
        ask1 = float(best_row[12]) if best_row[12] is not None else None
        if bid1 is None:
            continue
        team1_mid = (bid1 + (ask1 if ask1 else bid1 + 0.01)) / 2

        events.append({
            "match": filename,
            "sheet": sheet_name,
            "event_num": int(parts[0].replace("E", "")) if parts[0].startswith("E") else 0,
            "event_type": event_type,
            "runs": runs,
            "wickets": wickets,
            "overs": overs,
            "balls_in_over": balls_in_over,
            "balls_bowled": balls_bowled,
            "team1_name": team1_name,
            "team2_name": team2_name,
            "team1_mid": team1_mid,
        })

    wb.close()
    return events


def identify_innings(events):
    """Split events into 1st and 2nd innings and determine target."""
    if not events:
        return [], [], 0

    # Find the innings boundary: score resets
    inn1 = []
    inn2 = []
    prev_runs = 0
    innings = 1
    max_1st_innings_score = 0

    for e in sorted(events, key=lambda x: x["event_num"]):
        if innings == 1 and e["runs"] < prev_runs * 0.5 and prev_runs > 50:
            innings = 2
            max_1st_innings_score = prev_runs

        if innings == 1:
            inn1.append(e)
            max_1st_innings_score = max(max_1st_innings_score, e["runs"])
        else:
            inn2.append(e)

        prev_runs = e["runs"]

    # If no clear boundary found, try another heuristic
    if not inn2 and len(inn1) > 10:
        # Look for a score reset
        for i in range(1, len(inn1)):
            if inn1[i]["runs"] < inn1[i-1]["runs"] - 20:
                inn2 = inn1[i:]
                inn1 = inn1[:i]
                max_1st_innings_score = inn1[-1]["runs"] if inn1 else 0
                break

    target = max_1st_innings_score + 1  # need 1 more to win
    return inn1, inn2, target


# ====================================================================
# TEST 1: Ball Outcome Prediction
# ====================================================================
def test1_ball_prediction(events, phase_transitions):
    """For each event, predict the most likely outcome and check if correct."""
    from src.dp.states import TransitionProbs

    print("\n" + "=" * 70)
    print("TEST 1: BALL OUTCOME PREDICTION")
    print("=" * 70)
    print("For each captured event (4/6/W), what does our model predict?")
    print("The model outputs P(dot), P(single), P(4), P(6), P(wicket)...")
    print("We check: (a) most likely prediction, (b) probability assigned to actual outcome\n")

    correct = 0
    total = 0
    prob_assigned_to_actual = []
    event_type_counts = {"4": 0, "6": 0, "W": 0}
    event_type_correct = {"4": 0, "6": 0, "W": 0}
    event_type_prob = {"4": [], "6": [], "W": []}

    # Map event types to outcome classes
    event_to_outcome = {"4": "four", "6": "six", "W": "wicket"}

    for e in events:
        etype = e["event_type"]
        if etype not in event_to_outcome:
            continue

        actual_outcome = event_to_outcome[etype]
        balls = e["balls_bowled"]

        # Get phase
        over = balls // 6
        if over < 6:
            phase = "powerplay"
        elif over < 15:
            phase = "middle"
        else:
            phase = "death"

        # Get transition probabilities
        if phase in phase_transitions:
            tp = phase_transitions[phase]
        else:
            tp = TransitionProbs.from_phase_averages(phase)

        probs = tp.as_dict()

        # Most likely outcome
        predicted = max(probs, key=probs.get)
        is_correct = (predicted == actual_outcome)

        # Probability assigned to actual outcome
        prob_actual = probs.get(actual_outcome, 0)
        prob_assigned_to_actual.append(prob_actual)

        total += 1
        if is_correct:
            correct += 1

        if etype in event_type_counts:
            event_type_counts[etype] += 1
            if is_correct:
                event_type_correct[etype] += 1
            event_type_prob[etype].append(prob_actual)

    # Results
    accuracy = correct / total if total > 0 else 0
    avg_prob = np.mean(prob_assigned_to_actual) if prob_assigned_to_actual else 0

    print(f"Total events tested: {total}")
    print(f"Correct predictions: {correct}/{total} = {accuracy:.1%}")
    print(f"Avg probability assigned to actual outcome: {avg_prob:.4f}")
    print()

    # NOTE: This accuracy is expected to be LOW because:
    # - We only capture 4s, 6s, and Ws (the dramatic events)
    # - dots and singles are never captured as events
    # - So the model's most likely prediction (usually "single" or "dot") will almost
    #   never match because the event set is 100% boundaries/wickets
    print("IMPORTANT CONTEXT:")
    print("  Captured events are ONLY boundaries (4,6) and wickets (W).")
    print("  dots/singles (the most common outcomes ~65%) are never captured.")
    print("  So 'most likely prediction = single' will never match the captured data.")
    print("  The right metric is: probability the model assigns to actual outcome.\n")

    print(f"{'Event':>6} {'Count':>6} {'Correct':>8} {'Accuracy':>10} {'Avg P(actual)':>14}")
    print("-" * 50)
    for etype in ["4", "6", "W"]:
        cnt = event_type_counts[etype]
        corr = event_type_correct[etype]
        acc = corr / cnt if cnt > 0 else 0
        avg_p = np.mean(event_type_prob[etype]) if event_type_prob[etype] else 0
        print(f"{etype:>6} {cnt:>6} {corr:>8} {acc:>10.1%} {avg_p:>14.4f}")

    # Historical base rates for comparison
    print(f"\nFor reference — IPL historical base rates:")
    print(f"  P(four)   ≈ 11%   — model assigns: {np.mean(event_type_prob.get('4', [0])):.1%}")
    print(f"  P(six)    ≈ 6%    — model assigns: {np.mean(event_type_prob.get('6', [0])):.1%}")
    print(f"  P(wicket) ≈ 5%    — model assigns: {np.mean(event_type_prob.get('W', [0])):.1%}")
    print(f"  P(dot)    ≈ 36%   — (never in captured events)")
    print(f"  P(single) ≈ 28%   — (never in captured events)")

    return {
        "total_events": total,
        "correct": correct,
        "accuracy": accuracy,
        "avg_prob_actual": avg_prob,
        "by_type": {
            etype: {
                "count": event_type_counts[etype],
                "correct": event_type_correct[etype],
                "avg_prob": float(np.mean(event_type_prob[etype])) if event_type_prob[etype] else 0,
            }
            for etype in ["4", "6", "W"]
        },
    }


# ====================================================================
# TEST 2: Win Probability vs Polymarket
# ====================================================================
def test2_win_prob_vs_market(events_by_match, dp):
    """Compare model chase win probability vs Polymarket mid-price."""
    print("\n" + "=" * 70)
    print("TEST 2: WIN PROBABILITY vs POLYMARKET")
    print("=" * 70)
    print("For 2nd innings events, compare model P(chase win) vs market price.\n")

    all_model = []
    all_market = []
    all_details = []

    for match_name, events in events_by_match.items():
        inn1, inn2, target = identify_innings(events)
        if not inn2 or target <= 0:
            print(f"  {match_name}: Could not identify 2nd innings (inn1={len(inn1)}, inn2={len(inn2)})")
            continue

        team1 = events[0]["team1_name"]
        team2 = events[0]["team2_name"]

        print(f"\n  {match_name}")
        print(f"  Teams: {team1} vs {team2}")
        print(f"  Target: {target} (1st inn score: {target-1})")
        print(f"  2nd innings events: {len(inn2)}")

        # Determine which team is chasing
        # In 2nd innings, if team1_mid is HIGH when chase is going well (runs increasing,
        # wickets low), then team1 is the chasing team
        # If team1_mid is LOW when chase is going well, team2 is chasing

        # Identify chasing team using event reactions:
        # On a BOUNDARY: chaser's prob goes UP, fielder's goes DOWN
        # On a WICKET:   chaser's prob goes DOWN, fielder's goes UP
        # Count: does team1_mid go UP on boundaries (→ team1 chasing) or DOWN?
        if len(inn2) >= 3:
            boundary_deltas = []
            wicket_deltas = []
            for i in range(1, len(inn2)):
                delta = inn2[i]["team1_mid"] - inn2[i-1]["team1_mid"]
                if inn2[i]["event_type"] in ("4", "6"):
                    boundary_deltas.append(delta)
                elif inn2[i]["event_type"] == "W":
                    wicket_deltas.append(delta)

            # On boundaries, does team1_mid go UP (team1 is chasing) or DOWN?
            boundary_up = sum(1 for d in boundary_deltas if d > 0.005)
            boundary_down = sum(1 for d in boundary_deltas if d < -0.005)
            # On wickets, does team1_mid go DOWN (team1 is chasing) or UP?
            wicket_down = sum(1 for d in wicket_deltas if d < -0.005)
            wicket_up = sum(1 for d in wicket_deltas if d > 0.005)

            # Score: +1 for evidence team1 is chasing, -1 for evidence team1 is fielding
            score = (boundary_up - boundary_down) + (wicket_down - wicket_up)

            if score > 0:
                chasing_team = team1
                chase_prob_from_market = [e["team1_mid"] for e in inn2]
            else:
                chasing_team = team2
                chase_prob_from_market = [1 - e["team1_mid"] for e in inn2]

            print(f"  Chasing team: {chasing_team} (boundary_up={boundary_up}, boundary_down={boundary_down}, "
                  f"wicket_down={wicket_down}, wicket_up={wicket_up}, score={score})")
        else:
            chasing_team = team1
            chase_prob_from_market = [e["team1_mid"] for e in inn2]

        print(f"  {'Event':>6} {'Type':>5} {'Score':>10} {'Model':>8} {'Market':>8} {'Diff':>8}")
        print(f"  {'-'*50}")

        for i, e in enumerate(inn2):
            runs_scored = e["runs"]
            wickets_fallen = e["wickets"]
            balls_bowled = e["balls_bowled"]

            runs_needed = target - runs_scored
            balls_remaining = 120 - balls_bowled
            wickets_in_hand = 10 - wickets_fallen

            if runs_needed <= 0:
                model_prob = 1.0
            elif wickets_in_hand <= 0 or balls_remaining <= 0:
                model_prob = 0.0
            else:
                model_prob = dp.lookup(balls_remaining, min(runs_needed, dp.MAX_RUNS), wickets_in_hand)

            market_prob = chase_prob_from_market[i]

            diff = model_prob - market_prob
            all_model.append(model_prob)
            all_market.append(market_prob)
            all_details.append({
                "match": match_name,
                "event": e["sheet"],
                "type": e["event_type"],
                "score": f"{runs_scored}/{wickets_fallen}",
                "overs": f"{e['overs']}.{e['balls_in_over']}",
                "runs_needed": runs_needed,
                "balls_rem": balls_remaining,
                "wkts_hand": wickets_in_hand,
                "model": model_prob,
                "market": market_prob,
                "diff": diff,
            })

            print(f"  {e['sheet'].split('_')[0]:>6} {e['event_type']:>5} {runs_scored:>4}/{wickets_fallen} ({e['overs']}.{e['balls_in_over']}) "
                  f"{model_prob:>8.4f} {market_prob:>8.4f} {diff:>+8.4f}")

    # Overall metrics
    all_model = np.array(all_model)
    all_market = np.array(all_market)

    print(f"\n{'='*70}")
    print(f"OVERALL RESULTS ({len(all_model)} 2nd-innings events across {len(events_by_match)} matches)")
    print(f"{'='*70}")

    if len(all_model) < 2:
        print("Not enough data points for metrics")
        return {}

    rmse = np.sqrt(np.mean((all_model - all_market) ** 2))
    mae = np.mean(np.abs(all_model - all_market))
    corr = np.corrcoef(all_model, all_market)[0, 1]
    bias = np.mean(all_model - all_market)

    print(f"  RMSE:         {rmse:.4f}")
    print(f"  MAE:          {mae:.4f}")
    print(f"  Correlation:  {corr:.4f}")
    print(f"  Bias:         {bias:+.4f} ({'model overestimates' if bias > 0 else 'model underestimates'})")

    # Breakdown by event type
    details_df = pd.DataFrame(all_details)
    print(f"\n  By event type:")
    for etype, group in details_df.groupby("type"):
        m = np.array(group["model"])
        mk = np.array(group["market"])
        d = m - mk
        r = np.sqrt(np.mean(d**2))
        print(f"    {etype}: n={len(group):>3}, RMSE={r:.4f}, bias={d.mean():+.4f}, MAE={np.abs(d).mean():.4f}")

    # Breakdown by match phase
    print(f"\n  By chase phase:")
    for phase_name, ball_range in [("Early (0-6 ov)", (0, 36)), ("Middle (6-15)", (36, 90)), ("Death (15-20)", (90, 120))]:
        mask = (details_df["balls_rem"].values < (120 - ball_range[0])) & (details_df["balls_rem"].values >= (120 - ball_range[1]))
        if mask.sum() > 0:
            m = details_df.loc[mask, "model"].values
            mk = details_df.loc[mask, "market"].values
            d = m - mk
            r = np.sqrt(np.mean(d**2))
            print(f"    {phase_name}: n={mask.sum():>3}, RMSE={r:.4f}, bias={d.mean():+.4f}")

    return {
        "n_events": len(all_model),
        "rmse": float(rmse),
        "mae": float(mae),
        "correlation": float(corr),
        "bias": float(bias),
    }


# ====================================================================
# MAIN
# ====================================================================
def main():
    print("=" * 70)
    print("HONEST TESTING — Real Data Validation")
    print("=" * 70)

    # Load model
    dp, model, phase_transitions = load_dp_and_model()

    # Find the 3 largest event book files
    event_files = sorted(
        [f for f in CAPTURES.glob("*event_book_cricipl*.xlsx") if "visual" not in f.name],
        key=lambda f: f.stat().st_size,
        reverse=True,
    )[:6]  # take top 6 by size (includes duplicates, we'll handle)

    print(f"\nUsing {len(event_files)} captured match files:")
    for f in event_files:
        print(f"  {f.name}")

    # Parse all events
    all_events = []
    events_by_match = {}
    for f in event_files:
        events = parse_events_from_workbook(f)
        match_key = f.stem
        events_by_match[match_key] = events
        all_events.extend(events)
        print(f"  {f.name}: {len(events)} events")

    print(f"\nTotal events: {len(all_events)}")

    # Test 1: Ball outcome prediction
    test1_results = test1_ball_prediction(all_events, phase_transitions)

    # Test 2: Win probability vs market
    test2_results = test2_win_prob_vs_market(events_by_match, dp)

    # Summary
    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"\nTest 1 — Ball Outcome Prediction:")
    print(f"  Events: {test1_results['total_events']}")
    print(f"  P(actual) for boundaries: {test1_results['by_type']['4']['avg_prob']:.1%}")
    print(f"  P(actual) for sixes:      {test1_results['by_type']['6']['avg_prob']:.1%}")
    print(f"  P(actual) for wickets:    {test1_results['by_type']['W']['avg_prob']:.1%}")
    print(f"  (Base rates: 4→11%, 6→6%, W→5%)")

    print(f"\nTest 2 — Win Prob vs Polymarket:")
    if test2_results:
        print(f"  RMSE:        {test2_results['rmse']:.4f}")
        print(f"  MAE:         {test2_results['mae']:.4f}")
        print(f"  Correlation: {test2_results['correlation']:.4f}")
        print(f"  Bias:        {test2_results['bias']:+.4f}")

    # Save results
    results = {"test1": test1_results, "test2": test2_results}
    (ROOT / "HONEST_SCORES.json").write_text(json.dumps(results, indent=2, default=str))
    print(f"\nResults saved to HONEST_SCORES.json")


if __name__ == "__main__":
    main()
