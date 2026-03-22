#!/usr/bin/env python3
"""Polymarket Live Event Capture — records EVERY event at nanosecond precision.

Captures from ALL available data sources simultaneously:
  1. Market WS  — every book snapshot, price change, trade tick (real-time)
  2. REST trades — every trade with wallet attribution (2s poll)

Every event logged with:
  - Nanosecond timestamp (time.time_ns)
  - IST time string (Asia/Kolkata)
  - Full classification: event type, side, outcome, price, size, tokens, USD
  - Wallet address (when available from REST)
  - Order inference: FILL / CANCEL / ORDER_ADDED / ORDER_REMOVED
  - Correlation: WS trade ticks matched to REST wallet data via tx hash

Output: JSONL file + live colored terminal output

Usage:
  python capture.py <slug> [--duration 600]
  python capture.py crint-gbr-ind-2026-03-05 --duration 300
"""

import argparse
import asyncio
import json
import os
import sys
import time
from collections import defaultdict, deque
from datetime import datetime, timezone, timedelta
from pathlib import Path

import httpx
import websockets

# ── Constants ──────────────────────────────────────────────────────────

GAMMA_API = "https://gamma-api.polymarket.com"
DATA_API = "https://data-api.polymarket.com"
WS_URL = "wss://ws-subscriptions-clob.polymarket.com/ws/market"

IST = timezone(timedelta(hours=5, minutes=30))

# ANSI colors
C_RESET = "\033[0m"
C_RED = "\033[91m"
C_GREEN = "\033[92m"
C_YELLOW = "\033[93m"
C_BLUE = "\033[94m"
C_MAGENTA = "\033[95m"
C_CYAN = "\033[96m"
C_DIM = "\033[90m"
C_BOLD = "\033[1m"
C_BG_RED = "\033[41m"
C_BG_GREEN = "\033[42m"


# ── Helpers ────────────────────────────────────────────────────────────

def now_ns() -> int:
    return time.time_ns()

def now_ist() -> str:
    """IST timestamp with nanosecond fraction."""
    t = time.time()
    ns = time.time_ns() % 1_000_000_000
    dt = datetime.fromtimestamp(t, tz=IST)
    return dt.strftime("%Y-%m-%d %H:%M:%S") + f".{ns:09d} IST"

def ist_short() -> str:
    t = time.time()
    ns = time.time_ns() % 1_000_000
    dt = datetime.fromtimestamp(t, tz=IST)
    return dt.strftime("%H:%M:%S") + f".{ns:06d}"


# ── Market Fetch ───────────────────────────────────────────────────────

def fetch_market(slug: str) -> dict:
    resp = httpx.get(f"{GAMMA_API}/markets", params={"slug": slug}, timeout=15)
    resp.raise_for_status()
    markets = resp.json()
    if not markets:
        print(f"Error: No market found for slug '{slug}'")
        sys.exit(1)
    return markets[0]

def parse_tokens(market: dict) -> tuple[list[str], list[str]]:
    tokens = market.get("clobTokenIds", "")
    outcomes = market.get("outcomes", "[]")
    if isinstance(tokens, str):
        try: tokens = json.loads(tokens)
        except: tokens = tokens.split(",") if tokens else []
    if isinstance(outcomes, str):
        try: outcomes = json.loads(outcomes)
        except: outcomes = []
    return tokens, outcomes


# ── Event Logger ───────────────────────────────────────────────────────

class EventLogger:
    """Writes structured events to JSONL + prints colored terminal output."""

    def __init__(self, output_path: str):
        self.path = output_path
        self._file = open(output_path, "a")
        self.count = 0
        self.counts_by_type = defaultdict(int)

    def log(self, event_type: str, data: dict, color: str = "", prefix: str = ""):
        ts_ns = now_ns()
        ts = time.time()
        ist = now_ist()

        record = {
            "seq": self.count,
            "ts_ns": ts_ns,
            "ts": ts,
            "ist": ist,
            "type": event_type,
            **data,
        }
        self._file.write(json.dumps(record) + "\n")
        self._file.flush()
        self.count += 1
        self.counts_by_type[event_type] += 1

        # Terminal output
        short_t = ist_short()
        summary = self._format_summary(event_type, data)
        pfx = prefix or event_type.upper()
        c = color or C_DIM
        print(f"{C_DIM}{short_t}{C_RESET} {c}{pfx:>18s}{C_RESET} {summary}")

    def _format_summary(self, etype: str, d: dict) -> str:
        outcome = d.get("outcome", d.get("token_name", ""))
        side = d.get("side", "")
        price = d.get("price")
        size = d.get("size")
        notional = d.get("notional")
        wallet = d.get("wallet", d.get("taker", d.get("maker", "")))

        parts = []
        if side:
            c = C_GREEN if side == "BUY" else C_RED
            parts.append(f"{c}{side}{C_RESET}")
        if outcome:
            parts.append(f"{C_CYAN}{outcome}{C_RESET}")
        if price is not None:
            parts.append(f"@{C_YELLOW}{price}{C_RESET}")
        if size is not None:
            parts.append(f"{size} tokens")
        if notional is not None:
            parts.append(f"${notional:.2f}")
        if wallet:
            w = wallet[:8] + ".." if len(wallet) > 10 else wallet
            parts.append(f"{C_MAGENTA}{w}{C_RESET}")

        # Extra fields
        delta = d.get("delta")
        if delta is not None:
            c = C_GREEN if delta > 0 else C_RED
            parts.append(f"delta={c}{delta:+.0f}{C_RESET}")
        old_size = d.get("old_size")
        new_size = d.get("new_size")
        if old_size is not None and new_size is not None:
            parts.append(f"({old_size:.0f}→{new_size:.0f})")
        tx = d.get("tx_hash", d.get("transaction_hash", ""))
        if tx:
            parts.append(f"tx={C_DIM}{tx[:12]}{C_RESET}")
        fee = d.get("fee_rate_bps")
        if fee and fee != "0":
            parts.append(f"fee={fee}bps")

        return " ".join(parts) if parts else json.dumps(d)[:120]

    def close(self):
        self._file.close()

    def print_summary(self):
        print(f"\n{C_BOLD}{'='*70}{C_RESET}")
        print(f"{C_BOLD}  CAPTURE SUMMARY{C_RESET}")
        print(f"{'='*70}")
        print(f"  Total events: {C_BOLD}{self.count}{C_RESET}")
        print(f"  Output file:  {self.path}")
        for etype, cnt in sorted(self.counts_by_type.items(), key=lambda x: -x[1]):
            print(f"    {etype:>25s}: {cnt}")
        print(f"{'='*70}")


# ── Capture State ──────────────────────────────────────────────────────

class CaptureState:
    def __init__(self, token_ids: list[str], outcome_names: list[str], logger: EventLogger):
        self.token_ids = token_ids
        self.outcome_names = outcome_names
        self.token_to_outcome = dict(zip(token_ids, outcome_names))
        self.logger = logger

        # L2 book per token: {token_id: {"bids": {price: size}, "asks": {price: size}}}
        self.books: dict[str, dict[str, dict[float, float]]] = {
            tid: {"bids": {}, "asks": {}} for tid in token_ids
        }

        # Trade ticks indexed by (token_id, price_rounded) for fast lookup
        self._trade_ticks: deque[dict] = deque(maxlen=2000)
        self._classify_window = 2.0  # seconds to wait before classifying a level decrease

        # Pending level decreases waiting for classification
        self._pending_decreases: deque[dict] = deque(maxlen=5000)

        # REST trade dedup
        self._seen_tx: set[str] = set()

        # Stats
        self.ws_msg_count = 0
        self.trade_count = 0
        self.pure_cancel_count = 0
        self.pure_fill_count = 0
        self.snipe_mix_count = 0
        self.order_add_count = 0

    # ── WS message processing ──────────────────────────────────────

    def process_ws(self, raw_text: str):
        """Process a raw WS message — classify and log every event."""
        self.ws_msg_count += 1

        try:
            data = json.loads(raw_text)
        except json.JSONDecodeError:
            self.logger.log("ws_unparseable", {"raw": raw_text[:200]}, C_RED, "WS_ERROR")
            return

        events = data if isinstance(data, list) else [data]

        for event in events:
            if not isinstance(event, dict):
                continue

            # Market-level price_changes wrapper
            if "price_changes" in event:
                for change in event["price_changes"]:
                    self._process_price_change(change)
                continue

            asset_id = event.get("asset_id", "")
            event_type = event.get("event_type") or event.get("type", "")

            if event_type == "book":
                self._process_book(event, asset_id)

            elif event_type == "price_change":
                changes = event.get("changes", [])
                if changes:
                    for c in changes:
                        self._process_price_change(c, asset_id)
                else:
                    # Fallback: bids/asks in event body
                    for b in event.get("bids", []):
                        self._process_price_change(
                            {"price": b.get("price",0), "size": b.get("size",0), "side": "BUY"},
                            asset_id
                        )
                    for a in event.get("asks", []):
                        self._process_price_change(
                            {"price": a.get("price",0), "size": a.get("size",0), "side": "SELL"},
                            asset_id
                        )

            elif event_type == "last_trade_price":
                self._process_trade_tick(event, asset_id)

            elif event_type == "tick_size_change":
                outcome = self.token_to_outcome.get(asset_id, "")
                self.logger.log("tick_size_change", {
                    "token_id": asset_id, "outcome": outcome,
                    "old_tick_size": event.get("old_tick_size"),
                    "new_tick_size": event.get("new_tick_size"),
                }, C_YELLOW, "TICK_CHANGE")

    def _process_book(self, event: dict, asset_id: str):
        """Full book snapshot — log it, rebuild internal book."""
        outcome = self.token_to_outcome.get(asset_id, "")
        bids = event.get("bids", [])
        asks = event.get("asks", [])

        # Rebuild internal book
        if asset_id in self.books:
            self.books[asset_id]["bids"] = {}
            self.books[asset_id]["asks"] = {}
            for b in bids:
                p, s = float(b.get("price",0)), float(b.get("size",0))
                if s > 0: self.books[asset_id]["bids"][p] = s
            for a in asks:
                p, s = float(a.get("price",0)), float(a.get("size",0))
                if s > 0: self.books[asset_id]["asks"][p] = s

        bid_depth = sum(float(b.get("price",0))*float(b.get("size",0)) for b in bids)
        ask_depth = sum(float(a.get("price",0))*float(a.get("size",0)) for a in asks)
        best_bid = float(bids[0]["price"]) if bids else 0
        best_ask = float(asks[0]["price"]) if asks else 0

        self.logger.log("book_snapshot", {
            "token_id": asset_id, "outcome": outcome,
            "bid_levels": len(bids), "ask_levels": len(asks),
            "best_bid": best_bid, "best_ask": best_ask,
            "bid_depth_usd": round(bid_depth, 2),
            "ask_depth_usd": round(ask_depth, 2),
            "spread": round(best_ask - best_bid, 4) if best_bid and best_ask else None,
        }, C_BLUE, "BOOK_SNAPSHOT")

    def _process_price_change(self, change: dict, fallback_asset: str = ""):
        """Single price level change — classify as order add, remove, or cancel."""
        asset_id = change.get("asset_id", fallback_asset)
        price = float(change.get("price", 0))
        new_size = float(change.get("size", 0))
        side = change.get("side", "")
        book_hash = change.get("hash", "")
        outcome = self.token_to_outcome.get(asset_id, "")

        if asset_id not in self.books:
            return

        side_key = "bids" if side == "BUY" else "asks"
        old_size = self.books[asset_id][side_key].get(price, 0.0)
        delta = new_size - old_size

        # Skip noise
        if abs(delta) < 0.001:
            return

        # Update internal book
        if new_size <= 0:
            self.books[asset_id][side_key].pop(price, None)
        else:
            self.books[asset_id][side_key][price] = new_size

        notional = abs(delta) * price

        if delta > 0:
            self.order_add_count += 1
            self.logger.log("level_increase", {
                "token_id": asset_id, "outcome": outcome,
                "side": side, "price": price,
                "old_size": round(old_size, 2), "new_size": round(new_size, 2),
                "delta": round(delta, 2),
                "notional": round(notional, 2),
                "hash": book_hash,
                "source": "ws_price_change",
                "meaning": "resting_order_added",
                "wallet": "unknown — Market WS does not expose wallet on resting orders",
            }, C_GREEN, "LEVEL_UP")
        else:
            # Queue this decrease — classify after 2s when we know if trades followed
            removed = abs(delta)
            self._pending_decreases.append({
                "ts": time.time(),
                "token_id": asset_id, "outcome": outcome,
                "side": side, "price": price,
                "old_size": old_size, "new_size": new_size,
                "removed": removed,
                "notional": notional,
                "hash": book_hash,
            })

        # Classify any pending decreases whose 2s window has elapsed
        self._classify_pending()

    def _process_trade_tick(self, event: dict, asset_id: str):
        """last_trade_price — a definitive fill happened."""
        price = float(event.get("price", 0))
        size = float(event.get("size", 0))
        side = event.get("side", "")
        tx_hash = event.get("transaction_hash", "")
        fee_bps = event.get("fee_rate_bps", "0")
        outcome = self.token_to_outcome.get(asset_id, "")
        notional = price * size

        self.trade_count += 1

        # Store for deferred classification of level decreases
        self._trade_ticks.append({
            "ts": time.time(), "token_id": asset_id,
            "price": price, "size": size, "side": side,
        })

        if side == "BUY":
            taker_action = "TAKER_BUY"
            maker_action = "MAKER_SELL"
        else:
            taker_action = "TAKER_SELL"
            maker_action = "MAKER_BUY"

        self.logger.log("trade", {
            "token_id": asset_id, "outcome": outcome,
            "side": side, "price": price, "size": round(size, 6),
            "notional": round(notional, 2),
            "tokens": round(size, 6),
            "usd": round(notional, 2),
            "taker_side": side,
            "taker_action": taker_action,
            "maker_action": maker_action,
            "fee_rate_bps": fee_bps,
            "transaction_hash": tx_hash,
            "source": "ws_last_trade_price",
            "is_definitive": True,
            "wallet": "unknown — Market WS does not expose wallet on trades. See rest_trade events for wallet.",
        }, f"{C_BOLD}{C_BG_GREEN}" if side == "BUY" else f"{C_BOLD}{C_BG_RED}", "TRADE")

    def _classify_pending(self):
        """Classify level decreases after their 2s window has elapsed.

        For each pending decrease at price P:
          - Sum all trade ticks at price P within ±2s
          - traded_vol == 0       → PURE_CANCEL (no trade at this level)
          - traded_vol >= removed → PURE_FILL (trade consumed the level)
          - 0 < traded < removed  → SNIPE_MIX (partial fill + cancel = someone pulled)
        """
        now = time.time()
        while self._pending_decreases:
            d = self._pending_decreases[0]
            age = now - d["ts"]
            if age < self._classify_window:
                break  # not old enough yet
            self._pending_decreases.popleft()

            # Sum trades at this token+price within the window
            traded_vol = 0.0
            trade_count = 0
            for t in self._trade_ticks:
                if (t["token_id"] == d["token_id"]
                        and abs(t["price"] - d["price"]) < 0.0001
                        and abs(t["ts"] - d["ts"]) <= self._classify_window):
                    traded_vol += t["size"]
                    trade_count += 1

            removed = d["removed"]
            base = {
                "token_id": d["token_id"], "outcome": d["outcome"],
                "side": d["side"], "price": d["price"],
                "old_size": round(d["old_size"], 2),
                "new_size": round(d["new_size"], 2),
                "delta": round(-removed, 2),
                "removed": round(removed, 2),
                "notional": round(d["notional"], 2),
                "hash": d["hash"],
                "source": "ws_price_change",
                "traded_at_level_2s": round(traded_vol, 2),
                "trades_at_level_2s": trade_count,
                "wallet": "unknown — cannot see who cancelled/placed resting orders",
            }

            if traded_vol < 0.001:
                # NO trades at this price → pure cancel
                self.pure_cancel_count += 1
                base["tag"] = "PURE_CANCEL"
                base["reason"] = "level decreased with zero trades at this price in 2s window — order was cancelled (by user, GTD expiry, or system)"
                self.logger.log("pure_cancel", base, C_RED, "CANCEL")
            elif traded_vol >= removed * 0.95:
                # Trades >= decrease → pure fill
                self.pure_fill_count += 1
                base["tag"] = "PURE_FILL"
                base["reason"] = f"level decreased by {removed:.0f} and {traded_vol:.0f} traded at this price — fully consumed by fill"
                self.logger.log("pure_fill", base, C_GREEN, "FILL")
            else:
                # Partial trade + cancel = snipe/mix
                cancelled = removed - traded_vol
                self.snipe_mix_count += 1
                base["tag"] = "SNIPE_MIX"
                base["filled_portion"] = round(traded_vol, 2)
                base["cancelled_portion"] = round(cancelled, 2)
                base["cancel_pct"] = round(cancelled / removed * 100, 1)
                base["reason"] = (
                    f"level decreased by {removed:.0f} but only {traded_vol:.0f} traded — "
                    f"{cancelled:.0f} tokens ({cancelled/removed*100:.0f}%) were pulled before getting filled. "
                    f"Likely: maker saw aggressor coming and sniped out, or partial cancel+fill."
                )
                self.logger.log("snipe_mix", base, C_YELLOW, "SNIPE")

    def flush_pending(self):
        """Force-classify all remaining pending decreases (called at capture end)."""
        self._classify_window = 0  # classify everything
        self._classify_pending()

    # ── REST trade processing ──────────────────────────────────────

    def process_rest_trades(self, trades: list[dict]):
        """Process trades from REST API — wallet attribution."""
        for t in trades:
            tx_hash = t.get("transactionHash", "")
            if not tx_hash or tx_hash in self._seen_tx:
                continue
            self._seen_tx.add(tx_hash)

            size = float(t.get("size", 0))
            price = float(t.get("price", 0))
            notional = size * price
            side = t.get("side", "")
            wallet = (t.get("proxyWallet", "") or "").lower()
            outcome = t.get("outcome", t.get("asset", ""))
            ts_raw = t.get("timestamp", "")
            fee = t.get("feeRateBps", "0")
            maker_addr = (t.get("maker_address", "") or "").lower()

            self.logger.log("rest_trade", {
                "outcome": outcome,
                "side": side,
                "price": price,
                "size": round(size, 6),
                "tokens": round(size, 6),
                "usd": round(notional, 2),
                "notional": round(notional, 2),
                "wallet": wallet,
                "maker_address": maker_addr,
                "tx_hash": tx_hash,
                "fee_rate_bps": fee,
                "timestamp_api": ts_raw,
                "source": "rest_data_api",
                "is_definitive": True,
                # This is a REAL trade with wallet attribution (taker's proxy wallet)
                # We know: wallet, side, price, size, outcome, tx hash
                # We DON'T know: order type used (GTC/GTD/FOK/FAK)
            }, C_MAGENTA, "REST_TRADE")


# ── Async Tasks ────────────────────────────────────────────────────────

async def ws_capture(state: CaptureState, stop_event: asyncio.Event):
    """Connect to Market WS and capture every message."""
    while not stop_event.is_set():
        try:
            async with websockets.connect(WS_URL, ping_interval=None) as ws:
                sub = json.dumps({
                    "assets_ids": state.token_ids,
                    "type": "market",
                })
                await ws.send(sub)
                state.logger.log("ws_connected", {
                    "tokens": len(state.token_ids),
                    "outcomes": state.outcome_names,
                }, C_BLUE, "WS_CONNECTED")

                async def keepalive():
                    while not stop_event.is_set():
                        await asyncio.sleep(10)
                        try: await ws.send("PING")
                        except: return

                ping_task = asyncio.create_task(keepalive())
                try:
                    async for raw_msg in ws:
                        if stop_event.is_set():
                            break
                        if raw_msg == "PONG":
                            continue
                        state.process_ws(raw_msg)
                finally:
                    ping_task.cancel()

        except asyncio.CancelledError:
            break
        except Exception as e:
            state.logger.log("ws_error", {"error": str(e)}, C_RED, "WS_ERROR")
            try:
                await asyncio.wait_for(stop_event.wait(), timeout=2)
                break
            except asyncio.TimeoutError:
                pass


async def rest_poller(state: CaptureState, condition_id: str, stop_event: asyncio.Event):
    """Poll REST API for trades with wallet attribution."""
    while not stop_event.is_set():
        try:
            async with httpx.AsyncClient(timeout=15) as client:
                resp = await client.get(
                    f"{DATA_API}/trades",
                    params={"market": condition_id, "limit": 100},
                )
                if resp.status_code == 200:
                    state.process_rest_trades(resp.json())
        except asyncio.CancelledError:
            break
        except Exception as e:
            state.logger.log("rest_error", {"error": str(e)}, C_RED, "REST_ERROR")

        try:
            await asyncio.wait_for(stop_event.wait(), timeout=2)
            break
        except asyncio.TimeoutError:
            pass


async def status_printer(state: CaptureState, stop_event: asyncio.Event, duration: int):
    """Print periodic status updates."""
    start = time.time()
    while not stop_event.is_set():
        try:
            await asyncio.wait_for(stop_event.wait(), timeout=15)
            break
        except asyncio.TimeoutError:
            elapsed = int(time.time() - start)
            remaining = max(0, duration - elapsed)
            print(
                f"\n{C_DIM}── {ist_short()} | "
                f"Events: {state.logger.count} | "
                f"WS: {state.ws_msg_count} | "
                f"Trades: {state.trade_count} | "
                f"Cancels: {C_RED}{state.pure_cancel_count}{C_DIM} | "
                f"Fills: {C_GREEN}{state.pure_fill_count}{C_DIM} | "
                f"Snipes: {C_YELLOW}{state.snipe_mix_count}{C_DIM} | "
                f"Adds: {state.order_add_count} | "
                f"Remaining: {remaining//60}m{remaining%60}s"
                f" ──{C_RESET}\n"
            )


async def timer(stop_event: asyncio.Event, duration: int):
    """Stop capture after duration seconds."""
    await asyncio.sleep(duration)
    stop_event.set()


# ── Main ───────────────────────────────────────────────────────────────

def parse_ist_time(time_str: str) -> float:
    """Parse HH:MM IST time string into a unix timestamp (today or tomorrow)."""
    from datetime import datetime, timezone, timedelta
    ist = timezone(timedelta(hours=5, minutes=30))
    now = datetime.now(ist)
    parts = time_str.strip().split(":")
    target = now.replace(hour=int(parts[0]), minute=int(parts[1]), second=0, microsecond=0)
    if target < now:
        target += timedelta(days=1)
    return target.timestamp()


async def run(slug: str, duration: int, start_at: str = None):
    print(f"\n{C_BOLD}Polymarket Live Event Capture{C_RESET}")

    if start_at:
        target_ts = parse_ist_time(start_at)
        wait_secs = target_ts - time.time()
        if wait_secs > 0:
            target_ist = datetime.fromtimestamp(target_ts, tz=IST).strftime("%Y-%m-%d %H:%M:%S IST")
            print(f"Scheduled start: {C_YELLOW}{target_ist}{C_RESET}")
            print(f"Waiting {C_BOLD}{int(wait_secs//3600)}h {int((wait_secs%3600)//60)}m{C_RESET}...")
            print(f"(leave this running, capture begins automatically)\n")
            await asyncio.sleep(wait_secs)
            print(f"{C_GREEN}Start time reached — beginning capture{C_RESET}\n")

    print(f"Fetching market: {slug}")

    market = fetch_market(slug)
    condition_id = market.get("conditionId", "")
    question = market.get("question", slug)
    token_ids, outcome_names = parse_tokens(market)

    if not token_ids:
        print("Error: no token IDs")
        sys.exit(1)

    print(f"Market:    {question}")
    print(f"Outcomes:  {', '.join(outcome_names)}")
    print(f"Condition: {condition_id}")
    print(f"Tokens:    {len(token_ids)}")
    print(f"Duration:  {duration}s ({duration//60}m)")

    # Output file
    ts = time.strftime("%Y%m%d_%H%M%S")
    safe_slug = slug.replace("/", "_")[:60]
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(project_root, "data", safe_slug)
    os.makedirs(data_dir, exist_ok=True)
    outfile = os.path.join(data_dir, f"{ts}_{safe_slug}.jsonl")

    logger = EventLogger(outfile)
    logger.log("capture_start", {
        "slug": slug,
        "question": question,
        "condition_id": condition_id,
        "token_ids": token_ids,
        "outcome_names": outcome_names,
        "duration_seconds": duration,
        "ist_start": now_ist(),
    }, C_BOLD, "CAPTURE_START")

    state = CaptureState(token_ids, outcome_names, logger)

    print(f"\nOutput: {outfile}")
    print(f"Started at {now_ist()}")
    print(f"{'='*70}\n")

    stop_event = asyncio.Event()

    try:
        await asyncio.gather(
            ws_capture(state, stop_event),
            rest_poller(state, condition_id, stop_event),
            status_printer(state, stop_event, duration),
            timer(stop_event, duration),
        )
    except KeyboardInterrupt:
        stop_event.set()

    # Flush any remaining pending level decreases
    state.flush_pending()

    logger.log("capture_stop", {
        "total_events": logger.count,
        "ws_messages": state.ws_msg_count,
        "trades": state.trade_count,
        "pure_cancels": state.pure_cancel_count,
        "pure_fills": state.pure_fill_count,
        "snipe_mix": state.snipe_mix_count,
        "order_adds": state.order_add_count,
        "ist_end": now_ist(),
    }, C_BOLD, "CAPTURE_STOP")

    logger.print_summary()
    logger.close()


def export_to_excel(jsonl_path: str):
    """Convert JSONL capture to a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
    from openpyxl.utils import get_column_letter

    events = []
    with open(jsonl_path) as f:
        for line in f:
            events.append(json.loads(line))

    xlsx_path = jsonl_path.replace(".jsonl", ".xlsx")
    wb = Workbook()

    # ── Styling ────────────────────────────────────────────────
    hdr_font = Font(bold=True, size=10)
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font_w = Font(bold=True, size=10, color="FFFFFF")
    green_font = Font(color="22C55E")
    red_font = Font(color="EF4444")
    dim_font = Font(color="9CA3AF", size=9)
    num_fmt_price = '0.0000'
    num_fmt_size = '#,##0.00'
    num_fmt_usd = '$#,##0.00'
    thin_border = Border(
        bottom=XlSide(style='thin', color='374151'),
    )

    def style_header(ws, cols):
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.font = hdr_font_w
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')

    def auto_width(ws, cols):
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 4, 12)

    # ── Sheet 1: All Events ────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Events"
    cols = ["Seq", "IST Timestamp", "Type", "Outcome", "Side", "Price", "Size (tokens)", "USD", "Delta", "Old Size", "New Size", "Wallet", "Tx Hash", "Source", "Confidence/Meaning"]
    style_header(ws_all, cols)
    auto_width(ws_all, cols)

    for e in events:
        if e["type"] in ("capture_start", "capture_stop", "ws_connected"):
            continue
        row = [
            e.get("seq", ""),
            e.get("ist", ""),
            e.get("type", ""),
            e.get("outcome", ""),
            e.get("side", ""),
            e.get("price"),
            e.get("size") or e.get("tokens"),
            e.get("usd") or e.get("notional"),
            e.get("delta"),
            e.get("old_size"),
            e.get("new_size"),
            e.get("wallet", ""),
            e.get("tx_hash") or e.get("transaction_hash", ""),
            e.get("source", ""),
            e.get("meaning") or e.get("confidence", ""),
        ]
        r = ws_all.append(row)

    # ── Sheet 2: Trades Only (WS + REST) ───────────────────────
    ws_trades = wb.create_sheet("Trades")
    trades_cols = ["IST Timestamp", "Source", "Outcome", "Side", "Price", "Size", "USD", "Wallet", "Tx Hash", "Fee BPS", "Taker Action"]
    style_header(ws_trades, trades_cols)
    auto_width(ws_trades, trades_cols)

    for e in events:
        if e["type"] not in ("trade", "rest_trade"):
            continue
        ws_trades.append([
            e.get("ist", ""),
            e.get("source", e["type"]),
            e.get("outcome", ""),
            e.get("side", ""),
            e.get("price"),
            e.get("size") or e.get("tokens"),
            e.get("usd") or e.get("notional"),
            e.get("wallet", ""),
            e.get("tx_hash") or e.get("transaction_hash", ""),
            e.get("fee_rate_bps", ""),
            e.get("taker_action", ""),
        ])

    # ── Sheet 3: Level Changes (book mutations) ────────────────
    ws_levels = wb.create_sheet("Level Changes")
    level_cols = ["IST Timestamp", "Type", "Outcome", "Side", "Price", "Old Size", "New Size", "Delta", "Notional", "Meaning"]
    style_header(ws_levels, level_cols)
    auto_width(ws_levels, level_cols)

    for e in events:
        if e["type"] not in ("level_increase", "pure_cancel", "pure_fill", "snipe_mix"):
            continue
        ws_levels.append([
            e.get("ist", ""),
            e.get("tag", e.get("type", "")),
            e.get("outcome", ""),
            e.get("side", ""),
            e.get("price"),
            e.get("old_size"),
            e.get("new_size"),
            e.get("delta") or e.get("removed"),
            e.get("notional"),
            e.get("traded_at_level_2s", ""),
            e.get("cancelled_portion", ""),
            e.get("cancel_pct", ""),
            e.get("reason", e.get("meaning", "")),
        ])
    # Fix header for level changes
    level_cols = ["IST Timestamp", "Tag", "Outcome", "Side", "Price", "Old Size", "New Size", "Delta", "Notional", "Traded in 2s", "Cancelled Portion", "Cancel %", "Reason"]
    # Re-create header since we changed columns
    for i, col in enumerate(level_cols, 1):
        c = ws_levels.cell(row=1, column=i, value=col)
        c.font = hdr_font_w
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
    auto_width(ws_levels, level_cols)

    # ── Sheet 4: Book Snapshots ────────────────────────────────
    ws_books = wb.create_sheet("Book Snapshots")
    book_cols = ["IST Timestamp", "Outcome", "Bid Levels", "Ask Levels", "Best Bid", "Best Ask", "Spread", "Bid Depth $", "Ask Depth $"]
    style_header(ws_books, book_cols)
    auto_width(ws_books, book_cols)

    for e in events:
        if e["type"] != "book_snapshot":
            continue
        ws_books.append([
            e.get("ist", ""),
            e.get("outcome", ""),
            e.get("bid_levels"),
            e.get("ask_levels"),
            e.get("best_bid"),
            e.get("best_ask"),
            e.get("spread"),
            e.get("bid_depth_usd"),
            e.get("ask_depth_usd"),
        ])

    # ── Sheet 5: Summary ───────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40

    meta = next((e for e in events if e["type"] == "capture_start"), {})
    stop = next((e for e in events if e["type"] == "capture_stop"), {})

    summary_rows = [
        ("Market", meta.get("question", "")),
        ("Slug", meta.get("slug", "")),
        ("Condition ID", meta.get("condition_id", "")),
        ("Outcomes", ", ".join(meta.get("outcome_names", []))),
        ("Capture Start (IST)", meta.get("ist_start", "")),
        ("Capture End (IST)", stop.get("ist_end", "")),
        ("Duration (seconds)", stop.get("total_events", "")),
        ("", ""),
        ("Total Events", len(events)),
        ("WS Messages", stop.get("ws_messages", "")),
        ("WS Trade Ticks", stop.get("trades", "")),
        ("Level Increases", stop.get("order_adds", "")),
        ("Level Decreases (cancel)", stop.get("cancels", "")),
        ("REST Trades (with wallet)", len([e for e in events if e["type"] == "rest_trade"])),
        ("", ""),
        ("DATA SOURCES", ""),
        ("trade", "WS last_trade_price — DEFINITIVE fill, no wallet"),
        ("rest_trade", "REST data-api — DEFINITIVE fill WITH taker wallet"),
        ("level_increase", "WS price_change — INFERRED, size went up at level"),
        ("level_decrease_cancel", "WS price_change — INFERRED, size went down, no matching trade"),
        ("level_decrease_fill", "WS price_change — INFERRED, size went down WITH matching trade"),
        ("book_snapshot", "WS book — full L2 snapshot"),
        ("", ""),
        ("LIMITATIONS", ""),
        ("Order Type", "NOT available — GTC/GTD/FOK/FAK unknown from Market WS"),
        ("Individual Orders", "NOT available — only aggregated levels"),
        ("Cancel vs Fill", "HEURISTIC — based on trade tick correlation (2s window)"),
        ("Wallet on WS events", "NOT available — only REST trades have wallet"),
    ]
    for label, value in summary_rows:
        ws_summary.append([label, value])
        if label and label == label.upper():
            ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)

    wb.save(xlsx_path)
    print(f"\n{C_BOLD}Excel exported: {xlsx_path}{C_RESET}")
    print(f"  Sheets: All Events, Trades, Level Changes, Book Snapshots, Summary")
    return xlsx_path


def main():
    parser = argparse.ArgumentParser(description="Polymarket Live Event Capture")
    parser.add_argument("slug", help="Market slug")
    parser.add_argument("--duration", type=int, default=600, help="Capture duration in seconds (default: 600)")
    parser.add_argument("--start-at", type=str, default=None, help="IST time to start capture, e.g. '11:40' (waits until then)")
    args = parser.parse_args()

    try:
        asyncio.run(run(args.slug, args.duration, args.start_at))
    except KeyboardInterrupt:
        print("\nStopped.")

    # Export latest capture to Excel
    import glob
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    captures = sorted(glob.glob(os.path.join(project_root, "data", "**", "*.jsonl"), recursive=True))
    if captures:
        latest = captures[-1]
        try:
            export_to_excel(latest)
        except Exception as e:
            print(f"Excel export error: {e}")


if __name__ == "__main__":
    main()
