#!/usr/bin/env python3
"""Combined live cricket score logger — ESPN + Cricbuzz.

Polls both sources concurrently, merges into a single JSONL stream.
If one source dies the other continues. Deduplicates by score state.

Usage:
    python combined_score.py --espn 1491739 --cricbuzz 122731
    python combined_score.py --espn 1491739 --cricbuzz 122731 --duration 18000
"""

import argparse
import asyncio
import json
import os
import re
import time
from datetime import datetime, timezone, timedelta

import httpx

IST = timezone(timedelta(hours=5, minutes=30))

# ESPN
ESPN_URL = "https://site.web.api.espn.com/apis/site/v2/sports/cricket/{league_id}/summary?event={match_id}"
LEAGUE_IDS = [8676, 1410, 1359, 1457, 1001, 1283]

# Terminal colors
C_RESET = "\033[0m"
C_GREEN = "\033[92m"
C_RED = "\033[91m"
C_YELLOW = "\033[93m"
C_CYAN = "\033[96m"
C_MAGENTA = "\033[95m"
C_BOLD = "\033[1m"
C_DIM = "\033[90m"


def now_ist():
    t = time.time()
    ns = time.time_ns() % 1_000_000_000
    dt = datetime.fromtimestamp(t, tz=IST)
    return dt.strftime("%Y-%m-%d %H:%M:%S") + f".{ns:09d} IST"


def ist_short():
    ns = time.time_ns() % 1_000_000
    dt = datetime.now(IST)
    return dt.strftime("%H:%M:%S") + f".{ns:06d}"


class CombinedLogger:
    """Thread-safe (asyncio-safe) combined event logger."""

    def __init__(self, output_path: str):
        self.output_path = output_path
        self._file = open(output_path, "a")
        self.events: list[dict] = []
        self.seq = 0
        self._last_state = ""  # dedup key across sources
        self._lock = asyncio.Lock()

        # Per-source health tracking
        self.source_stats = {
            "espn": {"ok": 0, "err": 0, "last_ok": 0.0, "alive": True},
            "cricbuzz": {"ok": 0, "err": 0, "last_ok": 0.0, "alive": True},
        }

    async def log_event(self, source: str, event_type: str, data: dict) -> bool:
        """Log an event. Returns True if it was new (not deduped)."""
        # Build a dedup key from the score state
        dedup_key = f"{data.get('score', '')}|{data.get('overs', '')}|{data.get('last_ball', '')}"

        async with self._lock:
            if dedup_key == self._last_state and dedup_key:
                return False  # duplicate

            self._last_state = dedup_key
            self.seq += 1

            record = {
                "seq": self.seq,
                "ts_ns": time.time_ns(),
                "ts": time.time(),
                "ist": now_ist(),
                "source": source,
                "type": event_type,
                **data,
            }

            self._file.write(json.dumps(record) + "\n")
            self._file.flush()
            self.events.append(record)

            # Print
            src_color = C_MAGENTA if source == "espn" else C_CYAN
            src_tag = f"{src_color}{source:>8}{C_RESET}"

            if event_type == "wicket":
                ev_color = C_RED
            elif event_type == "boundary":
                ev_color = C_GREEN
            elif event_type == "ball":
                ev_color = C_CYAN
            elif event_type in ("complete", "post"):
                ev_color = C_YELLOW
            else:
                ev_color = C_DIM

            t = ist_short()
            score_str = data.get("score", "?")
            overs_str = data.get("overs", "")
            overs_part = f" ({overs_str} ov)" if overs_str else ""
            extras = ""
            if data.get("batsman"):
                extras += f" | {data['batsman']} {data.get('bat_runs', '')}({data.get('bat_balls', '')})"
            if data.get("bowler"):
                extras += f" | {data['bowler']} {data.get('bowl_overs', '')}-{data.get('bowl_runs', '')}-{data.get('bowl_wickets', '')}"
            if data.get("last_ball"):
                extras += f" | last: {data['last_ball']}"
            if data.get("recent_overs"):
                extras += f" | {data['recent_overs'][-30:]}"

            print(
                f"{C_DIM}{t}{C_RESET} {src_tag} {ev_color}{event_type:>8}{C_RESET} "
                f"{C_BOLD}{score_str}{C_RESET}{overs_part}{extras}"
            )
            return True

    def mark_source(self, source: str, ok: bool):
        s = self.source_stats[source]
        if ok:
            s["ok"] += 1
            s["last_ok"] = time.time()
            if not s["alive"]:
                s["alive"] = True
                print(f"{C_DIM}{ist_short()}{C_RESET} {C_GREEN}  {source} recovered{C_RESET}")
        else:
            s["err"] += 1

    def mark_dead(self, source: str, reason: str):
        s = self.source_stats[source]
        if s["alive"]:
            s["alive"] = False
            print(
                f"{C_DIM}{ist_short()}{C_RESET} {C_RED}  {source} DEAD: {reason}{C_RESET}"
            )

    def any_alive(self) -> bool:
        return any(s["alive"] for s in self.source_stats.values())

    def close(self):
        self._file.close()

    def export_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        xlsx_path = self.output_path.replace(".jsonl", ".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Ball by Ball"

        cols = [
            "IST Timestamp", "Source", "Event", "Score", "Overs", "Batting Team",
            "Batsman", "Bat Runs", "Bat Balls", "Bowler", "Bowl Overs",
            "Bowl Runs", "Bowl Wkts", "Run Rate", "Req Rate", "Target",
            "Last Ball", "Recent Overs", "Description",
        ]

        hdr_fill = PatternFill("solid", fgColor="1F2937")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.font = hdr_font
            c.fill = hdr_fill

        wicket_fill = PatternFill("solid", fgColor="FEE2E2")
        boundary_fill = PatternFill("solid", fgColor="DCFCE7")

        for e in self.events:
            row = [
                e.get("ist", ""), e.get("source", ""), e.get("type", ""),
                e.get("score", ""), e.get("overs", ""), e.get("batting_team", ""),
                e.get("batsman", ""), e.get("bat_runs", ""), e.get("bat_balls", ""),
                e.get("bowler", ""), e.get("bowl_overs", ""), e.get("bowl_runs", ""),
                e.get("bowl_wickets", ""), e.get("run_rate", ""), e.get("required_rate", ""),
                e.get("target", ""), e.get("last_ball", ""), e.get("recent_overs", ""),
                e.get("description", ""),
            ]
            ws.append(row)
            if e.get("type") == "wicket":
                for cell in ws[ws.max_row]:
                    cell.fill = wicket_fill
            elif e.get("type") == "boundary":
                for cell in ws[ws.max_row]:
                    cell.fill = boundary_fill

        for i, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 14)

        wb.save(xlsx_path)
        print(f"\n{C_BOLD}Excel: {xlsx_path}{C_RESET}")


# ─── ESPN Source ────────────────────────────────────────────────────────────


async def espn_find_league(client: httpx.AsyncClient, match_id: int) -> int | None:
    for lid in LEAGUE_IDS:
        try:
            url = ESPN_URL.format(league_id=lid, match_id=match_id)
            resp = await client.get(url)
            if resp.status_code == 200:
                d = resp.json()
                if d.get("header"):
                    return lid
        except Exception:
            continue
    return None


def parse_espn(data: dict) -> dict | None:
    header = data.get("header", {})
    competitions = header.get("competitions", [{}])
    if not competitions:
        return None

    comp = competitions[0]
    status_obj = comp.get("status", {})
    status_type = status_obj.get("type", {})
    status = status_type.get("description", "")
    state = status_type.get("state", "")

    competitors = comp.get("competitors", [])
    batting_team = ""
    scores = []
    for team in competitors:
        abbr = team.get("team", {}).get("abbreviation", "?")
        linescores = team.get("linescores", [])
        if linescores:
            for ls in linescores:
                sc = ls.get("score", "")
                if sc:
                    scores.append(f"{abbr} {sc}")
                    if ls.get("isBatting") and ls.get("isCurrent"):
                        batting_team = abbr
        elif team.get("score"):
            scores.append(f"{abbr} {team['score']}")

    score_str = " | ".join(scores) if scores else status

    situation = data.get("situation", {})
    last_ball = ""
    if isinstance(situation, dict):
        last_ball = situation.get("lastBallSummary", "")

    # Try to get overs from situation
    overs = ""
    if isinstance(situation, dict):
        overs = situation.get("currentOver", "")
        if not overs:
            overs = situation.get("oversRemaining", "")

    return {
        "score": score_str,
        "overs": str(overs),
        "batting_team": batting_team,
        "status": status,
        "state": state,
        "last_ball": last_ball,
    }


async def espn_loop(
    logger: CombinedLogger, match_id: int, poll_interval: float, stop_event: asyncio.Event
):
    max_consecutive_errors = 30  # ~30s at 1s poll before declaring dead

    async with httpx.AsyncClient(
        timeout=10,
        headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"},
    ) as client:
        print(f"{C_DIM}{ist_short()}{C_RESET} {C_MAGENTA}ESPN detecting league...{C_RESET}", end=" ", flush=True)
        league_id = await espn_find_league(client, match_id)
        if league_id is None:
            print(f"{C_RED}not found{C_RESET}")
            logger.mark_dead("espn", f"match {match_id} not found in any league")
            return
        print(f"{C_GREEN}league {league_id}{C_RESET}")

        url = ESPN_URL.format(league_id=league_id, match_id=match_id)
        consecutive_errors = 0

        while not stop_event.is_set():
            try:
                resp = await client.get(url)
                if resp.status_code != 200:
                    consecutive_errors += 1
                    if consecutive_errors >= max_consecutive_errors:
                        logger.mark_dead("espn", f"HTTP {resp.status_code} x{consecutive_errors}")
                    await asyncio.sleep(poll_interval)
                    continue

                data = resp.json()
                parsed = parse_espn(data)
                if not parsed:
                    await asyncio.sleep(poll_interval)
                    continue

                consecutive_errors = 0
                logger.mark_source("espn", True)

                event_type = "update"
                if parsed["state"] == "post":
                    event_type = "complete"
                elif "W" in parsed.get("last_ball", ""):
                    event_type = "wicket"

                await logger.log_event("espn", event_type, parsed)

                if parsed["state"] == "post":
                    logger.mark_dead("espn", "match complete")
                    return

            except asyncio.CancelledError:
                return
            except Exception as e:
                consecutive_errors += 1
                logger.mark_source("espn", False)
                if consecutive_errors >= max_consecutive_errors:
                    logger.mark_dead("espn", str(e))
                elif consecutive_errors % 5 == 0:
                    print(f"{C_DIM}{ist_short()}{C_RESET} {C_RED}ESPN err ({consecutive_errors}): {e}{C_RESET}")

            await asyncio.sleep(poll_interval)


# ─── Cricbuzz Source ────────────────────────────────────────────────────────


def _extract(pattern: str, html: str) -> str:
    m = re.search(pattern, html)
    return m.group(1) if m else ""


def parse_cricbuzz(html: str) -> dict | None:
    desc_match = re.search(r'<meta property="og:description" content="([^"]+)"', html)
    desc = desc_match.group(1).strip() if desc_match else ""
    if not desc:
        return None

    match_state = _extract(r'"matchState"\s*:\s*"([^"]+)"', html)
    score = _extract(r'"score"\s*:\s*"([^"]+)"', html)
    overs = _extract(r'"overs"\s*:\s*"([^"]+)"', html)
    target = _extract(r'"target"\s*:\s*"([^"]+)"', html)
    crr = _extract(r'"currentRunRate"\s*:\s*"([^"]+)"', html)
    rrr = _extract(r'"requiredRunRate"\s*:\s*"([^"]+)"', html)
    recent_overs = _extract(r'"recentOvsStats"\s*:\s*"([^"]+)"', html)
    bat_team = _extract(r'"batTeamName"\s*:\s*"([^"]+)"', html)
    batsman = _extract(r'"batName"\s*:\s*"([^"]+)"', html)
    bat_runs = _extract(r'"batRuns"\s*:\s*"([^"]+)"', html)
    bat_balls = _extract(r'"batBalls"\s*:\s*"([^"]+)"', html)
    bowler = _extract(r'"bowlName"\s*:\s*"([^"]+)"', html)
    bowl_overs = _extract(r'"bowlOvs"\s*:\s*"([^"]+)"', html)
    bowl_runs = _extract(r'"bowlRuns"\s*:\s*"([^"]+)"', html)
    bowl_wkts = _extract(r'"bowlWkts"\s*:\s*"([^"]+)"', html)

    # Parse score from description as fallback
    parsed_score = score
    parsed_overs = overs
    if not score:
        sc_match = re.search(r'(\w+)\s+(\d+/\d+)\s*\((\d+\.\d+)\)', desc)
        if sc_match:
            parsed_score = f"{sc_match.group(1)} {sc_match.group(2)}"
            parsed_overs = sc_match.group(3)

    # Detect event type
    event_type = "update"
    if match_state and "Complete" in match_state:
        event_type = "complete"
    # Wicket/boundary detection would need state tracking across calls,
    # handled by the loop below

    return {
        "description": desc,
        "match_state": match_state,
        "batting_team": bat_team,
        "score": parsed_score,
        "overs": parsed_overs,
        "target": target,
        "run_rate": crr,
        "required_rate": rrr,
        "recent_overs": recent_overs,
        "batsman": batsman,
        "bat_runs": bat_runs,
        "bat_balls": bat_balls,
        "bowler": bowler,
        "bowl_overs": bowl_overs,
        "bowl_runs": bowl_runs,
        "bowl_wickets": bowl_wkts,
    }


async def cricbuzz_loop(
    logger: CombinedLogger, match_id: int, poll_interval: float, stop_event: asyncio.Event
):
    url = f"https://www.cricbuzz.com/live-cricket-scores/{match_id}"
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
    max_consecutive_errors = 15  # ~45s at 3s poll

    last_recent_overs = ""
    last_overs = ""
    last_state_key = ""
    consecutive_errors = 0

    async with httpx.AsyncClient(timeout=10, headers=headers) as client:
        while not stop_event.is_set():
            try:
                bust_url = f"{url}?cb={int(time.time() * 1000)}"
                resp = await client.get(bust_url, headers={
                    "Cache-Control": "no-cache",
                    "Pragma": "no-cache",
                }, follow_redirects=True)

                if resp.status_code != 200:
                    consecutive_errors += 1
                    if consecutive_errors >= max_consecutive_errors:
                        logger.mark_dead("cricbuzz", f"HTTP {resp.status_code} x{consecutive_errors}")
                    await asyncio.sleep(poll_interval)
                    continue

                parsed = parse_cricbuzz(resp.text)
                if not parsed:
                    await asyncio.sleep(poll_interval)
                    continue

                consecutive_errors = 0
                logger.mark_source("cricbuzz", True)

                # Change detection
                state_key = f"{parsed.get('description', '')}|{parsed.get('score', '')}|{parsed.get('overs', '')}|{parsed.get('recent_overs', '')}"
                if state_key == last_state_key and state_key:
                    await asyncio.sleep(poll_interval)
                    continue
                last_state_key = state_key

                # Classify event
                ro = parsed.get("recent_overs", "")
                ov = parsed.get("overs", "")
                event_type = "update"

                if parsed.get("match_state") and "Complete" in parsed["match_state"]:
                    event_type = "complete"
                elif "W" in ro and "W" not in last_recent_overs:
                    event_type = "wicket"
                elif ov != last_overs and ov:
                    # Check for boundary in new part of recent overs
                    if ro and last_recent_overs:
                        new_part = ro[len(last_recent_overs):] if ro.startswith(last_recent_overs.rstrip()) else ro
                        if "4" in new_part or "6" in new_part:
                            event_type = "boundary"
                        else:
                            event_type = "ball"
                    else:
                        event_type = "ball"

                last_recent_overs = ro
                last_overs = ov

                await logger.log_event("cricbuzz", event_type, parsed)

                if event_type == "complete":
                    logger.mark_dead("cricbuzz", "match complete")
                    return

            except asyncio.CancelledError:
                return
            except Exception as e:
                consecutive_errors += 1
                logger.mark_source("cricbuzz", False)
                if consecutive_errors >= max_consecutive_errors:
                    logger.mark_dead("cricbuzz", str(e))
                elif consecutive_errors % 5 == 0:
                    print(f"{C_DIM}{ist_short()}{C_RESET} {C_RED}CB err ({consecutive_errors}): {e}{C_RESET}")

            await asyncio.sleep(poll_interval)


# ─── Main ───────────────────────────────────────────────────────────────────


async def run(espn_id: int | None, cricbuzz_id: int | None, duration: int,
              espn_poll: float, cb_poll: float):
    ts = time.strftime("%Y%m%d_%H%M%S")
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    scores_dir = os.path.join(project_root, "data", "_scores")
    os.makedirs(scores_dir, exist_ok=True)
    tag = espn_id or cricbuzz_id
    outfile = os.path.join(scores_dir, f"{ts}_combined_{tag}.jsonl")

    logger = CombinedLogger(outfile)
    stop = asyncio.Event()

    print(f"\n{C_BOLD}Combined Cricket Score Logger{C_RESET}")
    if espn_id:
        print(f"  ESPN:     match {espn_id} (poll {espn_poll}s)")
    if cricbuzz_id:
        print(f"  Cricbuzz: match {cricbuzz_id} (poll {cb_poll}s)")
    print(f"  Duration: {duration}s ({duration // 60}m)")
    print(f"  Output:   {outfile}")
    print(f"  Started:  {now_ist()}")
    print(f"{'=' * 80}\n")

    tasks = []
    if espn_id:
        tasks.append(asyncio.create_task(espn_loop(logger, espn_id, espn_poll, stop)))
    if cricbuzz_id:
        tasks.append(asyncio.create_task(cricbuzz_loop(logger, cricbuzz_id, cb_poll, stop)))

    if not tasks:
        print(f"{C_RED}No sources specified. Use --espn and/or --cricbuzz.{C_RESET}")
        return

    start = time.time()
    try:
        while time.time() - start < duration:
            # Check if all sources are dead
            if not logger.any_alive() and time.time() - start > 10:
                all_done = all(t.done() for t in tasks)
                if all_done:
                    print(f"\n{C_YELLOW}All sources finished.{C_RESET}")
                    break
            await asyncio.sleep(1)
    except KeyboardInterrupt:
        pass

    stop.set()
    for t in tasks:
        t.cancel()
    await asyncio.gather(*tasks, return_exceptions=True)

    elapsed = int(time.time() - start)
    print(f"\n{'=' * 80}")
    print(f"Captured {logger.seq} events in {elapsed}s")
    for src, stats in logger.source_stats.items():
        print(f"  {src}: {stats['ok']} polls ok, {stats['err']} errors")
    print(f"Output: {outfile}")

    logger.close()
    try:
        logger.export_excel()
    except Exception as e:
        print(f"Excel export error: {e}")


def main():
    parser = argparse.ArgumentParser(description="Combined Cricket Score Logger")
    parser.add_argument("--espn", type=int, default=None, help="ESPN Cricinfo match ID")
    parser.add_argument("--cricbuzz", type=int, default=None, help="Cricbuzz match ID")
    parser.add_argument("--duration", type=int, default=18000, help="Duration in seconds (default: 5h)")
    parser.add_argument("--espn-poll", type=float, default=1.0, help="ESPN poll interval (default: 1s)")
    parser.add_argument("--cb-poll", type=float, default=3.0, help="Cricbuzz poll interval (default: 3s)")
    args = parser.parse_args()
    asyncio.run(run(args.espn, args.cricbuzz, args.duration, args.espn_poll, args.cb_poll))


if __name__ == "__main__":
    main()
