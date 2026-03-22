#!/usr/bin/env python3
"""Live cricket ball-by-ball score logger.

Scrapes Cricbuzz every 3 seconds for live score updates.
Logs every ball/event change with IST timestamp to JSONL + Excel.

Usage:
    python cricket_score.py 122709 --duration 14400
    python cricket_score.py <cricbuzz_match_id> --duration <seconds>

Find match ID: go to cricbuzz.com/live-cricket-scores, the number in the URL is the match ID.
"""

import argparse
import asyncio
import json
import os
import re
import sys
import time
from collections import deque
from datetime import datetime, timezone, timedelta

import httpx

IST = timezone(timedelta(hours=5, minutes=30))

def now_ist():
    t = time.time()
    ns = time.time_ns() % 1_000_000_000
    dt = datetime.fromtimestamp(t, tz=IST)
    return dt.strftime("%Y-%m-%d %H:%M:%S") + f".{ns:09d} IST"

def ist_short():
    ns = time.time_ns() % 1_000_000
    dt = datetime.now(IST)
    return dt.strftime("%H:%M:%S") + f".{ns:06d}"

C_RESET = "\033[0m"
C_GREEN = "\033[92m"
C_RED = "\033[91m"
C_YELLOW = "\033[93m"
C_CYAN = "\033[96m"
C_BOLD = "\033[1m"
C_DIM = "\033[90m"


class CricketLogger:
    def __init__(self, match_id: int, output_path: str):
        self.match_id = match_id
        self.url = f"https://www.cricbuzz.com/live-cricket-scores/{match_id}"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        }
        self.output_path = output_path
        self._file = open(output_path, "a")
        self.events: list[dict] = []
        self.count = 0

        # State tracking — detect changes
        self._last_score = ""
        self._last_overs = ""
        self._last_recent_overs = ""
        self._last_desc = ""
        self._last_batsmen = ""
        self._last_status = ""

    def _log(self, event_type: str, data: dict):
        record = {
            "seq": self.count,
            "ts_ns": time.time_ns(),
            "ts": time.time(),
            "ist": now_ist(),
            "type": event_type,
            **data,
        }
        self._file.write(json.dumps(record) + "\n")
        self._file.flush()
        self.events.append(record)
        self.count += 1

    async def poll(self):
        """Fetch live score page once and extract all data."""
        # Cache-bust to get fresh data every poll
        bust_url = f"{self.url}?cb={int(time.time()*1000)}"
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(bust_url, headers={
                **self.headers,
                'Cache-Control': 'no-cache',
                'Pragma': 'no-cache',
            }, follow_redirects=True)
            if resp.status_code != 200:
                return

        html = resp.text

        # Extract score from og:description (most reliable, updates every ball)
        desc_match = re.search(r'<meta property="og:description" content="([^"]+)"', html)
        desc = desc_match.group(1).strip() if desc_match else ""

        # Extract detailed data from embedded JSON
        match_state = self._extract(r'"matchState"\s*:\s*"([^"]+)"', html)
        score = self._extract(r'"score"\s*:\s*"([^"]+)"', html)
        overs = self._extract(r'"overs"\s*:\s*"([^"]+)"', html)
        target = self._extract(r'"target"\s*:\s*"([^"]+)"', html)
        crr = self._extract(r'"currentRunRate"\s*:\s*"([^"]+)"', html)
        rrr = self._extract(r'"requiredRunRate"\s*:\s*"([^"]+)"', html)
        recent_overs = self._extract(r'"recentOvsStats"\s*:\s*"([^"]+)"', html)

        # Batsmen/bowler info
        batsman1 = self._extract(r'"batName"\s*:\s*"([^"]+)"', html)
        bat1_runs = self._extract(r'"batRuns"\s*:\s*"([^"]+)"', html)
        bat1_balls = self._extract(r'"batBalls"\s*:\s*"([^"]+)"', html)
        bowler = self._extract(r'"bowlName"\s*:\s*"([^"]+)"', html)
        bowl_overs = self._extract(r'"bowlOvs"\s*:\s*"([^"]+)"', html)
        bowl_runs = self._extract(r'"bowlRuns"\s*:\s*"([^"]+)"', html)
        bowl_wkts = self._extract(r'"bowlWkts"\s*:\s*"([^"]+)"', html)

        # Batting team
        bat_team = self._extract(r'"batTeamName"\s*:\s*"([^"]+)"', html)

        # Combine into state string for change detection
        state_key = f"{desc}|{score}|{overs}|{recent_overs}"

        if state_key == self._last_desc and desc:
            return  # No change

        # Detect what changed
        is_new_ball = (overs != self._last_overs and overs)
        is_wicket = "W" in (recent_overs or "") and "W" not in (self._last_recent_overs or "")
        is_boundary = False
        if recent_overs and self._last_recent_overs:
            new_part = recent_overs[len(self._last_recent_overs):] if recent_overs.startswith(self._last_recent_overs.rstrip()) else recent_overs
            is_boundary = "4" in new_part or "6" in new_part

        self._last_desc = state_key
        self._last_overs = overs
        self._last_recent_overs = recent_overs

        # Parse score from description
        # Format: "Follow RSA 36/3 (5.3) \n (Rubin Hermann 0(0) Jason Smith 5(3))"
        parsed_score = ""
        parsed_overs_desc = ""
        sc_match = re.search(r'(\w+)\s+(\d+/\d+)\s*\((\d+\.\d+)\)', desc)
        if sc_match:
            parsed_score = f"{sc_match.group(1)} {sc_match.group(2)}"
            parsed_overs_desc = sc_match.group(3)

        data = {
            "description": desc,
            "match_state": match_state,
            "batting_team": bat_team,
            "score": score or parsed_score,
            "overs": overs or parsed_overs_desc,
            "target": target,
            "run_rate": crr,
            "required_rate": rrr,
            "recent_overs": recent_overs,
            "batsman": batsman1,
            "bat_runs": bat1_runs,
            "bat_balls": bat1_balls,
            "bowler": bowler,
            "bowl_overs": bowl_overs,
            "bowl_runs": bowl_runs,
            "bowl_wickets": bowl_wkts,
        }

        # Classify event
        if is_wicket:
            event_type = "wicket"
            color = C_RED
        elif is_boundary:
            event_type = "boundary"
            color = C_GREEN
        elif is_new_ball:
            event_type = "ball"
            color = C_CYAN
        else:
            event_type = "update"
            color = C_DIM

        self._log(event_type, data)

        # Print
        t = ist_short()
        score_str = data["score"] or parsed_score or "?"
        overs_str = data["overs"] or "?"
        extras = ""
        if data["batsman"]:
            extras = f" | {data['batsman']} {data['bat_runs']}({data['bat_balls']})"
        if data["bowler"]:
            extras += f" | {data['bowler']} {data['bowl_overs']}-{data['bowl_runs']}-{data['bowl_wickets']}"
        if data["recent_overs"]:
            extras += f" | {data['recent_overs'][-30:]}"

        print(f"{C_DIM}{t}{C_RESET} {color}{event_type:>8}{C_RESET} "
              f"{C_BOLD}{score_str}{C_RESET} ({overs_str} ov)"
              f"{extras}")

    def _extract(self, pattern: str, html: str) -> str:
        m = re.search(pattern, html)
        return m.group(1) if m else ""

    def close(self):
        self._file.close()

    def export_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        xlsx_path = self.output_path.replace(".jsonl", ".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Ball by Ball"

        cols = ["IST Timestamp", "Event", "Score", "Overs", "Batting Team",
                "Batsman", "Bat Runs", "Bat Balls", "Bowler", "Bowl Overs",
                "Bowl Runs", "Bowl Wkts", "Run Rate", "Req Rate", "Target",
                "Recent Overs", "Description"]

        hdr_fill = PatternFill("solid", fgColor="1F2937")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.font = hdr_font
            c.fill = hdr_fill

        wicket_fill = PatternFill("solid", fgColor="FEE2E2")
        boundary_fill = PatternFill("solid", fgColor="DCFCE7")

        for e in self.events:
            row = [
                e.get("ist", ""),
                e.get("type", ""),
                e.get("score", ""),
                e.get("overs", ""),
                e.get("batting_team", ""),
                e.get("batsman", ""),
                e.get("bat_runs", ""),
                e.get("bat_balls", ""),
                e.get("bowler", ""),
                e.get("bowl_overs", ""),
                e.get("bowl_runs", ""),
                e.get("bowl_wickets", ""),
                e.get("run_rate", ""),
                e.get("required_rate", ""),
                e.get("target", ""),
                e.get("recent_overs", ""),
                e.get("description", ""),
            ]
            ws.append(row)
            if e.get("type") == "wicket":
                for cell in ws[ws.max_row]:
                    cell.fill = wicket_fill
            elif e.get("type") == "boundary":
                for cell in ws[ws.max_row]:
                    cell.fill = boundary_fill

        # Auto-width
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 14)

        wb.save(xlsx_path)
        print(f"\n{C_BOLD}Cricket Excel: {xlsx_path}{C_RESET}")
        return xlsx_path


async def run(match_id: int, duration: int, poll_interval: float = 3.0):
    ts = time.strftime("%Y%m%d_%H%M%S")
    os.makedirs("captures", exist_ok=True)
    outfile = f"captures/{ts}_cricket_{match_id}.jsonl"

    logger = CricketLogger(match_id, outfile)

    print(f"\n{C_BOLD}Cricket Ball-by-Ball Logger{C_RESET}")
    print(f"Match ID: {match_id}")
    print(f"Cricbuzz: https://www.cricbuzz.com/live-cricket-scores/{match_id}")
    print(f"Duration: {duration}s ({duration//60}m)")
    print(f"Poll:     every {poll_interval}s")
    print(f"Output:   {outfile}")
    print(f"Started:  {now_ist()}")
    print(f"{'='*70}\n")

    start = time.time()
    try:
        while time.time() - start < duration:
            try:
                await logger.poll()
            except Exception as e:
                print(f"{C_DIM}{ist_short()}{C_RESET} {C_RED}ERROR{C_RESET} {e}")
            await asyncio.sleep(poll_interval)
    except KeyboardInterrupt:
        pass

    print(f"\n{'='*70}")
    print(f"Captured {logger.count} events in {int(time.time()-start)}s")

    logger.close()
    try:
        logger.export_excel()
    except Exception as e:
        print(f"Excel export error: {e}")


def main():
    parser = argparse.ArgumentParser(description="Cricket Ball-by-Ball Logger")
    parser.add_argument("match_id", type=int, help="Cricbuzz match ID (from URL)")
    parser.add_argument("--duration", type=int, default=14400, help="Duration in seconds (default: 4h)")
    parser.add_argument("--poll", type=float, default=3.0, help="Poll interval seconds (default: 3)")
    args = parser.parse_args()
    asyncio.run(run(args.match_id, args.duration, args.poll))


if __name__ == "__main__":
    main()
