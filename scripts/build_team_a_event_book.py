"""For each of the 17 complete matches, build a per-event sheet with:

  Header rows (top of sheet):
    slug, team A (winner), team B, final score line, who won.

  Per-ball rows:
    time, score (TEAM run/wkt (overs)), event, then 8 odds columns:
      t-50, t-40, t-30, t-20, t-10, t, t+10, t+20  (team A best bid)

Event timeline uses the FASTER source per ball:
    t = min(capture.local_ts_ms, espn.bbb_ts_ms)
  When a ball is in only one feed, that source's timestamp is used.

Snapshot lookup: nearest book_snapshot for team A's asset, in either direction,
no time-window cap (we want to fill every cell).

Output: captures/team_a_event_book.xlsx (one sheet per match + _summary).
"""
from __future__ import annotations

import bisect
import json
import sqlite3
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CAPTURES = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures")
ESPN_WB = CAPTURES / "espn_ipl2026_ballbyball.xlsx"
OUT = CAPTURES / "team_a_event_book.xlsx"
IST = timezone(timedelta(hours=5, minutes=30))

# The 17 complete matches (slugs WITHOUT cricipl- prefix to match ESPN tab names)
COMPLETE_SLUGS = [
    "che-kol-2026-04-14",
    "del-pun-2026-04-25",
    "guj-kol-2026-04-17",
    "guj-mum-2026-04-20",
    "kol-luc-2026-04-09",
    "kol-raj-2026-04-19",
    "luc-raj-2026-04-22",
    "mum-che-2026-04-23",
    "mum-pun-2026-04-16",
    "pun-luc-2026-04-19",
    "pun-sun-2026-04-11",
    "raj-sun-2026-04-25",
    "roy-del-2026-04-18",
    "roy-guj-2026-04-24",
    "roy-luc-2026-04-15",
    "sun-che-2026-04-18",
    "sun-del-2026-04-21",
]

OFFSETS_S = [-50, -40, -30, -20, -10, 0, 10, 20]
OFFSET_LABELS = ["t-50", "t-40", "t-30", "t-20", "t-10", "t", "t+10", "t+20"]

# Outcome-name -> short tag used in slug
NAME_TO_SHORT = {
    "Chennai Super Kings": "CSK", "Kolkata Knight Riders": "KKR",
    "Mumbai Indians": "MI", "Royal Challengers Bengaluru": "RCB",
    "Royal Challengers Bangalore": "RCB",
    "Rajasthan Royals": "RR", "Delhi Capitals": "DC",
    "Gujarat Titans": "GT", "Lucknow Super Giants": "LSG",
    "Punjab Kings": "PBKS", "Sunrisers Hyderabad": "SRH",
}
SHORT_TO_SLUG_TAG = {
    "CSK": "che", "KKR": "kol", "MI": "mum", "RCB": "roy", "RR": "raj",
    "DC": "del", "GT": "guj", "LSG": "luc", "PBKS": "pun", "SRH": "sun",
}


def overs_to_balls(overs) -> int:
    if overs is None:
        return 0
    s = str(overs)
    whole, _, frac = s.partition(".")
    try:
        return int(whole) * 6 + (int(frac) if frac else 0)
    except ValueError:
        return 0


def detect_split(rows: list[tuple]) -> int | None:
    for i in range(1, len(rows)):
        gap = (rows[i][1] - rows[i - 1][1]) / 1000.0
        if (
            gap >= 300
            and overs_to_balls(rows[i][4]) < overs_to_balls(rows[i - 1][4])
            and rows[i][2] < rows[i - 1][2]
            and rows[i][2] <= 10
        ):
            return i
    return None


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

@dataclass
class CaptureBall:
    innings: int
    balls_idx: int
    overs_str: str
    ts_ms: int
    score_str: str  # "192/5 (20.0)"
    signal: str


def load_capture(slug: str) -> tuple[list[CaptureBall], dict, sqlite3.Connection]:
    db = next(CAPTURES.glob(f"match_capture_cricipl-{slug}_*.db"))
    conn = sqlite3.connect(db)
    meta = dict(conn.execute("SELECT key, value FROM match_meta").fetchall())
    rows = conn.execute(
        "SELECT id, local_ts_ms, COALESCE(runs, 0), COALESCE(wickets, 0), "
        "COALESCE(overs, '0.0'), score_str, signal_type "
        "FROM cricket_events WHERE signal_type != '?' ORDER BY id"
    ).fetchall()
    if not rows:
        return [], meta, conn
    split = detect_split(rows)
    out: list[CaptureBall] = []
    inn = 1
    for i, (rid, ts, runs, wkts, overs, score_str, sig) in enumerate(rows):
        if split is not None and i == split:
            inn = 2
        out.append(CaptureBall(
            innings=inn,
            balls_idx=overs_to_balls(overs),
            overs_str=str(overs),
            ts_ms=ts,
            score_str=score_str or f"{runs}/{wkts} ({overs})",
            signal=str(sig) if sig is not None else "",
        ))
    return out, meta, conn


@dataclass
class EspnBall:
    innings: int
    balls_idx: int
    overs_str: str
    ts_ms: int
    play_type: str
    score_value: int
    batting_team: str        # ESPN's team.abbreviation -> "SRH", "DC", etc.
    is_wide: bool
    is_noball: bool
    is_legbye: bool
    is_bye: bool
    home_score: str
    away_score: str
    short_text: str


def load_espn(ws) -> list[EspnBall]:
    headers = [c.value for c in ws[6]]
    out: list[EspnBall] = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] is None:
            continue
        rec = dict(zip(headers, row))
        out.append(EspnBall(
            innings=int(rec.get("innings") or 0),
            balls_idx=overs_to_balls(rec.get("over_actual")),
            overs_str=str(rec.get("over_actual")),
            ts_ms=int(rec.get("bbb_ts_ms")),
            play_type=str(rec.get("play_type") or ""),
            score_value=int(rec.get("score_value") or 0),
            batting_team=str(rec.get("batting_team") or ""),
            is_wide=bool(rec.get("is_wide")),
            is_noball=bool(rec.get("is_noball")),
            is_legbye=bool(rec.get("is_legbye")),
            is_bye=bool(rec.get("is_bye")),
            home_score=str(rec.get("home_score") or ""),
            away_score=str(rec.get("away_score") or ""),
            short_text=str(rec.get("short_text") or ""),
        ))
    return out


def _score_advances(s: str) -> bool:
    """True if score string is non-trivial (a team has scored at least one run)."""
    if not s:
        return False
    head = s.split("(")[0].strip()
    if not head or head == "0":
        return False
    parts = head.split("/")
    try:
        return int(parts[0]) > 0
    except ValueError:
        return False


def infer_inn1_home_bats(esp_balls: list[EspnBall]) -> bool:
    """Look at ESPN's home/away score progression in innings 1 to decide whether
    the HOME team batted first. Use the FIRST inn1 ball where either score
    has actually advanced past 0.
    """
    for e in esp_balls:
        if e.innings != 1:
            continue
        if _score_advances(e.home_score):
            return True
        if _score_advances(e.away_score):
            return False
    return True  # default if entire inn1 was 0/0 (impossible in practice)


# ---------------------------------------------------------------------------
# Event timeline construction
# ---------------------------------------------------------------------------

def normalize_event(cap: CaptureBall | None, esp: EspnBall | None) -> str:
    """Return a short event code: 0/1/2/3/4/5/6/W/WD/NB/LB/B."""
    if esp is not None:
        if esp.is_wide:
            return "WD"
        if esp.is_noball:
            return "NB"
        if esp.is_legbye:
            return "LB"
        if esp.is_bye:
            return "B"
        pt = esp.play_type.lower()
        if pt == "out":
            return "W"
        if pt == "four":
            return "4"
        if pt == "six":
            return "6"
        if pt == "no run":
            return "0"
        if pt == "run":
            return str(esp.score_value)
    # fallback to capture signal
    if cap is not None:
        return cap.signal
    return ""


@dataclass
class UnifiedBall:
    innings: int
    balls_idx: int
    overs_str: str
    t_ms: int                 # min(capture, espn)
    source: str               # "capture", "espn", "both", "capture-only", "espn-only"
    event_code: str
    batting_team_short: str   # e.g. "CSK"
    score_str_raw: str        # "192/5 (20.0)"
    short_text: str


def espn_batting_score(e: EspnBall) -> str:
    """ESPN puts the batting team's running score in either home_score or
    away_score — whichever currently contains '/'. The non-batting team's
    field is just '0'."""
    if "/" in e.home_score:
        return e.home_score
    if "/" in e.away_score:
        return e.away_score
    return "0/0"


def build_unified(
    cap_balls: list[CaptureBall],
    esp_balls: list[EspnBall],
    inn1_batting_short: str,
    inn2_batting_short: str,
) -> list[UnifiedBall]:
    """ESPN as canonical sequence. For each ESPN delivery we attach capture's
    timestamp ONLY when the delivery is a legal ball — that's the only event
    type whose ball_index alignment between the two feeds is reliable. (For
    extras the two feeds tag them differently: ESPN at the 6th-ball position
    on which the extra was bowled, capture at the previous legal ball's
    position. Pairing extras by ball_index would mismatch wides with legals
    in the same bucket.)

    Result: one row per ESPN delivery, with unified t = min(ESPN_ts,
    capture_legal_ts) for legals and = ESPN_ts for extras.
    """
    # Index capture's LEGAL balls by (innings, ball_idx) — these align cleanly.
    # We compute is_legal by walking the capture sequence and tracking ball
    # increments.
    cap_legal_by_key: dict[tuple, int] = {}
    last_balls_idx = {1: 0, 2: 0}
    for c in cap_balls:
        prev = last_balls_idx.get(c.innings, 0)
        if c.balls_idx == prev + 1:
            # legal delivery
            cap_legal_by_key.setdefault((c.innings, c.balls_idx), c.ts_ms)
            last_balls_idx[c.innings] = c.balls_idx

    pairs: list[UnifiedBall] = []
    for e in esp_balls:
        innings = e.innings
        batting = inn1_batting_short if innings == 1 else inn2_batting_short
        is_legal = not (e.is_wide or e.is_noball)  # wide/no-ball don't count
        if is_legal:
            cap_ts = cap_legal_by_key.get((innings, e.balls_idx))
            if cap_ts is not None:
                t = min(cap_ts, e.ts_ms)
                src = "both"
            else:
                t = e.ts_ms
                src = "espn-only"
        else:
            t = e.ts_ms
            src = "espn-only"  # extras: capture timing isn't reliably alignable
        running = espn_batting_score(e)
        score = f"{running} ({e.overs_str})"
        pairs.append(UnifiedBall(
            innings=innings,
            balls_idx=e.balls_idx,
            overs_str=e.overs_str,
            t_ms=t,
            source=src,
            event_code=normalize_event(None, e),
            batting_team_short=batting,
            score_str_raw=score,
            short_text=e.short_text,
        ))
    # Sort in CRICKET ORDER, not timestamp order. ESPN's bbbTimestamp is
    # occasionally out of order (commentator backfills a late entry minutes
    # later), which can flip the score progression. Sorting by
    # (innings, ball_idx) recovers cricket order; Python's stable sort
    # preserves the within-bucket ESPN seq order for extras at the same
    # ball position (e.g. consecutive wides).
    pairs.sort(key=lambda u: (u.innings, u.balls_idx))
    return pairs


# ---------------------------------------------------------------------------
# Team A (winner) determination + verification
# ---------------------------------------------------------------------------

def team_outcomes(meta: dict) -> tuple[list[str], list[str]]:
    token_ids = json.loads(meta.get("token_ids", "[]"))
    outcome_names = json.loads(meta.get("outcome_names", "[]"))
    return token_ids, outcome_names


def determine_winner(
    cap_balls: list[CaptureBall],
    esp_balls: list[EspnBall],
    home_short: str,
    away_short: str,
) -> tuple[str, str, str]:
    """Returns (winner_short, inn1_batting_short, inn2_batting_short)."""
    # Find innings 1 final (last cap ball with innings=1) and innings 2 final.
    inn1_final = None
    inn2_final = None
    for c in cap_balls:
        if c.innings == 1:
            inn1_final = c
        elif c.innings == 2:
            inn2_final = c

    # Use ESPN as backup if capture missed innings 2 ball
    inn1_runs = inn1_wkts = inn2_runs = inn2_wkts = 0
    if inn1_final:
        inn1_runs, inn1_wkts = parse_score(inn1_final.score_str)
    if inn2_final:
        inn2_runs, inn2_wkts = parse_score(inn2_final.score_str)
    if not inn2_final:
        # Fall back to ESPN
        for e in esp_balls:
            if e.innings == 2:
                inn2_runs, inn2_wkts = parse_score(e.home_score)

    # Identify which team batted first using ESPN: team field on first inn1 item.
    # If unavailable, default: home_short bats first.
    inn1_batting = home_short
    for e in esp_balls:
        if e.innings == 1 and e.batting_team:
            inn1_batting = e.batting_team
            break
    inn2_batting = away_short if inn1_batting == home_short else home_short

    # Decide winner from end states
    if inn2_runs > inn1_runs:
        winner = inn2_batting        # chase complete
    elif inn2_wkts >= 10 and inn2_runs <= inn1_runs:
        winner = inn1_batting        # all out, fell short
    elif inn1_runs > inn2_runs:
        winner = inn1_batting        # 20 overs done, lower score
    else:
        winner = inn1_batting        # tie default; rare

    return winner, inn1_batting, inn2_batting


def parse_score(s: str) -> tuple[int, int]:
    """'192/5 (20.0)' or '192/5' -> (192, 5)."""
    if not s:
        return 0, 0
    head = s.split("(")[0].strip()
    parts = head.split("/")
    try:
        return int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
    except (ValueError, IndexError):
        return 0, 0


def verify_winner_via_book(
    conn: sqlite3.Connection,
    token_ids: list[str],
    outcome_names: list[str],
    last_event_ts_ms: int,
    winner_short: str,
) -> tuple[str, dict]:
    """Look up the latest book snapshot for each token within ±10 min of match end
    and check which one has bid1_p > 0.95 — that's the winner per the book.
    """
    info = {}
    book_winner_short = None
    for tid, oname in zip(token_ids, outcome_names):
        row = conn.execute(
            """SELECT bid1_p, ask1_p, mid_price, local_ts_ms FROM book_snapshots
               WHERE asset_id = ? AND local_ts_ms <= ?
               ORDER BY local_ts_ms DESC LIMIT 1""",
            (tid, last_event_ts_ms + 600_000),
        ).fetchone()
        short = NAME_TO_SHORT.get(oname, oname)
        info[short] = row
        if row and row[0] and row[0] >= 0.95:
            book_winner_short = short
    verdict = "match" if book_winner_short == winner_short else \
              ("mismatch" if book_winner_short else "inconclusive")
    return verdict, info


# ---------------------------------------------------------------------------
# Snapshot lookup (nearest in either direction) for team A
# ---------------------------------------------------------------------------

def load_book_for_asset(conn: sqlite3.Connection, asset_id: str) -> tuple[list[int], list[float]]:
    rows = conn.execute(
        "SELECT local_ts_ms, bid1_p FROM book_snapshots "
        "WHERE asset_id = ? AND bid1_p IS NOT NULL ORDER BY local_ts_ms",
        (asset_id,),
    ).fetchall()
    return [r[0] for r in rows], [r[1] for r in rows]


def nearest_value(times: list[int], values: list[float], target_ms: int) -> float | None:
    if not times:
        return None
    i = bisect.bisect_left(times, target_ms)
    candidates = []
    if i < len(times):
        candidates.append((abs(times[i] - target_ms), values[i]))
    if i > 0:
        candidates.append((abs(times[i - 1] - target_ms), values[i - 1]))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
HEADER_FONT = Font(bold=True)


def write_sheet(
    wb: Workbook,
    slug: str,
    team_a_short: str,
    team_b_short: str,
    inn1_batting_short: str,
    inn2_batting_short: str,
    inn1_final_runs: int,
    inn1_final_wkts: int,
    inn2_final_runs: int,
    inn2_final_wkts: int,
    end_reason: str,
    timeline: list[UnifiedBall],
    book_times: list[int],
    book_bids: list[float],
    verify_verdict: str,
) -> None:
    ws = wb.create_sheet(title=slug[:31])

    ws.append([f"slug: {slug}",
               f"team A (winner): {team_a_short}",
               f"team B: {team_b_short}",
               f"verify-vs-book: {verify_verdict}"])
    ws.append([f"{inn1_batting_short} {inn1_final_runs}/{inn1_final_wkts}",
               f"{inn2_batting_short} {inn2_final_runs}/{inn2_final_wkts}",
               f"end: {end_reason}",
               f"events: {len(timeline)}"])
    ws.append([])
    headers = ["time_ist", "score", "event", "innings", "overs"] + OFFSET_LABELS
    ws.append(headers)
    for c in ws[4]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for u in timeline:
        ist_dt = datetime.fromtimestamp(u.t_ms / 1000, IST)
        # score_str_raw already has "(overs)"; strip it so we don't print twice.
        bare = u.score_str_raw.split(" (")[0]
        score = f"{u.batting_team_short} {bare} ({u.overs_str})"
        row = [
            ist_dt.strftime("%H:%M:%S.%f")[:-3],
            score,
            u.event_code,
            u.innings,
            u.overs_str,
        ]
        for off in OFFSETS_S:
            target = u.t_ms + off * 1000
            v = nearest_value(book_times, book_bids, target)
            row.append(round(v, 4) if v is not None else None)
        ws.append(row)

    widths = {"time_ist": 14, "score": 22, "event": 6, "innings": 6,
              "overs": 7}
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 9)


def build_summary(wb: Workbook, summary_rows: list[list]) -> None:
    if "_summary" in wb.sheetnames:
        wb.remove(wb["_summary"])
    ws = wb.create_sheet(title="_summary", index=0)
    ws.append(["slug", "team A (winner)", "team B", "inn1", "inn2",
               "end_reason", "events", "verify_vs_book"])
    for r in summary_rows:
        ws.append(r)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    if not ESPN_WB.exists():
        print(f"missing {ESPN_WB}", file=sys.stderr)
        return 1
    espn_wb = load_workbook(ESPN_WB, read_only=True)

    out_wb = Workbook()
    default = out_wb.active
    if default is not None:
        out_wb.remove(default)

    summary_rows: list[list] = []

    for slug in COMPLETE_SLUGS:
        print(f"\n=== {slug} ===")
        cap_balls, meta, conn = load_capture(slug)
        if slug not in espn_wb.sheetnames:
            print(f"  ESPN tab missing, skipping")
            conn.close()
            continue
        esp_balls = load_espn(espn_wb[slug])
        token_ids, outcome_names = team_outcomes(meta)
        if len(outcome_names) != 2:
            print("  bad meta, skipping")
            conn.close()
            continue

        # 1. Determine batting order DIRECTLY from ESPN's team field per ball.
        inn1_batting_set = {e.batting_team for e in esp_balls
                            if e.innings == 1 and e.batting_team}
        inn2_batting_set = {e.batting_team for e in esp_balls
                            if e.innings == 2 and e.batting_team}
        inn1_batting = next(iter(inn1_batting_set)) if inn1_batting_set else ""
        inn2_batting = next(iter(inn2_batting_set)) if inn2_batting_set else ""
        if not inn1_batting or not inn2_batting:
            print("  ESPN missing batting team field; skipping")
            conn.close()
            continue

        # 2. Identify winner from book WITHIN THE MATCH WINDOW (not all-time
        #    latest, which can drift hours after resolution into stale orders).
        last_event_ts = max(
            (c.ts_ms for c in cap_balls),
            default=max((e.ts_ms for e in esp_balls), default=0),
        )
        match_end_window = (last_event_ts - 30 * 60_000, last_event_ts + 5 * 60_000)
        winner_short = loser_short = None
        winner_token = None
        for tid, oname in zip(token_ids, outcome_names):
            r = conn.execute(
                "SELECT bid1_p, ask1_p FROM book_snapshots "
                "WHERE asset_id = ? AND local_ts_ms BETWEEN ? AND ? "
                "AND (bid1_p IS NOT NULL OR ask1_p IS NOT NULL) "
                "ORDER BY local_ts_ms DESC LIMIT 1",
                (tid, match_end_window[0], match_end_window[1]),
            ).fetchone()
            short = NAME_TO_SHORT.get(oname, oname)
            bid1 = r[0] if r else None
            ask1 = r[1] if r else None
            high = max(v for v in (bid1, ask1) if v is not None) if r else None
            low = min(v for v in (bid1, ask1) if v is not None) if r else None
            if high is not None and high >= 0.95:
                winner_short, winner_token = short, tid
            elif low is not None and low <= 0.05:
                loser_short = short
        if winner_short is None and loser_short is not None:
            # winner is the other team
            for tid, oname in zip(token_ids, outcome_names):
                short = NAME_TO_SHORT.get(oname, oname)
                if short != loser_short:
                    winner_short, winner_token = short, tid
                    break
        if winner_short is None:
            print(f"  could not determine winner from book; skipping")
            conn.close()
            continue
        if loser_short is None:
            for oname in outcome_names:
                short = NAME_TO_SHORT.get(oname, oname)
                if short != winner_short:
                    loser_short = short
                    break

        # 3. Compute final scores and end_reason from ESPN.
        inn1_finals = [e for e in esp_balls if e.innings == 1]
        inn2_finals = [e for e in esp_balls if e.innings == 2]
        inn1_runs, inn1_wkts = parse_score(espn_batting_score(inn1_finals[-1]))
        inn2_runs, inn2_wkts = parse_score(espn_batting_score(inn2_finals[-1]))

        if inn2_runs > inn1_runs:
            end_reason = "chase_complete"
        elif inn2_wkts >= 10:
            end_reason = "all_out"
        elif inn1_runs > inn2_runs:
            end_reason = "overs_complete_lower"
        else:
            end_reason = "tie"

        # 4. Cross-check: the team batting in inn1 should be either winner or
        #    loser per the book — if not, our team-name mapping is off.
        if winner_short is None:
            # fallback: derive from end_reason + ESPN batting attribution
            if end_reason == "chase_complete":
                winner_short = inn2_batting
                loser_short = inn1_batting
            else:
                winner_short = inn1_batting
                loser_short = inn2_batting
            winner_token = next(
                (tid for tid, oname in zip(token_ids, outcome_names)
                 if NAME_TO_SHORT.get(oname) == winner_short),
                None,
            )
            verify_status = "no-book-winner-fallback-to-espn"
        else:
            espn_winner = (inn2_batting if end_reason == "chase_complete"
                           else inn1_batting)
            verify_status = "match" if espn_winner == winner_short else "ESPN-vs-book-MISMATCH"
            if verify_status != "match":
                print(f"  WARN: book says {winner_short} won but ESPN says "
                      f"{espn_winner} won; trusting ESPN")
                winner_short, loser_short = espn_winner, (
                    inn1_batting if espn_winner == inn2_batting else inn2_batting
                )
                winner_token = next(
                    (tid for tid, oname in zip(token_ids, outcome_names)
                     if NAME_TO_SHORT.get(oname) == winner_short),
                    None,
                )

        if winner_token is None:
            print(f"  could not map {winner_short} to a token; skipping")
            conn.close()
            continue

        team_a = winner_short
        team_b = loser_short or (inn2_batting if team_a == inn1_batting else inn1_batting)

        # Build unified timeline with the (now correct) batting attribution.
        timeline = build_unified(cap_balls, esp_balls, inn1_batting, inn2_batting)
        if not timeline:
            print("  no events; skipping")
            conn.close()
            continue

        # Load team A's book series
        book_times, book_bids = load_book_for_asset(conn, winner_token)
        print(f"  team A (winner) = {team_a}  team B = {team_b}  "
              f"inn1 = {inn1_batting} {inn1_runs}/{inn1_wkts}  "
              f"inn2 = {inn2_batting} {inn2_runs}/{inn2_wkts}  "
              f"end = {end_reason}  events = {len(timeline)}  "
              f"book pts = {len(book_bids)}")

        write_sheet(out_wb, slug, team_a, team_b, inn1_batting, inn2_batting,
                    inn1_runs, inn1_wkts, inn2_runs, inn2_wkts, end_reason,
                    timeline, book_times, book_bids, verify_status)

        summary_rows.append([
            slug, team_a, team_b,
            f"{inn1_batting} {inn1_runs}/{inn1_wkts}",
            f"{inn2_batting} {inn2_runs}/{inn2_wkts}",
            end_reason, len(timeline), verify_status,
        ])

        conn.close()

    build_summary(out_wb, summary_rows)
    out_wb.save(OUT)
    print(f"\nSaved {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
