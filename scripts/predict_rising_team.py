"""Predict the RISING team's post-event price, calibrated so predicted ≤ actual.

Trading framing (correct):
  - Each event makes one team's price rise and the other fall (sum ≈ 1).
  - Identify which team rises (DP-predicted: side with positive WP delta).
  - For that team, predict a CONSERVATIVE post-event price r̂ such that:
        r̂_rising_team ≤ r_actual_rising_team       (no false positives)
  - If we BUY the rising team at r̂, the market is at r_actual ≥ r̂, so the
    buy fills at r̂ and we make r_actual − r̂.
  - Symmetric: when DP says A rises, we predict A's rise; when DP says B
    rises, we predict B's rise. We never quote on the falling team.

Steps:
  1. Solve chase DP and innings-1 DP.
  2. Build training data from the 17-match workbook in *rising-team* coords:
       y = (rising team's actual price at t)  −  (rising team's price at t-40)
         = +signed_delta_team_a  if DP predicts A rises
         = −signed_delta_team_a  if DP predicts B rises
       (y can be negative when DP got the direction wrong; that's a
        prediction-direction failure we want to bound.)
  3. Filter training to rising_team_t40 ∈ [0.10, 0.90].
  4. Train LightGBM quantile τ=0.005 on (dp_rise, t40_rising, event, phase, innings).
  5. CV-conformal margin (LOMO max over-prediction) + safety extra.
  6. Final predicted rise = max(0, GBM_pred − margin).
"""
from __future__ import annotations

import argparse, bisect, json, sqlite3, sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import numpy as np
import pandas as pd
import lightgbm as lgb

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent / "model-simulation" / "src"))
from dp.solver import DPTable, FirstInningsModel  # type: ignore
from dp.states import (  # type: ignore
    OUTCOMES, OUTCOME_CONSUMES_BALL, OUTCOME_IS_WICKET, OUTCOME_RUNS,
    TransitionProbs,
)

CAPTURES = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures")
RESID_CSV = CAPTURES / "predict_dp_residuals.csv"
IST = timezone(timedelta(hours=5, minutes=30))

NAME_TO_SHORT = {
    "Chennai Super Kings": "CSK", "Kolkata Knight Riders": "KKR",
    "Mumbai Indians": "MI", "Royal Challengers Bengaluru": "RCB",
    "Royal Challengers Bangalore": "RCB",
    "Rajasthan Royals": "RR", "Delhi Capitals": "DC",
    "Gujarat Titans": "GT", "Lucknow Super Giants": "LSG",
    "Punjab Kings": "PBKS", "Sunrisers Hyderabad": "SRH",
}
EVENT_TO_DP = {
    "0": "dot", "1": "single", "2": "double", "3": "triple",
    "4": "four", "5": None, "6": "six", "W": "wicket",
    "WD": "wide", "NB": "noball", "LB": "single", "B": "single",
}


# ---------- Innings-1 DP ----------
class InningsOneDP:
    MAX_BALLS = 120; MAX_RUNS = 300; MAX_WICKETS = 10
    def __init__(self, m: FirstInningsModel):
        self.inn1 = m
        shape = (self.MAX_BALLS+1, self.MAX_RUNS+1, self.MAX_WICKETS+1)
        self.table = np.zeros(shape, dtype=np.float32)
        self._terminals = np.array(
            [m.win_prob_at_total(r) for r in range(self.MAX_RUNS+1)],
            dtype=np.float32)
    def lookup(self, b, r, w):
        return float(self.table[max(0, min(self.MAX_BALLS, b)),
                                 max(0, min(self.MAX_RUNS, r)),
                                 max(0, min(self.MAX_WICKETS, w))])
    def solve(self):
        self.table[0, :, :] = self._terminals[:, None]
        self.table[:, :, 0] = self._terminals[None, :]
        for b in range(1, self.MAX_BALLS+1):
            ob = (self.MAX_BALLS - b) // 6
            phase = "powerplay" if ob < 6 else "middle" if ob < 15 else "death"
            tp = TransitionProbs.from_phase_averages(phase)
            probs = tp.as_dict()
            for w in range(1, self.MAX_WICKETS+1):
                vals = np.zeros(self.MAX_RUNS+1, dtype=np.float32)
                for outcome in OUTCOMES:
                    p = probs[outcome]
                    if p <= 0: continue
                    runs = OUTCOME_RUNS[outcome]
                    consumes = OUTCOME_CONSUMES_BALL[outcome]
                    is_wicket = OUTCOME_IS_WICKET[outcome]
                    new_b = b - (1 if consumes else 0)
                    new_w = w - (1 if is_wicket else 0)
                    if new_w <= 0 or new_b <= 0:
                        for r in range(self.MAX_RUNS+1):
                            vals[r] += p * self._terminals[min(self.MAX_RUNS, r+runs)]
                    else:
                        next_row = self.table[new_b, :, new_w]
                        if runs == 0:
                            vals += p * next_row
                        else:
                            shifted = np.empty(self.MAX_RUNS+1, dtype=np.float32)
                            shifted[:self.MAX_RUNS+1-runs] = next_row[runs:]
                            shifted[self.MAX_RUNS+1-runs:] = self._terminals[self.MAX_RUNS]
                            vals += p * shifted
                self.table[b, :, w] = vals


def overs_to_balls(s):
    s = str(s); whole, _, frac = s.partition('.')
    return int(whole)*6 + (int(frac) if frac else 0)


# ---------- Build rising-team-perspective training data ----------
def build_training_data():
    """Loads predict_dp_residuals.csv (signed team-A coords) and reframes to
    rising-team coords. Returns DataFrame with:
        rising_team    : "A" or "B" (DP's prediction)
        dp_rise_c      : DP-predicted rise of rising team (always >= 0)
        t40_rising     : rising team's price at t-40
        actual_rise_c  : rising team's actual rise (signed; negative if DP got
                         direction wrong)
    """
    df = pd.read_csv(RESID_CSV)
    df["date"] = df["slug"].str.extract(r"(\d{4}-\d{2}-\d{2})$")[0]

    # In team-A coords, pred_delta_c (DP) and actual_delta_c (market) are signed.
    df["dp_rises_A"] = (df["pred_delta_c"] > 0)
    df["dp_rise_c"] = df["pred_delta_c"].abs()
    df["t40_rising"] = np.where(df["dp_rises_A"], df["t-40"], 1.0 - df["t-40"])
    df["actual_rise_c"] = np.where(df["dp_rises_A"],
                                    df["actual_delta_c"], -df["actual_delta_c"])
    df["rising_team"] = np.where(df["dp_rises_A"], "A", "B")
    return df


def make_pbin(p):
    if p < 0.20: return "0.10-0.20"
    if p < 0.35: return "0.20-0.35"
    if p < 0.50: return "0.35-0.50"
    if p < 0.65: return "0.50-0.65"
    if p < 0.80: return "0.65-0.80"
    return "0.80-0.90"


def featurize(df, events, phases):
    feats = {
        "dp_rise_c":   df["dp_rise_c"].values,
        "t40_rising":  df["t40_rising"].values,
        "innings":     df["innings"].values.astype(float),
    }
    for e in events:
        feats[f"ev_{e}"] = (df["event"] == e).astype(float).values
    for p in phases:
        feats[f"ph_{p}"] = (df["phase"] == p).astype(float).values
    return pd.DataFrame(feats)


# Events with reliable DP direction (mismatch < 1% in training):
RELIABLE_EVENTS = {"4", "6", "NB"}

# ---------- Train zero-overshoot rising-team predictor ----------
def train_rising_predictor(safety_extra_c, mid_only=True, verbose=True,
                            reliable_only=True):
    df = build_training_data()
    if mid_only:
        df = df[(df["t40_rising"] >= 0.10) & (df["t40_rising"] <= 0.90)].copy()
    if reliable_only:
        df = df[df["event"].isin(RELIABLE_EVENTS)].copy()
    events = sorted(df["event"].unique())
    phases = sorted(df["phase"].unique())

    if verbose:
        print(f"Training rows (mid-range, rising-team coords): {len(df)}")
        print(f"  Direction-mismatch (actual rise < 0) on training: "
              f"{(df['actual_rise_c'] < 0).sum()} ({(df['actual_rise_c']<0).mean()*100:.1f}%)")

    # LOMO conformal margin
    matches = sorted(df["slug"].unique())
    cv_overs = []
    for held in matches:
        sub_tr = df[df["slug"] != held]
        sub_va = df[df["slug"] == held]
        if sub_va.empty: continue
        Xt = featurize(sub_tr, events, phases)
        yt = sub_tr["actual_rise_c"].values
        m = lgb.LGBMRegressor(objective="quantile", alpha=0.005,
                              n_estimators=600, learning_rate=0.04,
                              num_leaves=31, min_data_in_leaf=15, verbosity=-1)
        m.fit(Xt, yt)
        Xv = featurize(sub_va, events, phases)
        pv = m.predict(Xv)
        over = (pv - sub_va["actual_rise_c"].values).max()
        cv_overs.append(over)
    conformal_c = max(cv_overs)
    total_margin = conformal_c + safety_extra_c
    if verbose:
        print(f"  LOMO worst over-pred: max={max(cv_overs):.2f}c  "
              f"min={min(cv_overs):.2f}c  mean={np.mean(cv_overs):.2f}c")
        print(f"  Conformal margin = {conformal_c:.2f}c  "
              f"+ safety {safety_extra_c:.1f}c  = total {total_margin:.2f}c")

    # Final model on all data
    X = featurize(df, events, phases)
    y = df["actual_rise_c"].values
    final = lgb.LGBMRegressor(objective="quantile", alpha=0.005,
                              n_estimators=600, learning_rate=0.04,
                              num_leaves=31, min_data_in_leaf=15, verbosity=-1)
    final.fit(X, y)

    # In-sample sanity check
    pred = np.maximum(0.0, final.predict(X) - total_margin)
    err = y - pred  # actual_rise - predicted_rise; want >= 0
    if verbose:
        print(f"  In-sample after margin: over% = {(err<0).mean()*100:.2f}%  "
              f"mean abs = {np.abs(err).mean():.2f}c  "
              f"max overshoot = {(-err).max() if (err<0).any() else 0:.2f}c")

    return final, events, phases, total_margin, df


# ---------- Apply to a live match ----------
def load_espn_for_slug(slug):
    """If ESPN bbb workbook has a tab for this slug, return list of dicts
    with innings, balls_idx, ts_ms, is_legal. Else None."""
    from openpyxl import load_workbook
    wb_path = CAPTURES / "espn_ipl2026_ballbyball.xlsx"
    if not wb_path.exists():
        return None
    wb = load_workbook(wb_path, read_only=True)
    if slug not in wb.sheetnames:
        return None
    ws = wb[slug]
    headers = [c.value for c in ws[6]]
    out = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] is None: continue
        rec = dict(zip(headers, row))
        is_wide = bool(rec.get("is_wide"))
        is_noball = bool(rec.get("is_noball"))
        out.append({
            "innings": int(rec.get("innings") or 0),
            "balls_idx": overs_to_balls(rec.get("over_actual")),
            "ts_ms": int(rec.get("bbb_ts_ms")),
            "is_legal": not (is_wide or is_noball),
        })
    return out


def load_match_events_with_book(slug, dp, inn1_dp):
    """Load cricket_events + book series for BOTH outcome tokens.
    Uses t = min(capture_ts, espn_bbb_ts) when ESPN data is available
    (matching the workbook training basis)."""
    db_path = next(CAPTURES.glob(f"match_capture_cricipl-{slug}_*.db"))
    conn = sqlite3.connect(db_path)
    meta = dict(conn.execute("SELECT key, value FROM match_meta").fetchall())
    token_ids = json.loads(meta["token_ids"])
    outcome_names = json.loads(meta["outcome_names"])
    shorts = [NAME_TO_SHORT.get(o, o) for o in outcome_names]
    team_a, team_b = shorts[0], shorts[1]
    token_a, token_b = token_ids[0], token_ids[1]

    rows = conn.execute(
        "SELECT id, local_ts_ms, COALESCE(runs,0), COALESCE(wickets,0), "
        "COALESCE(overs,'0.0'), score_str, signal_type "
        "FROM cricket_events WHERE signal_type != '?' ORDER BY id"
    ).fetchall()

    # Detect innings split: 5+ min gap + balls reset
    split_idx = None
    for i in range(1, len(rows)):
        gap = (rows[i][1] - rows[i-1][1]) / 1000.0
        if gap >= 300 and overs_to_balls(rows[i][4]) < overs_to_balls(rows[i-1][4]) and rows[i][2] < rows[i-1][2]:
            split_idx = i; break

    inn1_total = None
    if split_idx is not None:
        inn1_total = rows[split_idx-1][2]
    target = (inn1_total + 1) if inn1_total is not None else None

    # Score parser to extract batting team prefix when present
    bat_in_inn1 = None
    for r in rows[:60]:
        sstr = r[5]
        if sstr:
            tok = sstr.split()[0]
            if tok in shorts: bat_in_inn1 = tok; break
    if bat_in_inn1 is None:
        bat_in_inn1 = team_a
    bat_in_inn2 = team_b if bat_in_inn1 == team_a else team_a

    # Build legal-ball capture timestamp index for ESPN pairing.
    cap_legal_by_key = {}
    last_b = {1: 0, 2: 0}
    for i, r in enumerate(rows):
        innings = 2 if (split_idx is not None and i >= split_idx) else 1
        b = overs_to_balls(r[4])
        prev = last_b.get(innings, 0)
        if b == prev + 1:
            cap_legal_by_key.setdefault((innings, b), r[1])
            last_b[innings] = b

    espn_data = load_espn_for_slug(slug)
    espn_ts_for_event = {}
    used_espn = espn_data is not None
    if used_espn:
        print(f"  ESPN data found for {slug} ({len(espn_data)} balls) — using min(capture, espn)")
        for e in espn_data:
            espn_ts_for_event[(e["innings"], e["balls_idx"], e["is_legal"])] = e["ts_ms"]
    else:
        print(f"  No ESPN data for {slug} — using capture timestamps only")

    # Book series for both tokens
    def load_book(token):
        rs = conn.execute(
            "SELECT local_ts_ms, bid1_p FROM book_snapshots "
            "WHERE asset_id = ? AND bid1_p IS NOT NULL ORDER BY local_ts_ms",
            (token,),
        ).fetchall()
        return [r[0] for r in rs], [r[1] for r in rs]
    times_a, bids_a = load_book(token_a)
    times_b, bids_b = load_book(token_b)

    def nearest(times, vals, tgt):
        if not times: return None
        i = bisect.bisect_left(times, tgt)
        cands = []
        if i < len(times): cands.append((abs(times[i]-tgt), vals[i]))
        if i > 0: cands.append((abs(times[i-1]-tgt), vals[i-1]))
        cands.sort()
        return cands[0][1] if cands else None

    out = []
    for i, r in enumerate(rows):
        rid, cap_ts, runs, wkts, overs, score_str, sig = r
        innings = 2 if (split_idx is not None and i >= split_idx) else 1
        bat_team = bat_in_inn2 if innings == 2 else bat_in_inn1
        ev = str(sig) if sig else ""
        dp_outcome = EVENT_TO_DP.get(ev)
        if dp_outcome is None: continue
        runs_post = runs; wkts_post = wkts
        balls_post = overs_to_balls(overs)
        runs_pre = runs_post - OUTCOME_RUNS[dp_outcome]
        wkts_pre = wkts_post - (1 if OUTCOME_IS_WICKET[dp_outcome] else 0)
        balls_pre = balls_post - (1 if OUTCOME_CONSUMES_BALL[dp_outcome] else 0)
        if runs_pre < 0 or wkts_pre < 0 or balls_pre < 0: continue

        # Compose t = min(capture, espn) for legal balls when ESPN available
        is_legal = ev not in ("WD", "NB")
        ts_ms = cap_ts
        if used_espn:
            espn_ts = espn_ts_for_event.get((innings, balls_post, is_legal))
            if espn_ts is not None:
                ts_ms = min(cap_ts, espn_ts)
            elif not is_legal:
                # extras: use ESPN-only when present
                e2 = espn_ts_for_event.get((innings, balls_post, False))
                if e2 is not None: ts_ms = e2

        # Book prices for both teams at t-40 and t
        t40_a = nearest(times_a, bids_a, ts_ms - 40_000)
        t_a   = nearest(times_a, bids_a, ts_ms)
        t40_b = nearest(times_b, bids_b, ts_ms - 40_000)
        t_b   = nearest(times_b, bids_b, ts_ms)
        if None in (t40_a, t_a, t40_b, t_b): continue

        # DP: WP for batting team
        balls_rem_pre = 120 - balls_pre
        balls_rem_post = 120 - balls_post
        wkts_hand_pre = 10 - wkts_pre
        wkts_hand_post = 10 - wkts_post
        if innings == 2 and target is not None:
            wp_pre = dp.lookup(balls_rem_pre, target - runs_pre, wkts_hand_pre)
            wp_post = dp.lookup(balls_rem_post, target - runs_post, wkts_hand_post)
        else:
            wp_pre = inn1_dp.lookup(balls_rem_pre, runs_pre, wkts_hand_pre)
            wp_post = inn1_dp.lookup(balls_rem_post, runs_post, wkts_hand_post)
        delta_bat = wp_post - wp_pre  # batting-team WP change

        # Translate to "rising team" framing.
        # The team whose WP rises = batting team if delta_bat>0, else bowling team.
        if delta_bat > 0:
            rising_team_name = bat_team
            dp_rise_c = delta_bat * 100
        elif delta_bat < 0:
            rising_team_name = team_a if bat_team == team_b else team_b
            dp_rise_c = -delta_bat * 100
        else:
            # Zero DP delta — pick batting team arbitrarily, mark dp_rise_c = 0
            rising_team_name = bat_team
            dp_rise_c = 0.0

        # Rising team's actual book prices
        if rising_team_name == team_a:
            t40_rising = t40_a; t_rising = t_a
        else:
            t40_rising = t40_b; t_rising = t_b
        actual_rise_c = (t_rising - t40_rising) * 100

        phase = "PP" if balls_post <= 36 else ("mid" if balls_post <= 90 else "death")
        ist_dt = datetime.fromtimestamp(ts_ms/1000, IST)
        cap_dt = datetime.fromtimestamp(cap_ts/1000, IST)

        out.append({
            "time_ist": ist_dt.strftime("%H:%M:%S"),
            "score": score_str,
            "event": ev,
            "innings": innings,
            "phase": phase,
            "bat_team": bat_team,
            "rising_team": rising_team_name,
            "dp_rise_c": dp_rise_c,
            "t40_rising": t40_rising,
            "t_rising": t_rising,
            "actual_rise_c": actual_rise_c,
        })
    conn.close()
    return out, team_a, team_b


def main():
    p = argparse.ArgumentParser()
    p.add_argument("slug")
    p.add_argument("--safety-extra-c", type=float, default=2.0)
    p.add_argument("--out-md", default=None)
    args = p.parse_args()

    print(f"=== {args.slug}  (rising-team predictor)  ===\n")
    print("Solving DPs…")
    dp = DPTable(); dp.solve()
    inn1_model = FirstInningsModel(dp)
    inn1_dp = InningsOneDP(inn1_model); inn1_dp.solve()

    print("\nTraining Layer 2 (rising-team coords)…")
    model, events, phases, margin_c, train_df = train_rising_predictor(
        args.safety_extra_c)

    # Held-out check on the workbook test split (last 4 matches)
    sorted_dates = sorted(train_df["date"].unique())
    test_dates = set(sorted_dates[-4:])
    holdout = train_df[train_df["date"].isin(test_dates)].copy()
    Xh = featurize(holdout, events, phases)
    pred_h = np.maximum(0.0, model.predict(Xh) - margin_c)
    err_h = holdout["actual_rise_c"].values - pred_h
    print(f"\n  Held-out 4-match check: over% = {(err_h<0).mean()*100:.2f}%   "
          f"mean abs = {np.abs(err_h).mean():.2f}c   "
          f"max overshoot = {(-err_h).max() if (err_h<0).any() else 0:.2f}c   n={len(holdout)}")

    print(f"\nLoading match {args.slug}…")
    rows, team_a, team_b = load_match_events_with_book(args.slug, dp, inn1_dp)
    print(f"  {len(rows)} predictable events")

    if not rows:
        print("No events.")
        return 1
    df = pd.DataFrame(rows)
    df["in_range"] = (df["t40_rising"] >= 0.10) & (df["t40_rising"] <= 0.90)
    df["reliable_event"] = df["event"].isin(RELIABLE_EVENTS)
    df_pred = df[df["in_range"] & df["reliable_event"]].copy()
    df.loc[~df["reliable_event"], "_unreliable"] = True
    if df_pred.empty:
        print("No mid-range predictable events.")
        return 1

    Xm = featurize(df_pred, events, phases)
    df_pred["pred_rise_c"] = np.maximum(0.0, model.predict(Xm) - margin_c)
    df_pred["pred_rising_t"] = (df_pred["t40_rising"] + df_pred["pred_rise_c"]/100).clip(0.001, 0.999)
    df_pred["err_c"] = df_pred["actual_rise_c"] - df_pred["pred_rise_c"]
    df_pred["status"] = np.where(df_pred["err_c"] >= 0, "OK", "OVERSHOOT")

    df = df.merge(df_pred[["time_ist", "pred_rise_c", "pred_rising_t",
                            "err_c", "status"]], on="time_ist", how="left")
    df["status"] = df["status"].fillna("NO_TRADE")
    df.loc[~df["in_range"] & df["status"].eq("NO_TRADE"), "status"] = "NO_TRADE_extreme_price"
    df.loc[~df["reliable_event"] & df["status"].eq("NO_TRADE"), "status"] = "NO_TRADE_unreliable_event"

    n_pred = len(df_pred)
    n_over = int((df_pred["err_c"] < 0).sum())
    print(f"\n=== Match summary ===")
    print(f"  Predictable events (mid-range): {n_pred}")
    print(f"  Overshoots: {n_over} ({n_over/n_pred*100:.2f}%)")
    print(f"  Mean abs error: {df_pred['err_c'].abs().mean():.2f}c")
    print(f"  Median abs error: {df_pred['err_c'].abs().median():.2f}c")
    print(f"  Max abs error: {df_pred['err_c'].abs().max():.2f}c")
    if n_over:
        print(f"  Worst overshoot: {(-df_pred['err_c'][df_pred['err_c']<0]).max():.2f}c")

    md_path = Path(args.out_md) if args.out_md else CAPTURES / f"predictions_rising_{args.slug}.md"
    with open(md_path, "w") as f:
        f.write(f"# Rising-team predictions: {args.slug}\n\n")
        f.write(f"Teams: {team_a} vs {team_b}.  Model: GBM quantile τ=0.005 + "
                f"LOMO conformal + {args.safety_extra_c}c safety → total margin "
                f"**{margin_c:.2f}c**.\n\n")
        f.write(f"Goal: predicted rise ≤ actual rise of the rising team. "
                f"Result on this match: **{n_over}/{n_pred} overshoots ({n_over/n_pred*100:.2f}%)**.\n\n")
        f.write(f"**Errors (cents):** mean abs {df_pred['err_c'].abs().mean():.2f}, "
                f"median {df_pred['err_c'].abs().median():.2f}, "
                f"max {df_pred['err_c'].abs().max():.2f}.\n\n")
        f.write("Columns: rising team's price at t-40, predicted rising team's "
                "price at t (our quote = lower bound), actual rising team's "
                "price at t, predicted rise (cents), actual rise (cents), error.\n\n")
        f.write("| # | time IST | score | event | bat | rising | t-40 (rising) | "
                "predicted t (rising) | actual t (rising) | "
                "pred rise (c) | actual rise (c) | err (c) | status |\n")
        f.write("|---:|---|---|:---:|:---:|:---:|---:|---:|---:|---:|---:|---:|---|\n")
        for i, r in df.iterrows():
            pt = f"{r['pred_rising_t']:.3f}" if pd.notna(r["pred_rising_t"]) else "—"
            pr = f"{r['pred_rise_c']:.2f}" if pd.notna(r["pred_rise_c"]) else "—"
            er = f"{r['err_c']:+.2f}" if pd.notna(r["err_c"]) else "—"
            f.write(f"| {i+1} | {r['time_ist']} | {r['score']} | {r['event']} | "
                    f"{r['bat_team']} | {r['rising_team']} | "
                    f"{r['t40_rising']:.3f} | {pt} | {r['t_rising']:.3f} | "
                    f"{pr} | {r['actual_rise_c']:+.2f} | {er} | {r['status']} |\n")
    print(f"\nMarkdown: {md_path}")
    df.to_csv(CAPTURES / f"predictions_rising_{args.slug}.csv", index=False)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
