#!/usr/bin/env python3
"""Cricket Event + Orderbook Snapshot Capture System.

Streams live cricket scores via Firebase SSE and Polymarket orderbook via WS.
Maintains a rolling 3-minute buffer of orderbook snapshots (one per 200ms).
When a special cricket event (4, 6, W) is detected:
  - Freezes the 3-minute pre-event buffer
  - Captures 2 more minutes of post-event snapshots
  - Dumps all 5 minutes to a new sheet in an Excel workbook

Usage:
    python event_book_capture.py --slug <polymarket-slug> --match <cricket-match-key>
    python event_book_capture.py --slug csk-vs-pbks-ipl --match a-rz--cricket--Og2031799084419014663

Requires CRICKET_API_KEY in .env.
"""

import argparse
import asyncio
import json
import os
import sys
import threading
import time
from collections import deque
from datetime import datetime, timezone, timedelta

import httpx
import websockets
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

# ── Constants ────────────────────────────────────────────────────────────────

WS_URL = "wss://ws-subscriptions-clob.polymarket.com/ws/market"
GAMMA_API = "https://gamma-api.polymarket.com"
IST = timezone(timedelta(hours=5, minutes=30))

BUFFER_SECONDS = 180       # 3 minutes rolling buffer
POST_EVENT_SECONDS = 120   # 2 minutes after event
SNAPSHOT_INTERVAL = 0.2    # 200ms

# Max snapshots in buffer: 180 / 0.2 = 900
MAX_BUFFER_SIZE = int(BUFFER_SECONDS / SNAPSHOT_INTERVAL)

# Colors
C_RESET = "\033[0m"
C_GREEN = "\033[92m"
C_RED = "\033[91m"
C_YELLOW = "\033[93m"
C_CYAN = "\033[96m"
C_BOLD = "\033[1m"
C_DIM = "\033[90m"
C_MAGENTA = "\033[95m"

SPECIAL_EVENTS = {"4", "6", "W", "w"}


def ist_now():
    return datetime.now(IST).strftime("%H:%M:%S.%f")[:-3]


# ── Polymarket Market Resolution ────────────────────────────────────────────

def fetch_market(slug: str) -> dict:
    import requests
    resp = requests.get(f"{GAMMA_API}/markets", params={"slug": slug}, timeout=15)
    resp.raise_for_status()
    markets = resp.json()
    if not markets:
        print(f"{C_RED}No market found for slug '{slug}'{C_RESET}")
        sys.exit(1)
    return markets[0]


def parse_tokens(market: dict) -> tuple[list[str], list[str]]:
    tokens = market.get("clobTokenIds", "")
    outcomes = market.get("outcomes", "[]")
    if isinstance(tokens, str):
        tokens = json.loads(tokens) if tokens else []
    if isinstance(outcomes, str):
        outcomes = json.loads(outcomes) if outcomes else []
    return tokens, outcomes


# ── Orderbook Snapshot ──────────────────────────────────────────────────────

class BookSnapshot:
    """A single point-in-time L2 orderbook snapshot (top 5 levels)."""
    __slots__ = ("ts", "ts_ns", "ist", "books")

    def __init__(self, books: dict[str, dict[str, dict[float, float]]],
                 outcome_names: list[str], token_ids: list[str]):
        self.ts = time.time()
        self.ts_ns = time.time_ns()
        self.ist = ist_now()
        # Freeze top 5 levels for each token
        self.books = {}
        for tid, oname in zip(token_ids, outcome_names):
            raw = books.get(tid, {"bids": {}, "asks": {}})
            top_bids = sorted(raw["bids"].items(), key=lambda x: -x[0])[:5]
            top_asks = sorted(raw["asks"].items(), key=lambda x: x[0])[:5]
            self.books[oname] = {
                "bids": [(p, s) for p, s in top_bids],
                "asks": [(p, s) for p, s in top_asks],
            }


# ── Shared State ────────────────────────────────────────────────────────────

class CaptureState:
    def __init__(self, token_ids: list[str], outcome_names: list[str]):
        self.token_ids = token_ids
        self.outcome_names = outcome_names

        # Live L2 book: {token_id: {"bids": {price: size}, "asks": {price: size}}}
        self.books: dict[str, dict[str, dict[float, float]]] = {
            tid: {"bids": {}, "asks": {}} for tid in token_ids
        }
        self.book_lock = threading.Lock()

        # Rolling snapshot buffer (thread-safe deque)
        self.snapshot_buffer: deque[BookSnapshot] = deque(maxlen=MAX_BUFFER_SIZE)

        # Cricket state
        self.last_score_str = ""
        self.last_wickets = 0
        self.last_runs = 0
        self.last_overs = ""
        self.cricket_connected = False

        # Event capture tracking
        self.event_count = 0
        self.active_captures = 0  # how many post-event captures running

        # WS connected
        self.ws_connected = False

    def take_snapshot(self) -> BookSnapshot:
        with self.book_lock:
            return BookSnapshot(self.books, self.outcome_names, self.token_ids)

    def apply_book_snapshot(self, asset_id: str, bids: list, asks: list):
        if asset_id not in self.books:
            return
        with self.book_lock:
            self.books[asset_id]["bids"] = {}
            self.books[asset_id]["asks"] = {}
            for b in bids:
                p, s = float(b.get("price", 0)), float(b.get("size", 0))
                if s > 0:
                    self.books[asset_id]["bids"][p] = s
            for a in asks:
                p, s = float(a.get("price", 0)), float(a.get("size", 0))
                if s > 0:
                    self.books[asset_id]["asks"][p] = s

    def apply_price_change(self, asset_id: str, changes: list):
        if asset_id not in self.books:
            return
        with self.book_lock:
            book = self.books[asset_id]
            for change in changes:
                price = float(change.get("price", 0))
                new_size = float(change.get("size", 0))
                side_str = change.get("side", "")
                side_key = "bids" if side_str == "BUY" else "asks"
                if new_size <= 0:
                    book[side_key].pop(price, None)
                else:
                    book[side_key][price] = new_size


# ── Polymarket WebSocket ────────────────────────────────────────────────────

async def ws_orderbook(state: CaptureState):
    """Connect to Polymarket WS and maintain live orderbook."""
    while True:
        try:
            async with websockets.connect(WS_URL, ping_interval=None) as ws:
                sub = json.dumps({"assets_ids": state.token_ids, "type": "market"})
                await ws.send(sub)
                state.ws_connected = True
                print(f"{C_DIM}{ist_now()}{C_RESET} {C_GREEN}WS connected{C_RESET} — subscribed to {len(state.token_ids)} tokens")

                async def keepalive():
                    while True:
                        await asyncio.sleep(10)
                        try:
                            await ws.send("PING")
                        except Exception:
                            return

                ping_task = asyncio.create_task(keepalive())
                try:
                    async for raw in ws:
                        if raw == "PONG":
                            continue
                        _process_ws(state, raw)
                finally:
                    ping_task.cancel()
                    state.ws_connected = False

        except (websockets.ConnectionClosed, ConnectionError, OSError) as e:
            state.ws_connected = False
            print(f"{C_DIM}{ist_now()}{C_RESET} {C_RED}WS disconnected:{C_RESET} {e}")
        except Exception as e:
            state.ws_connected = False
            print(f"{C_DIM}{ist_now()}{C_RESET} {C_RED}WS error:{C_RESET} {e}")

        await asyncio.sleep(2)


def _process_ws(state: CaptureState, text: str):
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return

    events = data if isinstance(data, list) else [data]
    for event in events:
        if not isinstance(event, dict):
            continue

        if "price_changes" in event:
            for change in event["price_changes"]:
                aid = change.get("asset_id", "")
                if aid in state.books:
                    state.apply_price_change(aid, [change])
            continue

        asset_id = event.get("asset_id")
        event_type = event.get("event_type") or event.get("type")
        if not asset_id or asset_id not in state.books:
            continue

        if event_type == "book":
            state.apply_book_snapshot(asset_id, event.get("bids", []), event.get("asks", []))
        elif event_type == "price_change":
            changes = event.get("changes", [])
            if changes:
                state.apply_price_change(asset_id, changes)
            else:
                bid_changes = [{"price": b.get("price", 0), "size": b.get("size", 0), "side": "BUY"}
                               for b in event.get("bids", [])]
                ask_changes = [{"price": a.get("price", 0), "size": a.get("size", 0), "side": "SELL"}
                               for a in event.get("asks", [])]
                state.apply_price_change(asset_id, bid_changes + ask_changes)


# ── Snapshot Ticker (200ms) ─────────────────────────────────────────────────

async def snapshot_ticker(state: CaptureState):
    """Take an orderbook snapshot every 200ms into the rolling buffer."""
    while True:
        snap = state.take_snapshot()
        state.snapshot_buffer.append(snap)
        await asyncio.sleep(SNAPSHOT_INTERVAL)


# ── Cricket SSE Stream ──────────────────────────────────────────────────────

async def cricket_sse(state: CaptureState, match_key: str, wb: Workbook,
                      output_path: str, loop: asyncio.AbstractEventLoop):
    """Stream live cricket score via Firebase SSE. Detect 4/6/W events."""
    base = os.getenv("CRICKET_API_KEY", "").rstrip("/")
    if not base:
        print(f"{C_RED}CRICKET_API_KEY not set in .env{C_RESET}")
        return

    score_url = f"{base}/recent-matches/{match_key}/play/live/score.json"

    print(f"{C_DIM}{ist_now()}{C_RESET} {C_CYAN}Cricket SSE connecting...{C_RESET}")
    print(f"{C_DIM}         URL: {score_url}{C_RESET}")

    while True:
        try:
            async with httpx.AsyncClient(timeout=None) as client:
                async with client.stream("GET", score_url,
                                         headers={"Accept": "text/event-stream"}) as resp:
                    resp.raise_for_status()
                    state.cricket_connected = True
                    print(f"{C_DIM}{ist_now()}{C_RESET} {C_GREEN}Cricket SSE connected{C_RESET}")

                    event_type = None
                    data_buf = []

                    async for line in resp.aiter_lines():
                        if line.startswith("event:"):
                            event_type = line[6:].strip()
                            continue
                        if line.startswith("data:"):
                            data_buf.append(line[5:].strip())
                            continue

                        if line == "" and data_buf:
                            raw = "\n".join(data_buf)
                            data_buf = []

                            if raw == "null" or event_type == "keep-alive":
                                continue

                            try:
                                payload = json.loads(raw)
                            except json.JSONDecodeError:
                                continue

                            _handle_cricket_update(
                                state, payload, event_type, wb, output_path, loop
                            )

        except Exception as e:
            state.cricket_connected = False
            print(f"{C_DIM}{ist_now()}{C_RESET} {C_RED}Cricket SSE error:{C_RESET} {e}")

        await asyncio.sleep(2)


def _handle_cricket_update(state: CaptureState, payload, event_type: str | None,
                           wb: Workbook, output_path: str,
                           loop: asyncio.AbstractEventLoop):
    """Process a cricket SSE update. Detect special events."""
    if not isinstance(payload, dict):
        return

    # Firebase SSE sends {path, data} for put/patch
    data = payload.get("data", payload)
    if isinstance(data, dict) and "path" in data:
        data = data.get("data", data)
    if not isinstance(data, dict):
        return

    runs = data.get("runs", state.last_runs)
    wickets = data.get("wickets", state.last_wickets)
    overs = data.get("overs", "")
    title = data.get("title", "")
    run_rate = data.get("run_rate", "")

    if isinstance(overs, list) and len(overs) == 2:
        overs_str = f"{overs[0]}.{overs[1]}"
    else:
        overs_str = str(overs)

    score_str = f"{runs}/{wickets} ({overs_str})"

    if score_str == state.last_score_str:
        return

    # Detect special events
    run_diff = runs - state.last_runs
    wicket_diff = wickets - state.last_wickets
    special = None

    if wicket_diff > 0:
        special = "W"
    elif run_diff == 6:
        special = "6"
    elif run_diff == 4:
        special = "4"

    # Update state
    prev_score = state.last_score_str
    state.last_score_str = score_str
    state.last_runs = runs
    state.last_wickets = wickets
    state.last_overs = overs_str

    # Print
    if special:
        color = C_RED if special == "W" else C_GREEN
        print(f"\n{C_BOLD}{color}{'='*60}{C_RESET}")
        print(f"{C_BOLD}{color}  EVENT: {special}  |  {score_str}  |  RR: {run_rate}{C_RESET}")
        print(f"{C_BOLD}{color}{'='*60}{C_RESET}\n")

        # Trigger capture in a background thread
        state.event_count += 1
        event_num = state.event_count
        event_ts = time.time()
        event_ist = ist_now()

        # Freeze pre-event buffer
        pre_snaps = list(state.snapshot_buffer)

        state.active_captures += 1
        threading.Thread(
            target=_capture_event_thread,
            args=(state, wb, output_path, pre_snaps, special, score_str,
                  event_num, event_ts, event_ist, loop),
            daemon=True,
        ).start()
    else:
        print(f"{C_DIM}{ist_now()}{C_RESET} {C_CYAN}score{C_RESET}  {score_str}  RR:{run_rate}")


# ── Event Capture Thread ────────────────────────────────────────────────────

def _capture_event_thread(state: CaptureState, wb: Workbook, output_path: str,
                          pre_snaps: list[BookSnapshot], event_type: str,
                          score_str: str, event_num: int, event_ts: float,
                          event_ist: str, loop: asyncio.AbstractEventLoop):
    """Runs in a separate thread. Captures 2 more minutes of snapshots then dumps to Excel."""
    print(f"{C_DIM}{ist_now()}{C_RESET} {C_MAGENTA}capture #{event_num}{C_RESET} "
          f"started — collecting {POST_EVENT_SECONDS}s post-event snapshots...")

    post_snaps = []
    end_time = time.time() + POST_EVENT_SECONDS

    while time.time() < end_time:
        snap = state.take_snapshot()
        post_snaps.append(snap)
        time.sleep(SNAPSHOT_INTERVAL)

    all_snaps = pre_snaps + post_snaps
    total = len(all_snaps)

    print(f"{C_DIM}{ist_now()}{C_RESET} {C_MAGENTA}capture #{event_num}{C_RESET} "
          f"done — {len(pre_snaps)} pre + {len(post_snaps)} post = {total} snapshots")

    # Write to Excel sheet
    sheet_name = f"E{event_num}_{event_type}_{event_ist.replace(':', '').replace('.', '_')}"
    sheet_name = sheet_name[:31]  # Excel limit

    _write_event_sheet(wb, sheet_name, all_snaps, event_type, score_str,
                       event_ist, event_ts, state.outcome_names)

    # Save workbook
    try:
        wb.save(output_path)
        print(f"{C_DIM}{ist_now()}{C_RESET} {C_GREEN}saved{C_RESET} sheet '{sheet_name}' → {output_path}")
    except Exception as e:
        print(f"{C_DIM}{ist_now()}{C_RESET} {C_RED}save error:{C_RESET} {e}")

    state.active_captures -= 1


def _write_event_sheet(wb: Workbook, sheet_name: str, snaps: list[BookSnapshot],
                       event_type: str, score_str: str, event_ist: str,
                       event_ts: float, outcome_names: list[str]):
    """Write snapshot data to a new Excel sheet."""
    ws = wb.create_sheet(title=sheet_name)

    # Header styles
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    event_fill = PatternFill("solid", fgColor="FEF3C7")
    pre_fill = PatternFill("solid", fgColor="EFF6FF")
    post_fill = PatternFill("solid", fgColor="F0FDF4")

    # Metadata rows
    ws.append([f"Event: {event_type}", f"Score: {score_str}", f"Time: {event_ist}",
               f"Snapshots: {len(snaps)}"])
    ws.append([])

    # Build header: IST | ms_offset | outcome1_bid1..bid5 | outcome1_ask1..ask5 | outcome2_...
    headers = ["IST", "ms_from_event"]
    for oname in outcome_names:
        for i in range(1, 6):
            headers.append(f"{oname}_bid{i}_price")
            headers.append(f"{oname}_bid{i}_size")
        for i in range(1, 6):
            headers.append(f"{oname}_ask{i}_price")
            headers.append(f"{oname}_ask{i}_size")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill

    # Data rows
    for snap in snaps:
        ms_offset = int((snap.ts - event_ts) * 1000)
        row = [snap.ist, ms_offset]

        for oname in outcome_names:
            book = snap.books.get(oname, {"bids": [], "asks": []})
            # Bids (top 5)
            for i in range(5):
                if i < len(book["bids"]):
                    row.extend([book["bids"][i][0], book["bids"][i][1]])
                else:
                    row.extend(["", ""])
            # Asks (top 5)
            for i in range(5):
                if i < len(book["asks"]):
                    row.extend([book["asks"][i][0], book["asks"][i][1]])
                else:
                    row.extend(["", ""])

        ws.append(row)

        # Color pre vs post event rows
        row_num = ws.max_row
        fill = pre_fill if ms_offset < 0 else post_fill
        if abs(ms_offset) < 500:
            fill = event_fill
        for col_idx in range(1, 3):
            ws.cell(row=row_num, column=col_idx).fill = fill

    # Auto-width for first 2 columns
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14


# ── Status Printer ──────────────────────────────────────────────────────────

async def status_printer(state: CaptureState):
    """Print periodic status so you know things are alive."""
    while True:
        await asyncio.sleep(10)
        buf_len = len(state.snapshot_buffer)
        buf_secs = buf_len * SNAPSHOT_INTERVAL
        ws_status = f"{C_GREEN}connected{C_RESET}" if state.ws_connected else f"{C_RED}disconnected{C_RESET}"
        sse_status = f"{C_GREEN}connected{C_RESET}" if state.cricket_connected else f"{C_RED}disconnected{C_RESET}"
        active = f" | {C_MAGENTA}{state.active_captures} captures running{C_RESET}" if state.active_captures else ""
        print(f"{C_DIM}{ist_now()}{C_RESET} {C_DIM}status{C_RESET}  "
              f"WS:{ws_status} SSE:{sse_status} "
              f"buf:{buf_len} ({buf_secs:.0f}s) events:{state.event_count}{active}")


# ── Main ────────────────────────────────────────────────────────────────────

async def run(args):
    # Resolve Polymarket market
    print(f"{C_BOLD}Resolving market '{args.slug}'...{C_RESET}")
    market = fetch_market(args.slug)
    question = market.get("question", args.slug)
    token_ids, outcome_names = parse_tokens(market)

    if not token_ids:
        print(f"{C_RED}No token IDs found{C_RESET}")
        sys.exit(1)

    print(f"Market:   {question}")
    print(f"Outcomes: {', '.join(outcome_names)}")
    print(f"Tokens:   {len(token_ids)}")
    print(f"Match:    {args.match}")
    print()

    state = CaptureState(token_ids, outcome_names)

    # Prepare Excel workbook
    os.makedirs("captures", exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join("captures", f"{ts}_event_book_{args.slug[:30]}.xlsx")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print(f"Output:   {output_path}")
    print(f"Buffer:   {BUFFER_SECONDS}s rolling ({MAX_BUFFER_SIZE} snapshots)")
    print(f"Capture:  {BUFFER_SECONDS}s pre + {POST_EVENT_SECONDS}s post = "
          f"{BUFFER_SECONDS + POST_EVENT_SECONDS}s per event")
    print(f"Interval: {SNAPSHOT_INTERVAL * 1000:.0f}ms")
    print(f"{'='*60}")
    print(f"Waiting for special events (4, 6, W)...\n")

    loop = asyncio.get_event_loop()

    await asyncio.gather(
        ws_orderbook(state),
        snapshot_ticker(state),
        cricket_sse(state, args.match, wb, output_path, loop),
        status_printer(state),
    )


def main():
    parser = argparse.ArgumentParser(description="Cricket Event + Orderbook Snapshot Capture")
    parser.add_argument("--slug", required=True, help="Polymarket market slug")
    parser.add_argument("--match", required=True, help="Cricket match key from Firebase")
    args = parser.parse_args()

    try:
        asyncio.run(run(args))
    except KeyboardInterrupt:
        print(f"\n{C_BOLD}Stopped.{C_RESET}")


if __name__ == "__main__":
    main()
