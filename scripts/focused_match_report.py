#!/usr/bin/env python3
"""focused_match_report.py — 5-sheet focused analysis from a match_capture DB.

Sheets:
  1_special_events   W/4/6 events, IST times
  2_event_books      book snapshots around each event (stacked tables)
  3_trades_clob      all CLOB trades
  4_trades_chain     all subgraph chain fills
  5_wallets          wallets with >=100 trades, >=10 BUYs and >=10 SELLs,
                     with PnL, volume, bps, taker vs maker breakdown

Usage:
  python3 focused_match_report.py <path_to_match_capture.db>
"""

import sqlite3, json, datetime, argparse, sys, urllib.request
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))


def ms_to_ist(ms):
    if ms is None: return ''
    return datetime.datetime.fromtimestamp(ms / 1000, tz=IST).strftime('%H:%M:%S.%f')[:-3]


def s_to_ist(s):
    if s is None: return ''
    return datetime.datetime.fromtimestamp(s, tz=IST).strftime('%H:%M:%S')


def fetch_settlement(slug):
    try:
        req = urllib.request.Request(
            f'https://gamma-api.polymarket.com/events?slug={slug}',
            headers={'User-Agent': 'Mozilla/5.0'},
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
        for m in data[0].get('markets', []):
            if m.get('slug') == slug:
                prices = m.get('outcomePrices')
                if isinstance(prices, str): prices = json.loads(prices)
                return float(prices[0]), float(prices[1])
    except Exception as e:
        print(f'warning: gamma api failed ({e}), falling back to 0.5/0.5', file=sys.stderr)
    return 0.5, 0.5


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('db')
    ap.add_argument('--out-dir', default='/Users/sobhagyaxd/DeepWork/totem-mm/captures')
    ap.add_argument('--min-trades', type=int, default=100)
    ap.add_argument('--min-side', type=int, default=10)
    ap.add_argument('--settle', help='Override settlement, e.g. "1,0" for T1-wins or "0,1" for T2-wins')
    args = ap.parse_args()

    con = sqlite3.connect(args.db)
    cur = con.cursor()
    meta = dict(cur.execute('SELECT key, value FROM match_meta').fetchall())
    slug = meta['slug']
    tokens = json.loads(meta['token_ids'])
    names = json.loads(meta['outcome_names'])
    T1, T2 = tokens[0], tokens[1]
    N1, N2 = names[0], names[1]
    N1S = (N1.split()[0] if ' ' in N1 else N1)[:3].upper()
    N2S = (N2.split()[0] if ' ' in N2 else N2)[:3].upper()

    if args.settle:
        parts = args.settle.split(',')
        t1_settle, t2_settle = float(parts[0]), float(parts[1])
    else:
        t1_settle, t2_settle = fetch_settlement(slug)
    print(f'Match: {N1} vs {N2}')
    print(f'Settle: {N1S}=${t1_settle}  {N2S}=${t2_settle}')

    start_ms = int(datetime.datetime.fromisoformat(meta['start_time']).timestamp() * 1000)

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    hdr_fill = PatternFill('solid', fgColor='D9E1F2')
    sub_fill = PatternFill('solid', fgColor='FFF2CC')

    # ===================== SHEET 1: special_events =====================
    print('Sheet 1: special_events ...')
    ws1 = wb.create_sheet('1_special_events')
    headers = ['event_id', 'ist_time', 'type', 'runs', 'wickets', 'overs', 'score', 'innings', 't_from_start_min']
    for i, h in enumerate(headers):
        c = ws1.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill

    events = cur.execute("""
        SELECT id, local_ts_ms, signal_type, runs, wickets, overs, score_str, innings
        FROM cricket_events WHERE signal_type IN ('W','4','6') ORDER BY local_ts_ms
    """).fetchall()

    for i, ev in enumerate(events, 2):
        mins = round((ev[1] - start_ms) / 60000, 1) if ev[1] else None
        row = [ev[0], ms_to_ist(ev[1]), ev[2], ev[3], ev[4], ev[5], ev[6], ev[7], mins]
        for j, v in enumerate(row):
            ws1.cell(row=i, column=j+1, value=v)
    ws1.freeze_panes = 'A2'
    ws1.column_dimensions['B'].width = 16

    # ===================== SHEET 2: event_books =====================
    print('Sheet 2: event_books (loading snapshots) ...')
    snaps = {T1: [], T2: []}
    for row in cur.execute("""
        SELECT asset_id, local_ts_ms,
               bid3_p, bid3_s, bid2_p, bid2_s, bid1_p, bid1_s,
               ask1_p, ask1_s, ask2_p, ask2_s, ask3_p, ask3_s,
               mid_price, spread, total_bid_depth, total_ask_depth
        FROM book_snapshots WHERE asset_id IN (?,?) ORDER BY local_ts_ms
    """, (T1, T2)):
        snaps[row[0]].append(row[1:])
    print(f'  T1 snaps={len(snaps[T1])}  T2 snaps={len(snaps[T2])}')

    def snap_at(asset, ts):
        arr = snaps[asset]
        lo, hi = 0, len(arr)
        while lo < hi:
            m = (lo+hi)//2
            if arr[m][0] <= ts: lo = m+1
            else: hi = m
        return arr[lo-1] if lo > 0 else None

    OFFSETS = [-120_000, -60_000, -30_000, -15_000, -10_000, -5_000, -3_000, -1_000,
               0, 1_000, 3_000, 5_000, 10_000, 15_000, 30_000, 60_000]

    ws2 = wb.create_sheet('2_event_books')
    row_idx = 1
    snap_headers = ['offset', 'token',
                    'bid3', 'b3sz', 'bid2', 'b2sz', 'bid1', 'b1sz',
                    'ask1', 'a1sz', 'ask2', 'a2sz', 'ask3', 'a3sz',
                    'mid', 'spread', 'bid_depth', 'ask_depth']

    for ev in events:
        ev_id, ev_ts, stype = ev[0], ev[1], ev[2]
        title = f'Event #{ev_id}  type={stype}  time={ms_to_ist(ev_ts)}  score={ev[6]}  overs={ev[5]}'
        c = ws2.cell(row=row_idx, column=1, value=title)
        c.font = Font(bold=True, size=11); c.fill = sub_fill
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(snap_headers))
        row_idx += 1

        for j, h in enumerate(snap_headers):
            c = ws2.cell(row=row_idx, column=j+1, value=h); c.font = bold; c.fill = hdr_fill
        row_idx += 1

        for off in OFFSETS:
            target = ev_ts + off
            for tok, lbl in [(T1, N1S), (T2, N2S)]:
                s = snap_at(tok, target)
                if s is None: continue
                # s: ts, b3p, b3s, b2p, b2s, b1p, b1s, a1p, a1s, a2p, a2s, a3p, a3s, mid, spread, tot_b, tot_a
                vals = [f'{off//1000:+d}s', lbl,
                        s[1], s[2], s[3], s[4], s[5], s[6],
                        s[7], s[8], s[9], s[10], s[11], s[12],
                        round(s[13], 4) if s[13] else None,
                        round(s[14], 4) if s[14] else None,
                        round(s[15], 1) if s[15] else None,
                        round(s[16], 1) if s[16] else None]
                for j, v in enumerate(vals):
                    ws2.cell(row=row_idx, column=j+1, value=v)
                row_idx += 1
        row_idx += 1  # blank between events

    for letter in 'ABCDEFGHIJKLMNOPQR':
        ws2.column_dimensions[letter].width = 9
    ws2.column_dimensions['A'].width = 10

    # ===================== SHEET 3: trades_clob =====================
    print('Sheet 3: trades_clob ...')
    ws3 = wb.create_sheet('3_trades_clob')
    headers = ['id', 'ist_time', 'outcome', 'side', 'price', 'size', 'notional_usdc',
               'fee_rate_bps', 'taker_wallet', 'tx_hash']
    for i, h in enumerate(headers):
        c = ws3.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill

    for i, r in enumerate(cur.execute("""
        SELECT id, clob_ts_ms, asset_id, side, price, size, notional_usdc, fee_rate_bps, taker_wallet, transaction_hash
        FROM trades ORDER BY clob_ts_ms, id
    """), 2):
        outcome = N1S if r[2] == T1 else N2S
        vals = [r[0], ms_to_ist(r[1]), outcome, r[3],
                round(r[4], 4) if r[4] else None,
                round(r[5], 4) if r[5] else None,
                round(r[6], 4) if r[6] else None,
                r[7],
                (r[8][:14]+'...') if r[8] else '',
                (r[9][:14]+'...') if r[9] else '']
        for j, v in enumerate(vals):
            ws3.cell(row=i, column=j+1, value=v)
    ws3.freeze_panes = 'A2'
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['I'].width = 20
    ws3.column_dimensions['J'].width = 20
    print(f'  {ws3.max_row - 1} rows')

    # ===================== SHEET 4: trades_chain =====================
    print('Sheet 4: trades_chain ...')
    ws4 = wb.create_sheet('4_trades_chain')
    headers = ['id', 'ist_time', 'maker', 'taker', 'direction', 'maker_amt', 'taker_amt',
               'derived_price', 'derived_size', 'fee', 'tx_hash']
    for i, h in enumerate(headers):
        c = ws4.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill

    chain_rows = cur.execute("""
        SELECT id, chain_timestamp, maker, taker, maker_asset_id, taker_asset_id,
               maker_amount, taker_amount, fee, transaction_hash
        FROM chain_fills ORDER BY chain_timestamp, id
    """).fetchall()

    for i, r in enumerate(chain_rows, 2):
        _id, ts, m, t, ma_id, ta_id, ma_amt, ta_amt, fee, tx = r
        # Derive direction from maker's perspective
        ma_id_s = str(ma_id) if ma_id is not None else ''
        ta_id_s = str(ta_id) if ta_id is not None else ''
        if ma_id_s == '0' and ta_id_s == T1:
            direction = f'USDC->{N1S} (mk BUY)'; price = (ma_amt/ta_amt) if ta_amt else 0; size = ta_amt or 0
        elif ma_id_s == '0' and ta_id_s == T2:
            direction = f'USDC->{N2S} (mk BUY)'; price = (ma_amt/ta_amt) if ta_amt else 0; size = ta_amt or 0
        elif ma_id_s == T1 and ta_id_s == '0':
            direction = f'{N1S}->USDC (mk SELL)'; price = (ta_amt/ma_amt) if ma_amt else 0; size = ma_amt or 0
        elif ma_id_s == T2 and ta_id_s == '0':
            direction = f'{N2S}->USDC (mk SELL)'; price = (ta_amt/ma_amt) if ma_amt else 0; size = ma_amt or 0
        else:
            direction = '?'; price = 0; size = 0
        vals = [_id, s_to_ist(ts),
                (m[:14]+'...') if m else '',
                (t[:14]+'...') if t else '',
                direction,
                round(ma_amt, 4) if ma_amt else 0,
                round(ta_amt, 4) if ta_amt else 0,
                round(price, 4),
                round(size, 4),
                round(fee, 4) if fee else 0,
                (tx[:14]+'...') if tx else '']
        for j, v in enumerate(vals):
            ws4.cell(row=i, column=j+1, value=v)
    ws4.freeze_panes = 'A2'
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 22
    print(f'  {ws4.max_row - 1} rows')

    # ===================== SHEET 5: wallets (with PnL) =====================
    print('Sheet 5: wallets (computing PnL) ...')
    # Per-wallet record list. Each record = dict with role, side, token, size, usdc, fee.
    wrec = defaultdict(list)

    # chain_fills — both maker and taker rows
    for r in chain_rows:
        _id, ts, m, t, ma_id, ta_id, ma_amt, ta_amt, fee, tx = r
        ma_id_s = str(ma_id) if ma_id is not None else ''
        ta_id_s = str(ta_id) if ta_id is not None else ''

        # Determine which side is which
        if ma_id_s == '0':
            # maker gave USDC → maker BUYS a token, taker SELLS it
            token = T1 if ta_id_s == T1 else (T2 if ta_id_s == T2 else None)
            if token is None: continue
            size = ta_amt or 0
            usdc = ma_amt or 0
            if m: wrec[m].append({'role':'maker','side':'BUY','token':token,'size':size,'usdc':usdc,'fee':0,'ts':ts,'tx':tx})
            if t: wrec[t].append({'role':'taker','side':'SELL','token':token,'size':size,'usdc':usdc,'fee':(fee or 0),'ts':ts,'tx':tx})
        elif ma_id_s in (T1, T2):
            # maker gave token → maker SELLS, taker BUYS
            token = ma_id_s
            size = ma_amt or 0
            usdc = ta_amt or 0
            if m: wrec[m].append({'role':'maker','side':'SELL','token':token,'size':size,'usdc':usdc,'fee':0,'ts':ts,'tx':tx})
            if t: wrec[t].append({'role':'taker','side':'BUY','token':token,'size':size,'usdc':usdc,'fee':(fee or 0),'ts':ts,'tx':tx})

    # Also pull from CLOB trades table for taker rows NOT already in chain_fills
    seen_tx = set()
    for r in chain_rows:
        if r[9]: seen_tx.add(r[9])

    for r in cur.execute("""
        SELECT clob_ts_ms, asset_id, side, price, size, notional_usdc, fee_rate_bps, taker_wallet, transaction_hash
        FROM trades WHERE taker_wallet IS NOT NULL
    """):
        ts, asset, side, p, sz, usdc, frb, w, tx = r
        if tx in seen_tx: continue  # already have it from chain_fills
        sz = sz or 0
        usdc = usdc or 0
        # Estimate fee via sports curve (0.03 * p * (1-p) * size)
        est_fee = 0.03 * (p or 0) * (1 - (p or 0)) * sz
        wrec[w].append({'role':'taker','side':side,'token':asset,'size':sz,'usdc':usdc,'fee':est_fee,'ts':ts/1000,'tx':tx})

    # Filter wallets
    def count_sides(recs):
        buys = sum(1 for r in recs if r['side'] == 'BUY')
        sells = sum(1 for r in recs if r['side'] == 'SELL')
        return buys, sells

    qualified = {}
    for w, recs in wrec.items():
        if len(recs) < args.min_trades: continue
        buys, sells = count_sides(recs)
        if buys < args.min_side or sells < args.min_side: continue
        qualified[w] = recs
    print(f'  qualified wallets: {len(qualified)}')

    def compute(recs):
        t1, t2, cash = 0.0, 0.0, 0.0
        t_cash, m_cash = 0.0, 0.0
        t_t1, t_t2 = 0.0, 0.0
        m_t1, m_t2 = 0.0, 0.0
        t_fees = 0.0
        t_vol = 0.0; m_vol = 0.0
        t_cnt = 0; m_cnt = 0
        buys = 0; sells = 0
        for r in recs:
            tok = r['token']; side = r['side']; sz = r['size']; us = r['usdc']; fee = r['fee']
            if side == 'BUY':
                buys += 1
                if r['role'] == 'taker':
                    t_cash -= us
                    if tok == T1: t_t1 += sz
                    else: t_t2 += sz
                    t_fees += fee
                    t_vol += us
                    t_cnt += 1
                else:
                    m_cash -= us
                    if tok == T1: m_t1 += sz
                    else: m_t2 += sz
                    m_vol += us
                    m_cnt += 1
            else:  # SELL
                sells += 1
                if r['role'] == 'taker':
                    t_cash += us
                    if tok == T1: t_t1 -= sz
                    else: t_t2 -= sz
                    t_fees += fee
                    t_vol += us
                    t_cnt += 1
                else:
                    m_cash += us
                    if tok == T1: m_t1 -= sz
                    else: m_t2 -= sz
                    m_vol += us
                    m_cnt += 1

        # Settle
        taker_pnl = t_cash + t_t1*t1_settle + t_t2*t2_settle - t_fees
        maker_pnl = m_cash + m_t1*t1_settle + m_t2*t2_settle
        total_pnl = taker_pnl + maker_pnl
        total_vol = t_vol + m_vol
        return {
            'n': len(recs), 'taker_cnt':t_cnt, 'maker_cnt':m_cnt,
            'buys':buys, 'sells':sells,
            'taker_vol':t_vol, 'maker_vol':m_vol, 'total_vol':total_vol,
            'net_t1': t_t1 + m_t1, 'net_t2': t_t2 + m_t2,
            'taker_fees': t_fees,
            'taker_pnl': taker_pnl, 'maker_pnl': maker_pnl, 'total_pnl': total_pnl,
            'taker_bps': (taker_pnl / t_vol * 10000) if t_vol else 0,
            'maker_bps': (maker_pnl / m_vol * 10000) if m_vol else 0,
            'total_bps': (total_pnl / total_vol * 10000) if total_vol else 0,
        }

    ws5 = wb.create_sheet('5_wallets')
    headers = ['rank', 'wallet', 'n_trades', 'maker_cnt', 'taker_cnt',
               'buy_cnt', 'sell_cnt', 'total_vol_usdc', 'maker_vol', 'taker_vol',
               f'net_{N1S}', f'net_{N2S}',
               'taker_fees', 'taker_pnl', 'maker_pnl', 'total_pnl',
               'taker_bps', 'maker_bps', 'total_bps']
    for i, h in enumerate(headers):
        c = ws5.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill

    rows_out = [(w, compute(r)) for w, r in qualified.items()]
    rows_out.sort(key=lambda x: -x[1]['total_vol'])

    for rank, (w, s) in enumerate(rows_out, 1):
        vals = [rank, w[:14]+'...',
                s['n'], s['maker_cnt'], s['taker_cnt'],
                s['buys'], s['sells'],
                round(s['total_vol'], 2), round(s['maker_vol'], 2), round(s['taker_vol'], 2),
                round(s['net_t1'], 2), round(s['net_t2'], 2),
                round(s['taker_fees'], 2),
                round(s['taker_pnl'], 2), round(s['maker_pnl'], 2), round(s['total_pnl'], 2),
                round(s['taker_bps'], 1), round(s['maker_bps'], 1), round(s['total_bps'], 1)]
        for j, v in enumerate(vals):
            ws5.cell(row=rank+1, column=j+1, value=v)
    ws5.freeze_panes = 'A2'
    ws5.column_dimensions['B'].width = 20

    # Save
    ts_now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = f"{args.out_dir}/focused_5sheet_{slug}_{ts_now}.xlsx"
    wb.save(out_path)
    print(f'\nSaved: {out_path}')
    print(f'  Sheet 1: {len(events)} special events (W/4/6)')
    print(f'  Sheet 2: {len(events)} event books')
    print(f'  Sheet 3: {ws3.max_row-1} CLOB trades')
    print(f'  Sheet 4: {ws4.max_row-1} chain fills')
    print(f'  Sheet 5: {len(qualified)} qualified wallets')
    print(f'  Match: {N1} vs {N2}  (settled {N1S}=${t1_settle} {N2S}=${t2_settle})')


if __name__ == '__main__':
    main()
