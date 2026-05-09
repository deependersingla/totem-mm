"""Sanity-test the team_a_event_book.xlsx output by checking that price moves
in the direction theory predicts:

  When team A is BATTING:
      4/6 -> price UP, W -> price DOWN
  When team A is BOWLING (i.e. team B is batting):
      4/6 -> price DOWN (team A's win prob falls), W -> price UP

We measure delta = price[t+20] - price[t-20] per event and aggregate.
"""
from __future__ import annotations

import statistics
from pathlib import Path

from openpyxl import load_workbook

WB = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures/team_a_event_book.xlsx")

# Column indexes (0-based) for price offsets in each row.
# headers row: time_ist, score, event, innings, overs, t-50, t-40, t-30, t-20, t-10, t, t+10, t+20
COL_T_MINUS_20 = 8
COL_T = 10
COL_T_PLUS_20 = 12


def percentile(s: list[float], p: float) -> float:
    if not s:
        return float("nan")
    k = (len(s) - 1) * p
    f = int(k)
    c = min(f + 1, len(s) - 1)
    if f == c:
        return s[f]
    return s[f] + (s[c] - s[f]) * (k - f)


def stats(label: str, deltas: list[float]) -> str:
    if not deltas:
        return f"  {label:<24}: (none)"
    s = sorted(deltas)
    pos = sum(1 for x in deltas if x > 0.005)
    neg = sum(1 for x in deltas if x < -0.005)
    return (
        f"  {label:<24}: n={len(s):>4}  "
        f"p10={percentile(s, 0.10):>+.3f}  "
        f"p50={percentile(s, 0.50):>+.3f}  "
        f"mean={statistics.mean(s):>+.3f}  "
        f"p90={percentile(s, 0.90):>+.3f}  "
        f"  up={pos:>3} ({100*pos/len(s):4.1f}%)  "
        f"down={neg:>3} ({100*neg/len(s):4.1f}%)"
    )


def main() -> None:
    wb = load_workbook(WB, read_only=True)
    pools = {
        "FOUR (team A batting)": [],
        "SIX  (team A batting)": [],
        "WKT  (team A batting)": [],
        "FOUR (team A bowling)": [],
        "SIX  (team A bowling)": [],
        "WKT  (team A bowling)": [],
    }

    for slug in wb.sheetnames:
        if slug == "_summary":
            continue
        ws = wb[slug]
        # Determine team A from header row 1 cell B
        meta = list(ws[1])
        team_a = (meta[1].value or "").replace("team A (winner): ", "").strip()
        # Determine inn1 batting from row 2: e.g. "CSK 192/5"
        meta2 = list(ws[2])
        inn1_batting = (meta2[0].value or "").split(" ")[0]
        # team A is batting in inn1 if team A == inn1_batting
        team_a_bats_inn1 = (team_a == inn1_batting)

        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            event = str(row[2] or "").upper()
            innings = row[3]
            t20m = row[COL_T_MINUS_20]
            tp20 = row[COL_T_PLUS_20]
            if t20m is None or tp20 is None:
                continue
            try:
                delta = float(tp20) - float(t20m)
            except (TypeError, ValueError):
                continue

            team_a_batting = (
                (innings == 1 and team_a_bats_inn1)
                or (innings == 2 and not team_a_bats_inn1)
            )
            mode = "batting" if team_a_batting else "bowling"
            if event == "4":
                pools[f"FOUR (team A {mode})"].append(delta)
            elif event == "6":
                pools[f"SIX  (team A {mode})"].append(delta)
            elif event == "W":
                pools[f"WKT  (team A {mode})"].append(delta)

    print("Delta = price(t+20) - price(t-20)  (team A best bid)")
    print("Expected when team A BATTING:  4/6 > 0,  W < 0")
    print("Expected when team A BOWLING:  4/6 < 0,  W > 0")
    print()
    for k, v in pools.items():
        print(stats(k, v))


if __name__ == "__main__":
    main()
