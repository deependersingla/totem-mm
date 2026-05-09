"""Build a workbook entry for a match using ESPN as the sole cricket signal
(falls back when cricket_events table is empty in the capture DB).

Used for evaluating the predictor on matches where our cricket API capture
failed but ESPN data + book snapshots are present.
"""
from __future__ import annotations

import argparse
import bisect
import json
import sqlite3
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CAP = Path('/Users/sobhagyaxd/DeepWork/totem-mm/captures')
ESPN_WB = CAP / 'espn_ipl2026_ballbyball.xlsx'
IST = timezone(timedelta(hours=5, minutes=30))
OFFSETS_S = [-50, -40, -30, -20, -10, 0, 10, 20]
OFFSET_LABELS = ["t-50","t-40","t-30","t-20","t-10","t","t+10","t+20"]

NAME_TO_SHORT = {
    "Chennai Super Kings": "CSK", "Kolkata Knight Riders": "KKR",
    "Mumbai Indians": "MI", "Royal Challengers Bengaluru": "RCB",
    "Royal Challengers Bangalore": "RCB",
    "Rajasthan Royals": "RR", "Delhi Capitals": "DC",
    "Gujarat Titans": "GT", "Lucknow Super Giants": "LSG",
    "Punjab Kings": "PBKS", "Sunrisers Hyderabad": "SRH",
}


def overs_to_balls(s):
    s = str(s); whole, _, frac = s.partition(".")
    try: return int(whole)*6 + (int(frac) if frac else 0)
    except: return 0


def normalize_event(rec):
    if rec.get('is_wide'): return 'WD'
    if rec.get('is_noball'): return 'NB'
    if rec.get('is_legbye'): return 'LB'
    if rec.get('is_bye'): return 'B'
    pt = (rec.get('play_type') or '').lower()
    if pt == 'out':  return 'W'
    if pt == 'four': return '4'
    if pt == 'six':  return '6'
    if pt == 'no run': return '0'
    if pt == 'run':  return str(rec.get('score_value') or '')
    return ''


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("slug")
    ap.add_argument("--out")
    args = ap.parse_args()
    slug = args.slug

    db = next(CAP.glob(f'match_capture_cricipl-{slug}_*.db'))
    conn = sqlite3.connect(db)
    meta = dict(conn.execute("SELECT key, value FROM match_meta").fetchall())
    token_ids = json.loads(meta['token_ids'])
    outcome_names = json.loads(meta['outcome_names'])
    shorts = [NAME_TO_SHORT.get(o, o) for o in outcome_names]

    # ESPN data
    espn_wb = load_workbook(ESPN_WB, read_only=True)
    if slug not in espn_wb.sheetnames:
        print(f"No ESPN tab for {slug}"); return 1
    ws = espn_wb[slug]
    headers = [c.value for c in ws[6]]
    espn_balls = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] is None: continue
        rec = dict(zip(headers, row))
        # batting team field (ESPN uses team.abbreviation)
        bat = rec.get('batting_team')
        score_running = (rec.get('home_score') if '/' in str(rec.get('home_score') or '') else
                          rec.get('away_score'))
        if score_running is None or '/' not in str(score_running): score_running = '0/0'
        espn_balls.append({
            'innings': int(rec.get('innings') or 0),
            'balls_idx': overs_to_balls(rec.get('over_actual')),
            'overs_str': str(rec.get('over_actual')),
            'ts_ms': int(rec.get('bbb_ts_ms')),
            'bat_short': bat,
            'event': normalize_event(rec),
            'score_str': f"{score_running} ({rec.get('over_actual')})",
        })
    espn_balls.sort(key=lambda b: (b['innings'], b['balls_idx'], b['ts_ms']))

    # Try to detect winner from book at end of last event
    if not espn_balls:
        print("No ESPN balls"); return 1
    last_ts = espn_balls[-1]['ts_ms']
    winner_short = None
    for tid, oname in zip(token_ids, outcome_names):
        r = conn.execute(
            "SELECT bid1_p FROM book_snapshots WHERE asset_id=? AND local_ts_ms BETWEEN ? AND ? "
            "AND bid1_p IS NOT NULL ORDER BY local_ts_ms DESC LIMIT 1",
            (tid, last_ts - 30*60_000, last_ts + 5*60_000)).fetchone()
        if r and r[0] >= 0.95:
            winner_short = NAME_TO_SHORT.get(oname, oname)
            winner_token = tid
    if winner_short is None:
        # Fallback: chase complete = inn2 batting team won
        inn1 = [b for b in espn_balls if b['innings']==1]
        inn2 = [b for b in espn_balls if b['innings']==2]
        if inn2:
            inn2_runs = int(inn2[-1]['score_str'].split('/')[0])
            inn1_runs = int(inn1[-1]['score_str'].split('/')[0]) if inn1 else 0
            if inn2_runs > inn1_runs:
                winner_short = inn2[-1]['bat_short']
            else:
                winner_short = inn1[-1]['bat_short']
        else:
            print(f"Cannot determine winner"); return 1
        winner_token = next(t for t,o in zip(token_ids, outcome_names)
                             if NAME_TO_SHORT.get(o,o)==winner_short)
    loser_short = next(s for s in shorts if s != winner_short)
    print(f"Team A (winner) = {winner_short}, Team B = {loser_short}")

    # Final scores
    inn1_balls = [b for b in espn_balls if b['innings']==1]
    inn2_balls = [b for b in espn_balls if b['innings']==2]
    def parse(s):
        s = s.split(' ')[0]; r,w = s.split('/'); return int(r), int(w)
    inn1_runs, inn1_wkts = parse(inn1_balls[-1]['score_str']) if inn1_balls else (0, 0)
    inn2_runs, inn2_wkts = parse(inn2_balls[-1]['score_str']) if inn2_balls else (0, 0)
    inn1_bat = inn1_balls[0]['bat_short'] if inn1_balls else shorts[0]
    inn2_bat = inn2_balls[0]['bat_short'] if inn2_balls else shorts[1]

    if inn2_runs > inn1_runs:
        end_reason = 'chase_complete'
    elif inn2_wkts >= 10:
        end_reason = 'all_out'
    elif inn1_runs > inn2_runs:
        end_reason = 'overs_complete_lower'
    else:
        end_reason = 'tie_or_no_result'

    # Load book series for team A
    book = conn.execute(
        "SELECT local_ts_ms, bid1_p FROM book_snapshots WHERE asset_id=? AND bid1_p IS NOT NULL ORDER BY local_ts_ms",
        (winner_token,)).fetchall()
    bk_ts = [b[0] for b in book]; bk_p = [b[1] for b in book]

    def near(target_ms):
        if not bk_ts: return None
        i = bisect.bisect_left(bk_ts, target_ms)
        cands = []
        if i < len(bk_ts): cands.append((abs(bk_ts[i]-target_ms), bk_p[i]))
        if i > 0: cands.append((abs(bk_ts[i-1]-target_ms), bk_p[i-1]))
        cands.sort()
        return cands[0][1] if cands else None

    # Build workbook entry
    out_path = Path(args.out) if args.out else CAP / f'team_a_event_book_{slug}.xlsx'
    wb = Workbook(); wb.remove(wb.active)
    ws_out = wb.create_sheet(title=slug[:31])
    ws_out.append([f"slug: {slug}", f"team A (winner): {winner_short}",
                    f"team B: {loser_short}", "verify-vs-book: espn-only"])
    ws_out.append([f"{inn1_bat} {inn1_runs}/{inn1_wkts}",
                    f"{inn2_bat} {inn2_runs}/{inn2_wkts}",
                    f"end: {end_reason}", f"events: {len(espn_balls)}"])
    ws_out.append([])
    headers_row = ["time_ist","score","event","innings","overs"] + OFFSET_LABELS
    ws_out.append(headers_row)
    HF = Font(bold=True); HFL = PatternFill("solid", fgColor="FFE7E6E6")
    for c in ws_out[4]:
        c.font = HF; c.fill = HFL

    for b in espn_balls:
        ist = datetime.fromtimestamp(b['ts_ms']/1000, IST)
        score = f"{b['bat_short']} {b['score_str'].split(' ')[0]} ({b['overs_str']})"
        row = [ist.strftime("%H:%M:%S.%f")[:-3], score, b['event'],
                b['innings'], b['overs_str']]
        for off in OFFSETS_S:
            v = near(b['ts_ms'] + off*1000)
            row.append(round(v, 4) if v is not None else None)
        ws_out.append(row)

    widths = {"time_ist":14,"score":22,"event":6,"innings":6,"overs":7}
    for i, h in enumerate(headers_row, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = widths.get(h, 9)

    # Summary sheet
    ss = wb.create_sheet(title="_summary", index=0)
    ss.append(["slug","team A (winner)","team B","inn1","inn2","end_reason","events","verify_vs_book"])
    ss.append([slug, winner_short, loser_short,
                f"{inn1_bat} {inn1_runs}/{inn1_wkts}",
                f"{inn2_bat} {inn2_runs}/{inn2_wkts}",
                end_reason, len(espn_balls), "espn-only"])
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
