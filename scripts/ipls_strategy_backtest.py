#!/usr/bin/env python3
"""IPL 2026 dip-buy strategy backtest — time-based (no cricket_events dep).

Assumptions
-----------
  * Match starts at 15:30 IST (afternoon) or 19:30 IST (evening). We pick
    whichever is closer to the capture's first book snapshot.
  * Innings 1 runs for 120 minutes from the scheduled start. Over boundaries
    are therefore linear: PP end = start + 36m, 10ov = +60m, 14ov = +84m,
    18ov = +108m, innings-break mid = +125m. Innings 2 starts at +130m.
  * Rain-delayed matches will be off by up to ~30-60m, which is acceptable
    per spec.

Strategy
--------
  1. Team A = pre-match favourite (first book snapshot, higher mid).
  2. At each of 5 checkpoints, if Team A best bid < 0.40, rest a $100 BUY
     limit at that bid.
  3. Buy fill  : best bid on Team A's token later strictly drops below the
                 placed price.
  4. Sell place: at (buy_price + 0.03) from innings-2 start onwards (or from
                 buy-fill time if that happens later).
  5. Sell fill : best ask on Team A's token later strictly rises above the
                 placed price.
  6. Unfilled sells settle to 1.00 if Team A won, 0.00 otherwise (judged by
                 final mid on Team A's token).

Inputs:   captures/match_capture_cricipl-*.db
Outputs:  captures/ipls_strategy_backtest.xlsx
"""

from __future__ import annotations

import glob
import json
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

IST = timezone(timedelta(hours=5, minutes=30))
ROOT = Path(__file__).resolve().parent.parent
CAPTURE_DIR = ROOT / "captures"
OUT_XLSX = CAPTURE_DIR / "ipls_strategy_backtest.xlsx"

# Minutes past scheduled match start for each checkpoint.
# Innings 1 = 120 min, break = 10 min, innings 2 starts at +130 min.
CHECKPOINTS = [
    ("PP (6ov)",    36),
    ("10ov",        60),
    ("14ov",        84),
    ("18ov",       108),
    ("IB (break)", 125),
]
INN2_OFFSET_MIN = 130

BUY_THRESHOLD = 0.40
BUY_NOTIONAL = 100.0
SELL_OFFSET = 0.03


# ─── helpers ──────────────────────────────────────────────────────────────

def fmt_ist(ts_ms: int | float | None) -> str:
    if ts_ms is None:
        return "--"
    return datetime.fromtimestamp(ts_ms / 1000, tz=IST).strftime("%H:%M:%S")


def fmt_ist_date(ts_ms: int | float | None) -> str:
    if ts_ms is None:
        return "--"
    return datetime.fromtimestamp(ts_ms / 1000, tz=IST).strftime("%Y-%m-%d %H:%M")


def infer_scheduled_start(first_snap_ts_ms: int) -> int:
    """Round the capture's first snapshot to the nearest scheduled IPL start
    on that date — 15:30 IST or 19:30 IST."""
    first = datetime.fromtimestamp(first_snap_ts_ms / 1000, tz=IST)
    day = first.date()
    afternoon = datetime.combine(day, datetime.min.time(), tzinfo=IST).replace(hour=15, minute=30)
    evening = datetime.combine(day, datetime.min.time(), tzinfo=IST).replace(hour=19, minute=30)
    pick = afternoon if abs((first - afternoon).total_seconds()) < abs((first - evening).total_seconds()) else evening
    return int(pick.timestamp() * 1000)


def book_at(cur: sqlite3.Cursor, token_id: str, ts_ms: int) -> dict | None:
    """Last book snapshot at or before ts_ms for this token."""
    row = cur.execute(
        """SELECT local_ts_ms, bid1_p, ask1_p, mid_price
             FROM book_snapshots
            WHERE asset_id = ? AND local_ts_ms <= ?
            ORDER BY local_ts_ms DESC LIMIT 1""",
        (token_id, ts_ms),
    ).fetchone()
    if not row:
        return None
    return {"ts": row[0], "bid1": row[1], "ask1": row[2], "mid": row[3]}


def first_book_after(cur: sqlite3.Cursor, token_id: str) -> dict | None:
    row = cur.execute(
        """SELECT local_ts_ms, bid1_p, ask1_p, mid_price
             FROM book_snapshots
            WHERE asset_id = ?
            ORDER BY local_ts_ms ASC LIMIT 1""",
        (token_id,),
    ).fetchone()
    if not row:
        return None
    return {"ts": row[0], "bid1": row[1], "ask1": row[2], "mid": row[3]}


def final_mid(cur: sqlite3.Cursor, token_id: str) -> float | None:
    """Last non-null mid for this token; falls back to last non-null bid
    if every later mid is null (common at match settlement)."""
    row = cur.execute(
        """SELECT mid_price FROM book_snapshots
           WHERE asset_id = ? AND mid_price IS NOT NULL
           ORDER BY local_ts_ms DESC LIMIT 1""",
        (token_id,),
    ).fetchone()
    if row and row[0] is not None:
        return float(row[0])
    row = cur.execute(
        """SELECT bid1_p FROM book_snapshots
           WHERE asset_id = ? AND bid1_p IS NOT NULL
           ORDER BY local_ts_ms DESC LIMIT 1""",
        (token_id,),
    ).fetchone()
    return float(row[0]) if row and row[0] is not None else None


def min_bid_after(cur: sqlite3.Cursor, token_id: str, start_ts: int, end_ts: int | None = None) -> float | None:
    """Lowest bid1_p observed for this token in (start_ts, end_ts]. If end_ts None, uses capture end."""
    if end_ts is None:
        row = cur.execute(
            """SELECT MIN(bid1_p) FROM book_snapshots
               WHERE asset_id = ? AND local_ts_ms > ?""",
            (token_id, start_ts),
        ).fetchone()
    else:
        row = cur.execute(
            """SELECT MIN(bid1_p) FROM book_snapshots
               WHERE asset_id = ? AND local_ts_ms > ? AND local_ts_ms <= ?""",
            (token_id, start_ts, end_ts),
        ).fetchone()
    return float(row[0]) if row and row[0] is not None else None


def max_ask_after(cur: sqlite3.Cursor, token_id: str, start_ts: int, end_ts: int | None = None) -> float | None:
    if end_ts is None:
        row = cur.execute(
            """SELECT MAX(ask1_p) FROM book_snapshots
               WHERE asset_id = ? AND local_ts_ms > ?""",
            (token_id, start_ts),
        ).fetchone()
    else:
        row = cur.execute(
            """SELECT MAX(ask1_p) FROM book_snapshots
               WHERE asset_id = ? AND local_ts_ms > ? AND local_ts_ms <= ?""",
            (token_id, start_ts, end_ts),
        ).fetchone()
    return float(row[0]) if row and row[0] is not None else None


def first_ts_bid_below(cur: sqlite3.Cursor, token_id: str, start_ts: int, price: float) -> int | None:
    row = cur.execute(
        """SELECT local_ts_ms FROM book_snapshots
           WHERE asset_id = ? AND local_ts_ms > ? AND bid1_p < ?
           ORDER BY local_ts_ms ASC LIMIT 1""",
        (token_id, start_ts, price),
    ).fetchone()
    return int(row[0]) if row else None


def first_ts_ask_above(cur: sqlite3.Cursor, token_id: str, start_ts: int, price: float) -> int | None:
    row = cur.execute(
        """SELECT local_ts_ms FROM book_snapshots
           WHERE asset_id = ? AND local_ts_ms > ? AND ask1_p > ?
           ORDER BY local_ts_ms ASC LIMIT 1""",
        (token_id, start_ts, price),
    ).fetchone()
    return int(row[0]) if row else None


# ─── per-match backtest ────────────────────────────────────────────────────

@dataclass
class Buy:
    checkpoint: str          # "PP (6ov)" etc.
    placed_ts: int
    placed_price: float      # bid we rested at
    shares: float            # BUY_NOTIONAL / placed_price
    filled: bool = False
    fill_ts: int | None = None
    sell_price: float | None = None
    sell_placed_ts: int | None = None
    sell_filled: bool = False
    sell_fill_ts: int | None = None
    pnl: float = 0.0
    settlement_value: float | None = None


@dataclass
class MatchResult:
    slug: str
    date: str
    teams: tuple[str, str]
    team_a_idx: int | None
    team_a_name: str
    team_b_name: str
    prematch_price_a: float | None
    prematch_price_b: float | None
    winner: str
    innings_break_ts: int | None
    inn2_end_ts: int | None
    checkpoints: list[dict] = field(default_factory=list)   # per checkpoint prices
    buys: list[Buy] = field(default_factory=list)
    total_pnl: float = 0.0
    total_invested: float = 0.0
    notes: str = ""


def analyse_match(db_path: Path) -> MatchResult | None:
    slug = db_path.stem.replace("match_capture_", "").rsplit("_", 1)[0]
    date = slug.rsplit("-", 3)
    date_str = "-".join(date[-3:]) if len(date) >= 3 else ""

    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    cur = conn.cursor()

    meta = {k: v for k, v in cur.execute("SELECT key, value FROM match_meta")}
    token_ids = json.loads(meta.get("token_ids", "[]"))
    outcomes = json.loads(meta.get("outcome_names", "[]"))
    if len(token_ids) != 2 or len(outcomes) != 2:
        return None

    pre_a = first_book_after(cur, token_ids[0])
    pre_b = first_book_after(cur, token_ids[1])
    if not pre_a or not pre_b or pre_a["bid1"] is None or pre_b["bid1"] is None:
        return MatchResult(
            slug=slug, date=date_str, teams=(outcomes[0], outcomes[1]),
            team_a_idx=None, team_a_name="", team_b_name="",
            prematch_price_a=None, prematch_price_b=None, winner="",
            innings_break_ts=None, inn2_end_ts=None,
            notes="SKIPPED: no pre-match book snapshot",
        )

    # Scheduled start inferred from earliest snapshot across both tokens.
    first_ts = min(pre_a["ts"], pre_b["ts"])
    sched_start_ts = infer_scheduled_start(first_ts)
    innings_break_mid = sched_start_ts + CHECKPOINTS[-1][1] * 60_000
    inn2_start_ts = sched_start_ts + INN2_OFFSET_MIN * 60_000

    # Capture end on Team A token (for fill-window caps).
    cap_end_row = cur.execute(
        "SELECT MAX(local_ts_ms) FROM book_snapshots WHERE asset_id IN (?, ?)",
        (token_ids[0], token_ids[1]),
    ).fetchone()
    inn2_end_ts = int(cap_end_row[0]) if cap_end_row and cap_end_row[0] else inn2_start_ts

    mid_a = pre_a["mid"] if pre_a["mid"] is not None else pre_a["bid1"]
    mid_b = pre_b["mid"] if pre_b["mid"] is not None else pre_b["bid1"]
    if mid_a >= mid_b:
        team_a_idx, team_b_idx = 0, 1
    else:
        team_a_idx, team_b_idx = 1, 0
    team_a_name = outcomes[team_a_idx]
    team_b_name = outcomes[team_b_idx]
    team_a_token = token_ids[team_a_idx]

    final_a = final_mid(cur, token_ids[team_a_idx])
    final_b = final_mid(cur, token_ids[team_b_idx])
    # Decide winner on whichever token settled near 1.0 (>=0.80 lenient
    # because some captures end slightly before the outright).
    if final_a is not None and final_a >= 0.80:
        winner = team_a_name
    elif final_b is not None and final_b >= 0.80:
        winner = team_b_name
    elif final_a is not None and final_b is not None:
        winner = team_a_name if final_a > final_b else team_b_name
    else:
        winner = "?"

    result = MatchResult(
        slug=slug, date=date_str, teams=(outcomes[0], outcomes[1]),
        team_a_idx=team_a_idx, team_a_name=team_a_name, team_b_name=team_b_name,
        prematch_price_a=mid_a, prematch_price_b=mid_b, winner=winner,
        innings_break_ts=innings_break_mid, inn2_end_ts=inn2_end_ts,
        notes=f"sched_start={fmt_ist(sched_start_ts)}",
    )

    for label, offset_min in CHECKPOINTS:
        cp_ts = sched_start_ts + offset_min * 60_000
        if cp_ts < first_ts or cp_ts > inn2_end_ts:
            continue

        book_a = book_at(cur, team_a_token, cp_ts)
        book_b = book_at(cur, token_ids[team_b_idx], cp_ts)
        if book_a is None or book_a["bid1"] is None:
            continue

        cp_info = {
            "label": label,
            "offset_min": offset_min,
            "ts": cp_ts,
            "score": f"T+{offset_min}m",
            "team_a_bid": book_a["bid1"],
            "team_a_ask": book_a["ask1"],
            "team_a_mid": book_a["mid"],
            "team_b_bid": book_b["bid1"] if book_b else None,
            "team_b_ask": book_b["ask1"] if book_b else None,
            "team_b_mid": book_b["mid"] if book_b else None,
        }
        result.checkpoints.append(cp_info)

        bid = book_a["bid1"]
        if bid is not None and bid < BUY_THRESHOLD:
            shares = BUY_NOTIONAL / bid
            buy = Buy(
                checkpoint=label, placed_ts=cp_ts,
                placed_price=float(bid), shares=shares,
            )
            fill_ts = first_ts_bid_below(cur, team_a_token, cp_ts, bid)
            if fill_ts is not None:
                buy.filled = True
                buy.fill_ts = fill_ts
                result.total_invested += BUY_NOTIONAL
            result.buys.append(buy)

    # Sell leg for each filled buy — placed at max(fill_ts, inn2_start_ts)
    for buy in result.buys:
        if not buy.filled:
            # unfilled buy = no position, no P&L
            continue
        sell_price = round(buy.placed_price + SELL_OFFSET, 4)
        sell_placed_ts = max(buy.fill_ts or 0, inn2_start_ts)
        buy.sell_price = sell_price
        buy.sell_placed_ts = sell_placed_ts
        fill_ts = first_ts_ask_above(cur, team_a_token, sell_placed_ts, sell_price)
        if fill_ts is not None:
            buy.sell_filled = True
            buy.sell_fill_ts = fill_ts
            buy.pnl = buy.shares * (sell_price - buy.placed_price)
        else:
            # MTM at settlement
            settle = 1.0 if winner == team_a_name else (0.0 if winner == team_b_name else (final_a or buy.placed_price))
            buy.settlement_value = settle
            buy.pnl = buy.shares * (settle - buy.placed_price)
        result.total_pnl += buy.pnl

    conn.close()
    return result


# ─── report ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="E7EEF7")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WHITE_BOLD = Font(bold=True, color="FFFFFF")


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = WHITE_BOLD
    cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_report(results: list[MatchResult], out_path: Path) -> None:
    wb = Workbook()

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    headers = [
        "Date", "Slug", "Team A (fav)", "Team B", "Pre-match A",
        "Pre-match B", "Winner", "Buys placed", "Buys filled",
        "Sells filled", "Shares held", "Invested $", "Realised $",
        "MTM $", "Net P&L $", "ROI %", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        style_header(cell)

    agg_invested = 0.0
    agg_pnl = 0.0
    agg_realised = 0.0
    agg_mtm = 0.0
    agg_buys_placed = 0
    agg_buys_filled = 0
    agg_sells_filled = 0

    for idx, r in enumerate(results, 2):
        buys_placed = len(r.buys)
        buys_filled = sum(1 for b in r.buys if b.filled)
        sells_filled = sum(1 for b in r.buys if b.sell_filled)
        shares_held = sum(b.shares for b in r.buys if b.filled)
        realised = sum(b.pnl for b in r.buys if b.sell_filled)
        mtm = sum(b.pnl for b in r.buys if b.filled and not b.sell_filled)
        net = r.total_pnl
        roi = (net / r.total_invested * 100) if r.total_invested else 0.0

        agg_invested += r.total_invested
        agg_pnl += net
        agg_realised += realised
        agg_mtm += mtm
        agg_buys_placed += buys_placed
        agg_buys_filled += buys_filled
        agg_sells_filled += sells_filled

        row = [
            r.date, r.slug, r.team_a_name, r.team_b_name,
            round(r.prematch_price_a, 4) if r.prematch_price_a else None,
            round(r.prematch_price_b, 4) if r.prematch_price_b else None,
            r.winner, buys_placed, buys_filled, sells_filled,
            round(shares_held, 2), round(r.total_invested, 2),
            round(realised, 2), round(mtm, 2), round(net, 2),
            round(roi, 2), r.notes,
        ]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=idx, column=c, value=v)
            if c == 15 and isinstance(v, (int, float)):
                if v > 0:
                    cell.fill = GOOD_FILL
                elif v < 0:
                    cell.fill = BAD_FILL

    # Totals row
    tot_row = len(results) + 3
    ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tot_row, column=8, value=agg_buys_placed).font = Font(bold=True)
    ws.cell(row=tot_row, column=9, value=agg_buys_filled).font = Font(bold=True)
    ws.cell(row=tot_row, column=10, value=agg_sells_filled).font = Font(bold=True)
    ws.cell(row=tot_row, column=12, value=round(agg_invested, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=13, value=round(agg_realised, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=14, value=round(agg_mtm, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=15, value=round(agg_pnl, 2)).font = Font(bold=True)
    ws.cell(row=tot_row, column=16,
            value=round(agg_pnl / agg_invested * 100, 2) if agg_invested else 0).font = Font(bold=True)

    autosize(ws, [12, 32, 25, 25, 12, 12, 22, 11, 11, 12, 12, 12, 12, 10, 11, 9, 42])

    # ── Checkpoints (per match, one row per checkpoint) ──
    ws2 = wb.create_sheet("Checkpoints")
    cp_headers = [
        "Date", "Slug", "Team A", "Checkpoint", "IST time", "Score",
        "A bid", "A ask", "A mid", "B bid", "B ask", "B mid",
        "Triggered buy?", "Buy price", "Filled?", "Fill IST",
        "Sell price", "Sell filled?", "Sell fill IST", "P&L $",
    ]
    for c, h in enumerate(cp_headers, 1):
        style_header(ws2.cell(row=1, column=c, value=h))

    row_i = 2
    for r in results:
        if not r.checkpoints:
            continue
        buy_by_cp = {b.checkpoint: b for b in r.buys}
        for cp in r.checkpoints:
            b = buy_by_cp.get(cp["label"])
            triggered = b is not None
            line = [
                r.date, r.slug, r.team_a_name, cp["label"], fmt_ist(cp["ts"]),
                cp["score"],
                round(cp["team_a_bid"], 4) if cp["team_a_bid"] is not None else None,
                round(cp["team_a_ask"], 4) if cp["team_a_ask"] is not None else None,
                round(cp["team_a_mid"], 4) if cp["team_a_mid"] is not None else None,
                round(cp["team_b_bid"], 4) if cp["team_b_bid"] is not None else None,
                round(cp["team_b_ask"], 4) if cp["team_b_ask"] is not None else None,
                round(cp["team_b_mid"], 4) if cp["team_b_mid"] is not None else None,
                "YES" if triggered else "",
                round(b.placed_price, 4) if triggered else None,
                ("YES" if b.filled else "NO") if triggered else "",
                fmt_ist(b.fill_ts) if (triggered and b.filled) else "",
                round(b.sell_price, 4) if (triggered and b.sell_price is not None) else None,
                ("YES" if b.sell_filled else "NO") if (triggered and b.filled) else "",
                fmt_ist(b.sell_fill_ts) if (triggered and b.sell_filled) else "",
                round(b.pnl, 2) if (triggered and b.filled) else None,
            ]
            for c, v in enumerate(line, 1):
                cell = ws2.cell(row=row_i, column=c, value=v)
                if c == 7 and isinstance(v, (int, float)) and v < BUY_THRESHOLD:
                    cell.fill = SUB_FILL
            row_i += 1

    autosize(ws2, [12, 32, 22, 12, 11, 16, 9, 9, 9, 9, 9, 9, 14, 11, 9, 11, 11, 12, 11, 10])

    # ── Trades (one row per buy) ──
    ws3 = wb.create_sheet("Trades")
    t_headers = [
        "Date", "Slug", "Team A", "Checkpoint", "Placed IST", "Buy price",
        "Shares", "Filled?", "Fill IST", "Sell price",
        "Sell placed IST", "Sell filled?", "Sell fill IST",
        "Settle value", "P&L $",
    ]
    for c, h in enumerate(t_headers, 1):
        style_header(ws3.cell(row=1, column=c, value=h))
    row_i = 2
    for r in results:
        for b in r.buys:
            line = [
                r.date, r.slug, r.team_a_name, b.checkpoint, fmt_ist(b.placed_ts),
                round(b.placed_price, 4), round(b.shares, 2),
                "YES" if b.filled else "NO", fmt_ist(b.fill_ts),
                round(b.sell_price, 4) if b.sell_price is not None else None,
                fmt_ist(b.sell_placed_ts), "YES" if b.sell_filled else ("NO" if b.filled else ""),
                fmt_ist(b.sell_fill_ts),
                round(b.settlement_value, 4) if b.settlement_value is not None else None,
                round(b.pnl, 2) if b.filled else 0.0,
            ]
            for c, v in enumerate(line, 1):
                cell = ws3.cell(row=row_i, column=c, value=v)
                if c == 15 and isinstance(v, (int, float)):
                    if v > 0:
                        cell.fill = GOOD_FILL
                    elif v < 0:
                        cell.fill = BAD_FILL
            row_i += 1
    autosize(ws3, [12, 32, 22, 12, 11, 11, 10, 9, 11, 11, 15, 12, 13, 12, 11])

    # ── Skipped matches ──
    ws4 = wb.create_sheet("Skipped")
    for c, h in enumerate(["Slug", "Reason"], 1):
        style_header(ws4.cell(row=1, column=c, value=h))
    row_i = 2
    for r in results:
        if r.notes:
            ws4.cell(row=row_i, column=1, value=r.slug)
            ws4.cell(row=row_i, column=2, value=r.notes)
            row_i += 1
    autosize(ws4, [36, 70])

    # ── Strategy description ──
    ws5 = wb.create_sheet("Strategy")
    desc = [
        ["IPL 2026 — Dip-buy on pre-match favourite (time-based)"],
        [""],
        ["Scheduled start", "Inferred: 15:30 IST (afternoon) or 19:30 IST (evening), whichever is closer to capture's first snapshot"],
        ["Team A selection", "Pre-match favourite (first book snapshot, higher mid)"],
        ["Checkpoints (min)", f"PP={CHECKPOINTS[0][1]}, 10ov={CHECKPOINTS[1][1]}, 14ov={CHECKPOINTS[2][1]}, 18ov={CHECKPOINTS[3][1]}, IB={CHECKPOINTS[4][1]} — past scheduled start"],
        ["Innings 2 start", f"Scheduled start + {INN2_OFFSET_MIN} min"],
        ["Buy trigger", f"Team A bid1 < {BUY_THRESHOLD}"],
        ["Buy fill rule", "Best bid later drops strictly below placed price (anywhere in the capture)"],
        ["Buy notional", f"${BUY_NOTIONAL:.0f} per triggered checkpoint"],
        ["Sell price", f"Buy price + ${SELL_OFFSET:.2f}"],
        ["Sell placed", "At max(buy fill ts, innings-2 start)"],
        ["Sell fill rule", "Best ask later rises strictly above placed price"],
        ["Unfilled sell", "Mark-to-market: 1.00 if Team A won, 0.00 if lost, final mid otherwise"],
        ["Rain delay", "Checkpoints may miss by ~30–60m; flagged only via scheduled-start note"],
    ]
    for r_i, row in enumerate(desc, 1):
        for c_i, v in enumerate(row, 1):
            cell = ws5.cell(row=r_i, column=c_i, value=v)
            if r_i == 1:
                cell.font = Font(bold=True, size=14)
    autosize(ws5, [28, 80])

    wb.save(out_path)


# ─── entrypoint ──────────────────────────────────────────────────────────

def main() -> None:
    dbs = sorted(CAPTURE_DIR.glob("match_capture_cricipl-*.db"))
    print(f"[ipls_strategy_backtest] scanning {len(dbs)} match captures …")

    results: list[MatchResult] = []
    for db in dbs:
        try:
            r = analyse_match(db)
        except Exception as exc:
            slug = db.stem.replace("match_capture_", "")
            print(f"  ERROR  {slug}: {exc}")
            continue
        if r is None:
            continue
        results.append(r)
        if r.notes.startswith("SKIPPED"):
            print(f"  SKIP   {r.slug:36s}  {r.notes}")
        else:
            filled = sum(1 for b in r.buys if b.filled)
            sells = sum(1 for b in r.buys if b.sell_filled)
            print(f"  OK     {r.slug:36s}  Team A={r.team_a_name:26s} "
                  f"buys={len(r.buys)} filled={filled} sells={sells} "
                  f"P&L=${r.total_pnl:+7.2f}  [{r.notes}]")

    write_report(results, OUT_XLSX)
    print(f"\nreport → {OUT_XLSX}")


if __name__ == "__main__":
    main()
