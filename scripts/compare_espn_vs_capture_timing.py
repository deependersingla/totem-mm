"""Compare per-ball timestamps between two cricket data sources:
  (A) the in-house cricket_events captured into match_capture_*.db
  (B) ESPN's playbyplay bbbTimestamp (cached in espn_ipl2026_ballbyball.xlsx)

For each match in the ESPN workbook that has a matching capture DB, we align
deliveries by (innings, over_actual, ball_in_over) and compute:
    delta_ms = capture.local_ts_ms - espn.bbb_ts_ms

Positive delta -> our cricket_events feed was LATER than ESPN.
Negative delta -> our cricket_events feed was EARLIER than ESPN.

For each pairing we record:
    match_slug, innings, over, ball, espn_ts, capture_ts, delta_ms,
    espn_play_type, capture_signal, espn_score_after, capture_score

Reports per-match stats and an aggregate.
"""
from __future__ import annotations

import sqlite3
import statistics
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

CAPTURES = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures")
WORKBOOK = CAPTURES / "espn_ipl2026_ballbyball.xlsx"
IST = timezone(timedelta(hours=5, minutes=30))


def overs_to_balls(overs: float | str | None) -> int:
    if overs is None:
        return 0
    try:
        s = str(overs)
        whole, _, frac = s.partition(".")
        return int(whole) * 6 + (int(frac) if frac else 0)
    except (ValueError, AttributeError):
        return 0


def detect_split(rows: list[tuple]) -> int | None:
    """Same heuristic as the other scripts: gap >= 5 min + score reset = innings 2."""
    for i in range(1, len(rows)):
        gap = (rows[i][1] - rows[i - 1][1]) / 1000.0
        if (
            gap >= 300
            and overs_to_balls(rows[i][4]) < overs_to_balls(rows[i - 1][4])
            and rows[i][2] < rows[i - 1][2]
            and rows[i][2] <= 10
        ):
            return i
    return None


def load_capture_balls(slug: str) -> list[dict]:
    """Return list of {innings, balls_idx, overs_str, ts_ms, signal} ordered."""
    candidates = list(CAPTURES.glob(f"match_capture_cricipl-{slug}_*.db"))
    if not candidates:
        return []
    db = candidates[0]
    conn = sqlite3.connect(db)
    rows = conn.execute(
        "SELECT id, local_ts_ms, COALESCE(runs, 0), COALESCE(wickets, 0), "
        "COALESCE(overs, '0.0'), signal_type FROM cricket_events ORDER BY id"
    ).fetchall()
    conn.close()
    if not rows:
        return []
    split = detect_split(rows)
    out: list[dict] = []
    inn = 1
    prev_balls = 0
    for i, (rid, ts, runs, wkts, overs, sig) in enumerate(rows):
        if split is not None and i == split:
            inn = 2
            prev_balls = 0
        cur_balls = overs_to_balls(overs)
        out.append({
            "innings": inn,
            "overs_str": str(overs),
            "balls_idx": cur_balls,
            "ts_ms": ts,
            "signal": sig,
            "score_str": f"{runs}/{wkts}",
            "is_legal": cur_balls == prev_balls + 1,
        })
        prev_balls = cur_balls
    return out


def load_espn_balls(ws) -> list[dict]:
    """Read the ESPN tab into per-ball dicts. Header row = 6, data starts row 7."""
    headers = [c.value for c in ws[6]]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] is None:
            continue
        rec = dict(zip(headers, row))
        out.append({
            "innings": rec.get("innings"),
            "overs_str": str(rec.get("over_actual")),
            "balls_idx": overs_to_balls(rec.get("over_actual")),
            "ts_ms": int(rec.get("bbb_ts_ms")),
            "play_type": rec.get("play_type"),
            "is_wide": int(rec.get("is_wide") or 0),
            "is_noball": int(rec.get("is_noball") or 0),
            "score_after": rec.get("home_score"),
            "ball_in_over": rec.get("ball_in_over"),
        })
    return out


def percentile(s: list[float], p: float) -> float:
    if not s:
        return float("nan")
    k = (len(s) - 1) * p
    f = int(k)
    c = min(f + 1, len(s) - 1)
    if f == c:
        return s[f]
    return s[f] + (s[c] - s[f]) * (k - f)


def stats_block(label: str, deltas_sec: list[float]) -> str:
    if not deltas_sec:
        return f"  {label:<28}: (none)"
    s = sorted(deltas_sec)
    return (
        f"  {label:<28}: n={len(s):>4}  "
        f"min={s[0]:>+7.1f}  p10={percentile(s, 0.10):>+7.1f}  "
        f"p50={percentile(s, 0.50):>+7.1f}  mean={statistics.mean(s):>+7.1f}  "
        f"p90={percentile(s, 0.90):>+7.1f}  p99={percentile(s, 0.99):>+7.1f}  "
        f"max={s[-1]:>+8.1f}"
    )


def align_match(slug: str, ws) -> dict:
    """Pair ESPN balls with capture balls by (innings, overs_str). Multiple
    capture events with the same overs_str (extras + the re-bowled legal) are
    matched in arrival order to the ESPN events in the same bucket.
    """
    capture = load_capture_balls(slug)
    espn = load_espn_balls(ws)
    if not capture or not espn:
        return {"slug": slug, "n_capture": len(capture), "n_espn": len(espn),
                "matched": [], "unmatched_capture": [], "unmatched_espn": []}

    # Group both feeds by (innings, overs_str)
    from collections import defaultdict
    cap_by_key: dict[tuple, list[dict]] = defaultdict(list)
    for c in capture:
        cap_by_key[(c["innings"], c["overs_str"])].append(c)
    espn_by_key: dict[tuple, list[dict]] = defaultdict(list)
    for e in espn:
        espn_by_key[(e["innings"], e["overs_str"])].append(e)

    matched: list[dict] = []
    unmatched_cap: list[dict] = []
    unmatched_espn: list[dict] = []

    keys = set(cap_by_key) | set(espn_by_key)
    for key in sorted(keys):
        cs = cap_by_key.get(key, [])
        es = espn_by_key.get(key, [])
        n = min(len(cs), len(es))
        for i in range(n):
            c = cs[i]
            e = es[i]
            matched.append({
                "innings": key[0],
                "overs": key[1],
                "espn_ts_ms": e["ts_ms"],
                "cap_ts_ms": c["ts_ms"],
                "delta_ms": c["ts_ms"] - e["ts_ms"],
                "espn_play_type": e["play_type"],
                "cap_signal": c["signal"],
                "espn_score": e["score_after"],
                "cap_score": c["score_str"],
            })
        unmatched_cap.extend(cs[n:])
        unmatched_espn.extend(es[n:])

    return {"slug": slug, "n_capture": len(capture), "n_espn": len(espn),
            "matched": matched, "unmatched_capture": unmatched_cap,
            "unmatched_espn": unmatched_espn}


def main() -> int:
    if not WORKBOOK.exists():
        print(f"missing {WORKBOOK} — run fetch_espn_ipl_ballbyball.py first")
        return 1

    wb = load_workbook(WORKBOOK, read_only=True)
    pooled: list[float] = []
    pooled_4: list[float] = []
    pooled_6: list[float] = []
    pooled_w: list[float] = []
    per_match: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "_summary":
            continue
        ws = wb[sheet_name]
        slug = sheet_name  # e.g. "che-kol-2026-04-14"
        result = align_match(slug, ws)
        if not result["matched"]:
            print(f"\n## {slug}: no overlap "
                  f"(capture={result['n_capture']}, espn={result['n_espn']})")
            continue
        per_match.append(result)

        deltas = [m["delta_ms"] / 1000 for m in result["matched"]]
        pooled.extend(deltas)
        pooled_4.extend(m["delta_ms"] / 1000 for m in result["matched"]
                        if (m["espn_play_type"] or "").lower() == "four")
        pooled_6.extend(m["delta_ms"] / 1000 for m in result["matched"]
                        if (m["espn_play_type"] or "").lower() == "six")
        pooled_w.extend(m["delta_ms"] / 1000 for m in result["matched"]
                        if (m["espn_play_type"] or "").lower() in ("wicket", "out"))

        print(f"\n## {slug}")
        print(f"   capture rows: {result['n_capture']}  "
              f"espn rows: {result['n_espn']}  "
              f"matched: {len(result['matched'])}  "
              f"unmatched capture/espn: "
              f"{len(result['unmatched_capture'])}/{len(result['unmatched_espn'])}")
        print(stats_block("delta (capture - espn) sec", deltas))

    print("\n" + "=" * 110)
    print(f"AGGREGATE across {len(per_match)} matches")
    print("=" * 110)
    print(stats_block("ALL deliveries", pooled))
    print(stats_block("ESPN play_type=four", pooled_4))
    print(stats_block("ESPN play_type=six", pooled_6))
    print(stats_block("ESPN play_type=wicket/out", pooled_w))

    print("\nInterpretation:")
    print("  delta > 0  -> our cricket_events feed published LATER than ESPN")
    print("  delta < 0  -> our cricket_events feed published EARLIER than ESPN")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
