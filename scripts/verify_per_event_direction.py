"""Per-event price direction verification.

The market reacts to a ball within a few seconds (book updates are real-time)
but our cricket_events feed lands the timestamp ~15-30s LATER (it's pulled
from an API, not live). So by 't' (our event time) the price has already
fully digested the move. The actual reaction is captured BEFORE our 't'.

We measure with three pre-event windows:
    Δa  = price(t-10) - price(t-50)        (40s of pre-window — main signal)
    Δb  = price(t-20) - price(t-50)        (30s of pre-window)
    Δc  = price(t-10) - price(t-30)        (20s of pre-window — tightest)

Theory:
    4/6  + team A batting -> UP
    4/6  + team A bowling -> DOWN
    W    + team A batting -> DOWN
    W    + team A bowling -> UP
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

WB = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures/team_a_event_book.xlsx")
FLAT_EPS = 0.005

# Column indices (0-based) in each row:
# 0=time, 1=score, 2=event, 3=innings, 4=overs,
# 5=t-50 6=t-40 7=t-30 8=t-20 9=t-10 10=t 11=t+10 12=t+20
COL = {-50: 5, -40: 6, -30: 7, -20: 8, -10: 9, 0: 10, 10: 11, 20: 12}


def expected_direction(event: str, team_a_batting: bool) -> str:
    if event in ("4", "6"):
        return "UP" if team_a_batting else "DOWN"
    if event == "W":
        return "DOWN" if team_a_batting else "UP"
    return ""


def classify(delta: float) -> str:
    if delta > FLAT_EPS:
        return "UP"
    if delta < -FLAT_EPS:
        return "DOWN"
    return "FLAT"


def main() -> None:
    wb = load_workbook(WB, read_only=True)
    per_match_rows = []
    agg_short = Counter()
    agg_medium = Counter()
    agg_long = Counter()
    counter_theory_examples = []

    for slug in wb.sheetnames:
        if slug == "_summary":
            continue
        ws = wb[slug]
        team_a = (list(ws[1])[1].value or "").replace("team A (winner): ", "").strip()
        inn1_meta = (list(ws[2])[0].value or "").strip()
        inn1_batting = inn1_meta.split(" ")[0]
        team_a_bats_inn1 = (team_a == inn1_batting)

        local_short = Counter()
        local_medium = Counter()
        local_long = Counter()

        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None:
                continue
            event = str(row[2] or "").upper()
            if event not in ("4", "6", "W"):
                continue
            innings = row[3]
            team_a_batting = (
                (innings == 1 and team_a_bats_inn1)
                or (innings == 2 and not team_a_bats_inn1)
            )
            expected = expected_direction(event, team_a_batting)

            for window_label, (a_off, b_off), counter in [
                ("a_50_to_10", (-50, -10), local_short),
                ("b_50_to_20", (-50, -20), local_medium),
                ("c_30_to_10", (-30, -10), local_long),
            ]:
                a = row[COL[a_off]]
                b = row[COL[b_off]]
                if a is None or b is None:
                    continue
                d = float(b) - float(a)
                cls = classify(d)
                bucket = (event, "bat" if team_a_batting else "bowl")
                if cls == "FLAT":
                    counter[(bucket, "flat")] += 1
                elif cls == expected:
                    counter[(bucket, "match")] += 1
                else:
                    counter[(bucket, "anti")] += 1
                    if window_label == "a_50_to_10":
                        counter_theory_examples.append({
                            "slug": slug, "innings": innings,
                            "overs": row[4], "event": event,
                            "team_a_role": "bat" if team_a_batting else "bowl",
                            "expected": expected, "got": cls,
                            "delta": d, "score": row[1],
                            "time": row[0],
                        })

        per_match_rows.append({
            "slug": slug,
            "short": local_short,
            "medium": local_medium,
            "long": local_long,
        })
        for k, v in local_short.items():
            agg_short[k] += v
        for k, v in local_medium.items():
            agg_medium[k] += v
        for k, v in local_long.items():
            agg_long[k] += v

    def render_table(title: str, agg: Counter) -> None:
        print(f"\n{title}")
        print(f"{'event/role':<22} {'match':>7} {'flat':>6} {'anti':>6} {'total':>6} "
              f"{'match%':>8} {'anti%':>7}")
        print("-" * 70)
        for event, role, label in [
            ("4", "bat", "FOUR — team A batting"),
            ("6", "bat", "SIX  — team A batting"),
            ("W", "bat", "WKT  — team A batting"),
            ("4", "bowl", "FOUR — team A bowling"),
            ("6", "bowl", "SIX  — team A bowling"),
            ("W", "bowl", "WKT  — team A bowling"),
        ]:
            m = agg.get(((event, role), "match"), 0)
            f = agg.get(((event, role), "flat"), 0)
            a = agg.get(((event, role), "anti"), 0)
            t = m + f + a
            mp = (100 * m / t) if t else 0
            ap = (100 * a / t) if t else 0
            print(f"{label:<22} {m:>7} {f:>6} {a:>6} {t:>6} "
                  f"{mp:>7.1f}% {ap:>6.1f}%")

    render_table("=== AGGREGATE — Δ = p(t-10) − p(t-50)  (40s pre-window) ===", agg_short)
    render_table("=== AGGREGATE — Δ = p(t-20) − p(t-50)  (30s pre-window) ===", agg_medium)
    render_table("=== AGGREGATE — Δ = p(t-10) − p(t-30)  (20s pre-window) ===", agg_long)

    # Per-match summary using the 40s pre-window (main signal)
    print("\n=== PER-MATCH (Δ = p(t-10) − p(t-50)) ===")
    print(f"{'match':<26} {'4_bat':>8} {'6_bat':>8} {'W_bat':>8} "
          f"{'4_bowl':>8} {'6_bowl':>8} {'W_bowl':>8}")
    print("-" * 90)
    for r in per_match_rows:
        cells = []
        for ev, role in [("4","bat"),("6","bat"),("W","bat"),
                         ("4","bowl"),("6","bowl"),("W","bowl")]:
            m = r["medium"].get(((ev, role), "match"), 0)
            f = r["medium"].get(((ev, role), "flat"), 0)
            a = r["medium"].get(((ev, role), "anti"), 0)
            t = m + f + a
            cells.append(f"{m}/{t}" if t else "—")
        print(f"{r['slug']:<26} " + " ".join(f"{c:>8}" for c in cells))
    print("\nFormat: matches/total. \"matches\" = rows where Δ direction matched theory.")

    # Counter-theory examples
    print(f"\n=== {len(counter_theory_examples)} COUNTER-THEORY events "
          f"(40s pre-window: t-50 to t-10) ===")
    print("(events where the price moved AGAINST theory by >0.5%; samples to spot-check)")
    print()
    counter_theory_examples.sort(key=lambda x: abs(x["delta"]), reverse=True)
    print(f"{'slug':<28} {'inn':>3} {'ovs':>5} {'ev':>3} {'role':>5} "
          f"{'expected':>8} {'got':>5} {'delta':>7} {'score':<22} {'time':>14}")
    print("-" * 120)
    for ex in counter_theory_examples[:30]:
        print(f"{ex['slug']:<28} {ex['innings']:>3} {ex['overs']:>5} {ex['event']:>3} "
              f"{ex['team_a_role']:>5} {ex['expected']:>8} {ex['got']:>5} "
              f"{ex['delta']:>+7.3f} {str(ex['score'])[:22]:<22} {ex['time']:>14}")


if __name__ == "__main__":
    main()
