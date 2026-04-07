#!/usr/bin/env python3
"""
Spread & Depth Analysis of Polymarket Cricket Orderbook Data
DC vs MI, IPL, April 4 2026
"""

import openpyxl
import numpy as np
from collections import defaultdict
import sys
import os

XLSX = '/Users/sobhagyaxd/DeepWork/totem-mm/captures/20260404_153643_event_book_cricipl-del-mum-2026-04-04.xlsx'
OUTPUT = '/Users/sobhagyaxd/DeepWork/totem-mm/polymarket-taker/docs/spread_depth_analysis.md'

# Column indices (0-based) from header row:
# 0: IST, 1: ms_from_event
# DC: bid1_p=2, bid1_s=3, bid2_p=4, bid2_s=5, ..., bid5_p=10, bid5_s=11
#     ask1_p=12, ask1_s=13, ask2_p=14, ask2_s=15, ..., ask5_p=20, ask5_s=21
# MI: bid1_p=22, bid1_s=23, bid2_p=24, bid2_s=25, ..., bid5_p=30, bid5_s=31
#     ask1_p=32, ask1_s=33, ask2_p=34, ask2_s=35, ..., ask5_p=40, ask5_s=41

COL = {
    'ms': 1,
    'dc_bid1_p': 2, 'dc_bid1_s': 3,
    'dc_bid2_p': 4, 'dc_bid2_s': 5,
    'dc_bid3_p': 6, 'dc_bid3_s': 7,
    'dc_bid4_p': 8, 'dc_bid4_s': 9,
    'dc_bid5_p': 10, 'dc_bid5_s': 11,
    'dc_ask1_p': 12, 'dc_ask1_s': 13,
    'dc_ask2_p': 14, 'dc_ask2_s': 15,
    'dc_ask3_p': 16, 'dc_ask3_s': 17,
    'dc_ask4_p': 18, 'dc_ask4_s': 19,
    'dc_ask5_p': 20, 'dc_ask5_s': 21,
    'mi_bid1_p': 22, 'mi_bid1_s': 23,
    'mi_bid2_p': 24, 'mi_bid2_s': 25,
    'mi_bid3_p': 26, 'mi_bid3_s': 27,
    'mi_bid4_p': 28, 'mi_bid4_s': 29,
    'mi_bid5_p': 30, 'mi_bid5_s': 31,
    'mi_ask1_p': 32, 'mi_ask1_s': 33,
    'mi_ask2_p': 34, 'mi_ask2_s': 35,
    'mi_ask3_p': 36, 'mi_ask3_s': 37,
    'mi_ask4_p': 38, 'mi_ask4_s': 39,
    'mi_ask5_p': 40, 'mi_ask5_s': 41,
}


def load_all_data():
    """Load all sheets, return list of dicts per sheet."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    sheets_data = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))

        # Parse header row (row 0)
        header_info = rows[0] if len(rows) > 0 else None
        event_type = str(header_info[0]).replace('Event: ', '') if header_info else '?'
        score = str(header_info[1]).replace('Score: ', '') if header_info and header_info[1] else '?'
        event_time = str(header_info[2]).replace('Time: ', '') if header_info and header_info[2] else '?'

        snapshots = []
        for i, r in enumerate(rows):
            if i < 3:  # skip header rows
                continue
            if r[COL['dc_bid1_p']] is None and r[COL['dc_ask1_p']] is None:
                continue

            snap = {}
            for key, idx in COL.items():
                val = r[idx] if idx < len(r) else None
                if val is not None and isinstance(val, (int, float)):
                    snap[key] = float(val)
                else:
                    snap[key] = None
            snapshots.append(snap)

        sheets_data.append({
            'name': sname,
            'event_type': event_type,
            'score': score,
            'event_time': event_time,
            'snapshots': snapshots,
        })

    wb.close()
    return sheets_data


def analyze_spreads(all_snaps, sheets_data):
    """Section 1: Spread Distribution."""
    dc_spreads = []
    mi_spreads = []
    combined_spreads = []

    # Spread by price bucket
    spread_by_price = defaultdict(list)  # bucket -> list of combined spreads

    # Spread by event phase
    spread_before = []   # ms < -5000
    spread_during = []   # -5000 <= ms <= 5000
    spread_after = []    # ms > 5000

    for snap in all_snaps:
        dc_b = snap.get('dc_bid1_p')
        dc_a = snap.get('dc_ask1_p')
        mi_b = snap.get('mi_bid1_p')
        mi_a = snap.get('mi_ask1_p')
        ms = snap.get('ms')

        if dc_b is not None and dc_a is not None:
            dc_sp = round((dc_a - dc_b) * 100, 2)  # in cents
            dc_spreads.append(dc_sp)
        else:
            dc_sp = None

        if mi_b is not None and mi_a is not None:
            mi_sp = round((mi_a - mi_b) * 100, 2)
            mi_spreads.append(mi_sp)
        else:
            mi_sp = None

        if dc_b is not None and dc_a is not None and mi_b is not None and mi_a is not None:
            comb = round((dc_a + mi_a - dc_b - mi_b) * 100, 2)
            combined_spreads.append(comb)

            # Price bucket (use DC bid as proxy for price level)
            mid = (dc_b + dc_a) / 2
            if mid < 0.20:
                bucket = '<20c'
            elif mid < 0.35:
                bucket = '20-35c'
            elif mid < 0.50:
                bucket = '35-50c'
            elif mid < 0.65:
                bucket = '50-65c'
            elif mid < 0.80:
                bucket = '65-80c'
            else:
                bucket = '>80c'
            spread_by_price[bucket].append(comb)

            # Event phase
            if ms is not None:
                if ms < -5000:
                    spread_before.append(comb)
                elif ms <= 5000:
                    spread_during.append(comb)
                else:
                    spread_after.append(comb)

    out = []
    out.append("## 1. Spread Distribution\n")
    out.append(f"Total snapshots with valid book: {len(combined_spreads):,}\n")

    # DC spread distribution
    out.append("### DC Spread (ask1 - bid1)\n")
    dc_arr = np.array(dc_spreads)
    out.append(f"- Mean: {dc_arr.mean():.2f}c | Median: {np.median(dc_arr):.2f}c | Std: {dc_arr.std():.2f}c\n")
    for threshold in [1, 2, 3, 4, 5]:
        pct = np.sum(dc_arr == threshold) / len(dc_arr) * 100
        out.append(f"  - Exactly {threshold}c: {pct:.1f}%")
    pct5plus = np.sum(dc_arr >= 5) / len(dc_arr) * 100
    out.append(f"  - 5c+: {pct5plus:.1f}%\n")

    # MI spread distribution
    out.append("### MI Spread (ask1 - bid1)\n")
    mi_arr = np.array(mi_spreads)
    out.append(f"- Mean: {mi_arr.mean():.2f}c | Median: {np.median(mi_arr):.2f}c | Std: {mi_arr.std():.2f}c\n")
    for threshold in [1, 2, 3, 4, 5]:
        pct = np.sum(mi_arr == threshold) / len(mi_arr) * 100
        out.append(f"  - Exactly {threshold}c: {pct:.1f}%")
    pct5plus = np.sum(mi_arr >= 5) / len(mi_arr) * 100
    out.append(f"  - 5c+: {pct5plus:.1f}%\n")

    # Combined spread
    out.append("### Combined Spread (DC_ask + MI_ask - DC_bid - MI_bid)\n")
    comb_arr = np.array(combined_spreads)
    out.append(f"- Mean: {comb_arr.mean():.2f}c | Median: {np.median(comb_arr):.2f}c | Std: {comb_arr.std():.2f}c\n")
    for threshold in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        pct = np.sum(comb_arr == threshold) / len(comb_arr) * 100
        pct_leq = np.sum(comb_arr <= threshold) / len(comb_arr) * 100
        out.append(f"  - {threshold}c: {pct:.1f}% (cumulative <= {threshold}c: {pct_leq:.1f}%)")
    out.append("")

    # Spread by price level
    out.append("### Spread by DC Mid-Price Level\n")
    out.append("| Price Bucket | Count | Mean Comb Spread | Median | Std |")
    out.append("|---|---|---|---|---|")
    for bucket in ['<20c', '20-35c', '35-50c', '50-65c', '65-80c', '>80c']:
        vals = spread_by_price.get(bucket, [])
        if vals:
            a = np.array(vals)
            out.append(f"| {bucket} | {len(vals):,} | {a.mean():.2f}c | {np.median(a):.2f}c | {a.std():.2f}c |")
        else:
            out.append(f"| {bucket} | 0 | - | - | - |")
    out.append("")

    # Spread by event phase
    out.append("### Spread by Event Phase\n")
    out.append("| Phase | Count | Mean Comb Spread | Median | Std |")
    out.append("|---|---|---|---|---|")
    for label, vals in [('Before (>5s pre)', spread_before), ('During (-5s to +5s)', spread_during), ('After (>5s post)', spread_after)]:
        if vals:
            a = np.array(vals)
            out.append(f"| {label} | {len(vals):,} | {a.mean():.2f}c | {np.median(a):.2f}c | {a.std():.2f}c |")
    out.append("")

    return '\n'.join(out)


def analyze_depth(all_snaps):
    """Section 2: Depth Analysis."""
    out = []
    out.append("## 2. Depth Analysis\n")

    # Collect depth at each level
    dc_bid_sizes = {i: [] for i in range(1, 6)}
    dc_ask_sizes = {i: [] for i in range(1, 6)}
    mi_bid_sizes = {i: [] for i in range(1, 6)}
    mi_ask_sizes = {i: [] for i in range(1, 6)}

    # Depth by event phase
    depth_by_phase = {'before': [], 'during': [], 'after': []}

    for snap in all_snaps:
        for lvl in range(1, 6):
            for team, prefix in [('dc', dc_bid_sizes), ('mi', mi_bid_sizes)]:
                v = snap.get(f'{team}_bid{lvl}_s')
                if v is not None:
                    prefix[lvl].append(v)
            for team, prefix in [('dc', dc_ask_sizes), ('mi', mi_ask_sizes)]:
                v = snap.get(f'{team}_ask{lvl}_s')
                if v is not None:
                    prefix[lvl].append(v)

        # Total depth at L1 for phase analysis
        ms = snap.get('ms')
        dc_b1s = snap.get('dc_bid1_s')
        dc_a1s = snap.get('dc_ask1_s')
        mi_b1s = snap.get('mi_bid1_s')
        mi_a1s = snap.get('mi_ask1_s')
        if all(v is not None for v in [dc_b1s, dc_a1s, mi_b1s, mi_a1s]) and ms is not None:
            total_l1 = dc_b1s + dc_a1s + mi_b1s + mi_a1s
            if ms < -5000:
                depth_by_phase['before'].append(total_l1)
            elif ms <= 5000:
                depth_by_phase['during'].append(total_l1)
            else:
                depth_by_phase['after'].append(total_l1)

    # Average depth at L1
    out.append("### Average Depth at Best Bid/Ask\n")
    out.append("| Level | DC Bid Size | DC Ask Size | MI Bid Size | MI Ask Size |")
    out.append("|---|---|---|---|---|")
    for lvl in range(1, 6):
        dc_b = np.mean(dc_bid_sizes[lvl]) if dc_bid_sizes[lvl] else 0
        dc_a = np.mean(dc_ask_sizes[lvl]) if dc_ask_sizes[lvl] else 0
        mi_b = np.mean(mi_bid_sizes[lvl]) if mi_bid_sizes[lvl] else 0
        mi_a = np.mean(mi_ask_sizes[lvl]) if mi_ask_sizes[lvl] else 0
        out.append(f"| L{lvl} | {dc_b:,.0f} | {dc_a:,.0f} | {mi_b:,.0f} | {mi_a:,.0f} |")
    out.append("")

    # Median depth
    out.append("### Median Depth at Best Bid/Ask\n")
    out.append("| Level | DC Bid Size | DC Ask Size | MI Bid Size | MI Ask Size |")
    out.append("|---|---|---|---|---|")
    for lvl in range(1, 6):
        dc_b = np.median(dc_bid_sizes[lvl]) if dc_bid_sizes[lvl] else 0
        dc_a = np.median(dc_ask_sizes[lvl]) if dc_ask_sizes[lvl] else 0
        mi_b = np.median(mi_bid_sizes[lvl]) if mi_bid_sizes[lvl] else 0
        mi_a = np.median(mi_ask_sizes[lvl]) if mi_ask_sizes[lvl] else 0
        out.append(f"| L{lvl} | {dc_b:,.0f} | {dc_a:,.0f} | {mi_b:,.0f} | {mi_a:,.0f} |")
    out.append("")

    # Depth by event phase
    out.append("### Total L1 Depth by Event Phase\n")
    out.append("| Phase | Mean Total L1 | Median | Min | Max |")
    out.append("|---|---|---|---|---|")
    for label, key in [('Before (>5s pre)', 'before'), ('During (-5s to +5s)', 'during'), ('After (>5s post)', 'after')]:
        vals = depth_by_phase[key]
        if vals:
            a = np.array(vals)
            out.append(f"| {label} | {a.mean():,.0f} | {np.median(a):,.0f} | {a.min():,.0f} | {a.max():,.0f} |")
    out.append("")

    # Liquidity vacuum detection
    out.append("### Liquidity Vacuum Detection\n")
    out.append("Snapshots where DC_bid1_size OR MI_bid1_size < 100 tokens:\n")
    vacuum_count = 0
    total_valid = 0
    for snap in all_snaps:
        dc_b1s = snap.get('dc_bid1_s')
        mi_b1s = snap.get('mi_bid1_s')
        dc_a1s = snap.get('dc_ask1_s')
        mi_a1s = snap.get('mi_ask1_s')
        if dc_b1s is not None and mi_b1s is not None:
            total_valid += 1
            if dc_b1s < 100 or mi_b1s < 100 or (dc_a1s is not None and dc_a1s < 100) or (mi_a1s is not None and mi_a1s < 100):
                vacuum_count += 1
    out.append(f"- Vacuum snapshots (any L1 < 100): {vacuum_count:,} / {total_valid:,} ({vacuum_count/max(total_valid,1)*100:.1f}%)\n")

    # Also check < 500
    vac500 = sum(1 for s in all_snaps if s.get('dc_bid1_s') is not None and (
        (s['dc_bid1_s'] < 500) or ((s.get('mi_bid1_s') or 999999) < 500) or
        ((s.get('dc_ask1_s') or 999999) < 500) or ((s.get('mi_ask1_s') or 999999) < 500)))
    out.append(f"- Vacuum snapshots (any L1 < 500): {vac500:,} / {total_valid:,} ({vac500/max(total_valid,1)*100:.1f}%)\n")

    # Depth recovery: look at depth at ms=0, ms=5000, ms=10000, etc.
    out.append("### Depth Recovery Timeline (average total L1 depth by ms_from_event)\n")
    depth_by_ms_bucket = defaultdict(list)
    for snap in all_snaps:
        ms = snap.get('ms')
        dc_b1s = snap.get('dc_bid1_s')
        dc_a1s = snap.get('dc_ask1_s')
        mi_b1s = snap.get('mi_bid1_s')
        mi_a1s = snap.get('mi_ask1_s')
        if ms is not None and all(v is not None for v in [dc_b1s, dc_a1s, mi_b1s, mi_a1s]):
            total = dc_b1s + dc_a1s + mi_b1s + mi_a1s
            # Bucket by 5-second intervals
            bucket_sec = int(ms // 5000) * 5
            if -60 <= bucket_sec <= 120:
                depth_by_ms_bucket[bucket_sec].append(total)

    out.append("| Seconds from Event | Mean Total L1 | Count |")
    out.append("|---|---|---|")
    for sec in sorted(depth_by_ms_bucket.keys()):
        vals = depth_by_ms_bucket[sec]
        out.append(f"| {sec:+d}s to {sec+5:+d}s | {np.mean(vals):,.0f} | {len(vals):,} |")
    out.append("")

    # Which levels have most liquidity
    out.append("### Liquidity Distribution Across Levels (% of total depth)\n")
    out.append("| Side | L1 % | L2 % | L3 % | L4 % | L5 % |")
    out.append("|---|---|---|---|---|---|")
    for team_label, bid_sizes, ask_sizes in [
        ('DC Bid', dc_bid_sizes, None), ('DC Ask', None, dc_ask_sizes),
        ('MI Bid', mi_bid_sizes, None), ('MI Ask', None, mi_ask_sizes)]:
        sizes = bid_sizes if bid_sizes else ask_sizes
        means = [np.mean(sizes[i]) if sizes[i] else 0 for i in range(1, 6)]
        total = sum(means)
        if total > 0:
            pcts = [m / total * 100 for m in means]
            out.append(f"| {team_label} | {pcts[0]:.1f}% | {pcts[1]:.1f}% | {pcts[2]:.1f}% | {pcts[3]:.1f}% | {pcts[4]:.1f}% |")
    out.append("")

    return '\n'.join(out)


def analyze_complementary_pricing(all_snaps):
    """Section 3: Complementary Pricing Analysis."""
    out = []
    out.append("## 3. Complementary Pricing Analysis\n")

    cross_ba_1 = []  # DC_bid + MI_ask
    cross_ba_2 = []  # DC_ask + MI_bid
    combined_asks = []  # DC_ask + MI_ask
    combined_bids = []  # DC_bid + MI_bid

    buy_both_arb_snaps = []  # combined_asks < 1.00
    sell_both_arb_snaps = []  # combined_bids > 1.00

    for snap in all_snaps:
        dc_b = snap.get('dc_bid1_p')
        dc_a = snap.get('dc_ask1_p')
        mi_b = snap.get('mi_bid1_p')
        mi_a = snap.get('mi_ask1_p')

        if all(v is not None for v in [dc_b, dc_a, mi_b, mi_a]):
            cb1 = round(dc_b + mi_a, 4)
            cb2 = round(dc_a + mi_b, 4)
            ca = round(dc_a + mi_a, 4)
            cb = round(dc_b + mi_b, 4)

            cross_ba_1.append(cb1)
            cross_ba_2.append(cb2)
            combined_asks.append(ca)
            combined_bids.append(cb)

            if ca < 1.0:
                buy_both_arb_snaps.append({
                    'ms': snap.get('ms'),
                    'dc_a': dc_a, 'mi_a': mi_a, 'sum': ca,
                    'dc_a_s': snap.get('dc_ask1_s'), 'mi_a_s': snap.get('mi_ask1_s')
                })
            if cb > 1.0:
                sell_both_arb_snaps.append({
                    'ms': snap.get('ms'),
                    'dc_b': dc_b, 'mi_b': mi_b, 'sum': cb,
                    'dc_b_s': snap.get('dc_bid1_s'), 'mi_b_s': snap.get('mi_bid1_s')
                })

    ca_arr = np.array(combined_asks)
    cb_arr = np.array(combined_bids)
    cross1 = np.array(cross_ba_1)
    cross2 = np.array(cross_ba_2)

    out.append("### Cross-Side Sums (should be <= $1.00 for no arb)\n")
    out.append(f"- DC_bid + MI_ask: Mean={cross1.mean():.4f}, Min={cross1.min():.4f}, Max={cross1.max():.4f}")
    out.append(f"  - > $1.00: {np.sum(cross1 > 1.0):,} ({np.sum(cross1 > 1.0)/len(cross1)*100:.2f}%)")
    out.append(f"- DC_ask + MI_bid: Mean={cross2.mean():.4f}, Min={cross2.min():.4f}, Max={cross2.max():.4f}")
    out.append(f"  - > $1.00: {np.sum(cross2 > 1.0):,} ({np.sum(cross2 > 1.0)/len(cross2)*100:.2f}%)\n")

    out.append("### Combined Asks (DC_ask + MI_ask) -- maker profit margin\n")
    out.append(f"- Mean: {ca_arr.mean():.4f} | Median: {np.median(ca_arr):.4f} | Min: {ca_arr.min():.4f} | Max: {ca_arr.max():.4f}")
    out.append(f"- Std: {ca_arr.std():.4f}\n")
    out.append("Distribution:")
    for thresh in [0.98, 0.99, 1.00, 1.01, 1.02, 1.03, 1.04, 1.05, 1.06, 1.08, 1.10]:
        pct_leq = np.sum(ca_arr <= thresh) / len(ca_arr) * 100
        out.append(f"  - <= ${thresh:.2f}: {pct_leq:.1f}%")
    out.append("")

    out.append("### Combined Bids (DC_bid + MI_bid) -- should be < $1.00\n")
    out.append(f"- Mean: {cb_arr.mean():.4f} | Median: {np.median(cb_arr):.4f} | Min: {cb_arr.min():.4f} | Max: {cb_arr.max():.4f}\n")
    out.append("Distribution:")
    for thresh in [0.94, 0.95, 0.96, 0.97, 0.98, 0.99, 1.00]:
        pct_leq = np.sum(cb_arr <= thresh) / len(cb_arr) * 100
        out.append(f"  - <= ${thresh:.2f}: {pct_leq:.1f}%")
    out.append("")

    # Arb windows
    out.append("### Arbitrage Windows\n")
    out.append(f"**Buy-both-merge arb** (combined_asks < $1.00): {len(buy_both_arb_snaps):,} snapshots ({len(buy_both_arb_snaps)/max(len(combined_asks),1)*100:.2f}%)\n")
    if buy_both_arb_snaps:
        out.append("First 10 occurrences:")
        out.append("| ms_from_event | DC_ask | MI_ask | Sum | DC_ask_size | MI_ask_size |")
        out.append("|---|---|---|---|---|---|")
        for s in buy_both_arb_snaps[:10]:
            out.append(f"| {s['ms']} | {s['dc_a']:.2f} | {s['mi_a']:.2f} | {s['sum']:.4f} | {s.get('dc_a_s', '?')} | {s.get('mi_a_s', '?')} |")
        out.append("")

    out.append(f"**Sell-both arb** (combined_bids > $1.00): {len(sell_both_arb_snaps):,} snapshots ({len(sell_both_arb_snaps)/max(len(combined_bids),1)*100:.2f}%)\n")
    if sell_both_arb_snaps:
        out.append("First 10 occurrences:")
        out.append("| ms_from_event | DC_bid | MI_bid | Sum | DC_bid_size | MI_bid_size |")
        out.append("|---|---|---|---|---|---|")
        for s in sell_both_arb_snaps[:10]:
            out.append(f"| {s['ms']} | {s['dc_b']:.2f} | {s['mi_b']:.2f} | {s['sum']:.4f} | {s.get('dc_b_s', '?')} | {s.get('mi_b_s', '?')} |")
        out.append("")

    # Overround analysis
    out.append("### Overround (combined_asks - $1.00) -- the vigorish\n")
    overround = ca_arr - 1.0
    out.append(f"- Mean overround: {overround.mean()*100:.2f}c")
    out.append(f"- Median overround: {np.median(overround)*100:.2f}c")
    out.append(f"- This is the cost to buy both sides (market's edge)\n")

    return '\n'.join(out)


def analyze_mm_viability(all_snaps):
    """Section 4: Market Making Viability."""
    out = []
    out.append("## 4. Market Making Viability\n")

    total = 0
    ca_ge_102 = 0
    ca_ge_101 = 0
    spread_ge_1 = 0
    spread_ge_2 = 0

    dc_b1_sizes = []
    dc_a1_sizes = []
    mi_b1_sizes = []
    mi_a1_sizes = []

    for snap in all_snaps:
        dc_b = snap.get('dc_bid1_p')
        dc_a = snap.get('dc_ask1_p')
        mi_b = snap.get('mi_bid1_p')
        mi_a = snap.get('mi_ask1_p')

        if all(v is not None for v in [dc_b, dc_a, mi_b, mi_a]):
            total += 1
            ca = dc_a + mi_a
            cb = dc_b + mi_b
            dc_sp = dc_a - dc_b
            mi_sp = mi_a - mi_b

            if ca >= 1.02:
                ca_ge_102 += 1
            if ca >= 1.01:
                ca_ge_101 += 1
            if dc_sp >= 0.01 or mi_sp >= 0.01:
                spread_ge_1 += 1
            if dc_sp >= 0.02 or mi_sp >= 0.02:
                spread_ge_2 += 1

            if snap.get('dc_bid1_s') is not None:
                dc_b1_sizes.append(snap['dc_bid1_s'])
            if snap.get('dc_ask1_s') is not None:
                dc_a1_sizes.append(snap['dc_ask1_s'])
            if snap.get('mi_bid1_s') is not None:
                mi_b1_sizes.append(snap['mi_bid1_s'])
            if snap.get('mi_ask1_s') is not None:
                mi_a1_sizes.append(snap['mi_ask1_s'])

    out.append("### Spread-Based Opportunity\n")
    out.append(f"- Total valid snapshots: {total:,}")
    out.append(f"- Combined asks >= $1.02: {ca_ge_102:,} ({ca_ge_102/max(total,1)*100:.1f}%)")
    out.append(f"- Combined asks >= $1.01: {ca_ge_101:,} ({ca_ge_101/max(total,1)*100:.1f}%)")
    out.append(f"- Either side spread >= 1c: {spread_ge_1:,} ({spread_ge_1/max(total,1)*100:.1f}%)")
    out.append(f"- Either side spread >= 2c: {spread_ge_2:,} ({spread_ge_2/max(total,1)*100:.1f}%)\n")

    out.append("### Available Depth for Quoting\n")
    out.append(f"- DC bid1 size: Mean={np.mean(dc_b1_sizes):,.0f}, Median={np.median(dc_b1_sizes):,.0f}, P10={np.percentile(dc_b1_sizes,10):,.0f}, P90={np.percentile(dc_b1_sizes,90):,.0f}")
    out.append(f"- DC ask1 size: Mean={np.mean(dc_a1_sizes):,.0f}, Median={np.median(dc_a1_sizes):,.0f}, P10={np.percentile(dc_a1_sizes,10):,.0f}, P90={np.percentile(dc_a1_sizes,90):,.0f}")
    out.append(f"- MI bid1 size: Mean={np.mean(mi_b1_sizes):,.0f}, Median={np.median(mi_b1_sizes):,.0f}, P10={np.percentile(mi_b1_sizes,10):,.0f}, P90={np.percentile(mi_b1_sizes,90):,.0f}")
    out.append(f"- MI ask1 size: Mean={np.mean(mi_a1_sizes):,.0f}, Median={np.median(mi_a1_sizes):,.0f}, P10={np.percentile(mi_a1_sizes,10):,.0f}, P90={np.percentile(mi_a1_sizes,90):,.0f}\n")

    # Taker flow estimation: size changes between consecutive snapshots
    out.append("### Taker Flow Estimation (size changes at bid1/ask1)\n")
    out.append("Approximated by absolute change in L1 size between consecutive snapshots.\n")
    dc_b1_changes = []
    dc_a1_changes = []
    mi_b1_changes = []
    mi_a1_changes = []
    prev = None
    for snap in all_snaps:
        if prev is not None:
            for key, changes in [('dc_bid1_s', dc_b1_changes), ('dc_ask1_s', dc_a1_changes),
                                  ('mi_bid1_s', mi_b1_changes), ('mi_ask1_s', mi_a1_changes)]:
                if snap.get(key) is not None and prev.get(key) is not None:
                    # Only count if price didn't change (same level)
                    price_key = key.replace('_s', '_p')
                    if snap.get(price_key) == prev.get(price_key):
                        delta = abs(snap[key] - prev[key])
                        if delta > 0:
                            changes.append(delta)
        prev = snap

    for label, changes in [('DC bid1', dc_b1_changes), ('DC ask1', dc_a1_changes),
                            ('MI bid1', mi_b1_changes), ('MI ask1', mi_a1_changes)]:
        if changes:
            a = np.array(changes)
            out.append(f"- {label}: {len(changes):,} changes, mean={a.mean():,.0f}, median={np.median(a):,.0f}, P90={np.percentile(a,90):,.0f}")
    out.append("")

    # Practical MM assessment
    out.append("### Practical Assessment\n")
    min_depth = min(np.median(dc_b1_sizes), np.median(dc_a1_sizes), np.median(mi_b1_sizes), np.median(mi_a1_sizes))
    out.append(f"- Minimum median L1 depth across sides: {min_depth:,.0f} tokens")
    out.append(f"- Safe quoting size (10% of min median): {min_depth*0.1:,.0f} tokens")
    out.append(f"- At 2c overround, profit per roundtrip: ~2c per token pair")
    out.append(f"- If quoting {min_depth*0.1:,.0f} tokens at 2c edge: ${min_depth*0.1*0.02:,.0f} per fill\n")

    return '\n'.join(out)


def analyze_inter_event(sheets_data):
    """Section 5: Inter-Event Analysis."""
    out = []
    out.append("## 5. Inter-Event Analysis\n")

    # Extract pre-event state from each sheet
    events = []
    for sd in sheets_data:
        snaps = sd['snapshots']
        if not snaps:
            continue

        # Find snapshot closest to ms=0
        event_snap = None
        pre_snap = None
        min_abs_ms = float('inf')
        for s in snaps:
            ms = s.get('ms')
            if ms is not None and abs(ms) < min_abs_ms and s.get('dc_bid1_p') is not None:
                min_abs_ms = abs(ms)
                event_snap = s

        # Find last snapshot before event (ms < -1000)
        for s in snaps:
            ms = s.get('ms')
            if ms is not None and -10000 < ms < -1000 and s.get('dc_bid1_p') is not None:
                pre_snap = s

        if event_snap:
            events.append({
                'name': sd['name'],
                'event_type': sd['event_type'],
                'score': sd['score'],
                'event_time': sd['event_time'],
                'event_snap': event_snap,
                'pre_snap': pre_snap,
                'n_snaps': len(snaps),
            })

    out.append(f"Total events: {len(events)}\n")

    # Event type breakdown
    event_types = defaultdict(int)
    for e in events:
        event_types[e['event_type']] += 1
    out.append("### Event Type Distribution\n")
    for et, cnt in sorted(event_types.items(), key=lambda x: -x[1]):
        out.append(f"- {et}: {cnt}")
    out.append("")

    # Price drift between events
    out.append("### Price at Each Event (DC bid1)\n")
    out.append("| Event | Type | Score | DC_bid | DC_ask | MI_bid | MI_ask | Comb Ask |")
    out.append("|---|---|---|---|---|---|---|---|")
    for e in events:
        s = e['event_snap']
        ca = (s.get('dc_ask1_p', 0) or 0) + (s.get('mi_ask1_p', 0) or 0)
        out.append(f"| {e['name']} | {e['event_type']} | {e['score']} | "
                   f"{s.get('dc_bid1_p', '?')} | {s.get('dc_ask1_p', '?')} | "
                   f"{s.get('mi_bid1_p', '?')} | {s.get('mi_ask1_p', '?')} | {ca:.2f} |")
    out.append("")

    # Time between events (parse from sheet names)
    out.append("### Time Between Events\n")
    # Sheet name format: E1_4_153740_762 -> time is 15:37:40.762
    import re
    event_times_ms = []
    for e in events:
        m = re.match(r'E\d+_\w+_(\d{2})(\d{2})(\d{2})_(\d+)', e['name'])
        if m:
            h, mi, s, frac = m.groups()
            total_ms = int(h) * 3600000 + int(mi) * 60000 + int(s) * 1000 + int(frac)
            event_times_ms.append(total_ms)

    if len(event_times_ms) > 1:
        gaps = []
        for i in range(1, len(event_times_ms)):
            gap_s = (event_times_ms[i] - event_times_ms[i-1]) / 1000
            gaps.append(gap_s)
        gaps_arr = np.array(gaps)
        out.append(f"- Mean gap: {gaps_arr.mean():.0f}s ({gaps_arr.mean()/60:.1f}min)")
        out.append(f"- Median gap: {np.median(gaps_arr):.0f}s ({np.median(gaps_arr)/60:.1f}min)")
        out.append(f"- Min gap: {gaps_arr.min():.0f}s | Max gap: {gaps_arr.max():.0f}s")
        out.append(f"- Total match duration: {(event_times_ms[-1] - event_times_ms[0])/1000/60:.0f}min\n")

        out.append("Gap distribution:")
        for thresh in [30, 60, 120, 180, 300, 600]:
            pct = np.sum(gaps_arr <= thresh) / len(gaps_arr) * 100
            out.append(f"  - <= {thresh}s: {pct:.0f}%")
    out.append("")

    # Book state right before each event
    out.append("### Pre-Event Book State (last snapshot ~5s before event)\n")
    pre_depths = []
    pre_spreads = []
    for e in events:
        s = e.get('pre_snap')
        if s and s.get('dc_bid1_s') is not None and s.get('dc_ask1_p') is not None and s.get('dc_bid1_p') is not None:
            total = (s.get('dc_bid1_s', 0) or 0) + (s.get('dc_ask1_s', 0) or 0) + \
                    (s.get('mi_bid1_s', 0) or 0) + (s.get('mi_ask1_s', 0) or 0)
            pre_depths.append(total)
            dc_sp = (s['dc_ask1_p'] - s['dc_bid1_p']) * 100
            pre_spreads.append(dc_sp)

    if pre_depths:
        d = np.array(pre_depths)
        sp = np.array(pre_spreads)
        out.append(f"- Pre-event total L1 depth: Mean={d.mean():,.0f}, Median={np.median(d):,.0f}")
        out.append(f"- Pre-event DC spread: Mean={sp.mean():.1f}c, Median={np.median(sp):.1f}c\n")

    return '\n'.join(out)


def main():
    print("Loading data...")
    sheets_data = load_all_data()

    # Flatten all snapshots for global analysis
    all_snaps = []
    for sd in sheets_data:
        for snap in sd['snapshots']:
            all_snaps.append(snap)

    print(f"Loaded {len(sheets_data)} sheets, {len(all_snaps):,} total snapshots")

    # Run all analyses
    print("Analyzing spreads...")
    s1 = analyze_spreads(all_snaps, sheets_data)

    print("Analyzing depth...")
    s2 = analyze_depth(all_snaps)

    print("Analyzing complementary pricing...")
    s3 = analyze_complementary_pricing(all_snaps)

    print("Analyzing MM viability...")
    s4 = analyze_mm_viability(all_snaps)

    print("Analyzing inter-event patterns...")
    s5 = analyze_inter_event(sheets_data)

    # Compose output
    report = f"""# Spread & Depth Analysis: DC vs MI (IPL, April 4 2026)

**Source:** `20260404_153643_event_book_cricipl-del-mum-2026-04-04.xlsx`
**Events:** {len(sheets_data)} sheets, {len(all_snaps):,} total book snapshots

---

{s1}
---

{s2}
---

{s3}
---

{s4}
---

{s5}
"""

    with open(OUTPUT, 'w') as f:
        f.write(report)

    print(f"\nResults written to {OUTPUT}")
    print("\n" + "="*60)
    print("KEY FINDINGS SUMMARY")
    print("="*60)
    # Print a condensed version
    print(report[:3000])


if __name__ == '__main__':
    main()
