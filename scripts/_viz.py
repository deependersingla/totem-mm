import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if len(sys.argv) < 3:
    print("Usage: python _viz.py <events.json> <slug>")
    sys.exit(1)

EVENTS_FILE = sys.argv[1]
SLUG = sys.argv[2]
DATA_DIR = Path(__file__).parent.parent / "data" / SLUG
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT = str(DATA_DIR / f"match_events_report_{SLUG}.xlsx")

with open(EVENTS_FILE) as f:
    events = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "Match Events"

# Colors
HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
WICKET_FILL = PatternFill("solid", fgColor="FEE2E2")
FOUR_FILL = PatternFill("solid", fgColor="DCFCE7")
SIX_FILL = PatternFill("solid", fgColor="FEF3C7")
BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
)
INNINGS_FILL = PatternFill("solid", fgColor="DBEAFE")

headers = [
    "#",
    "Score Feed Time",
    "Market Move Time",
    "Latency (s)",
    "Over",
    "Event",
    "Innings",
    "Score Before",
    "Score After",
    "Side",
    "NZ Price Before",
    "NZ Price After",
    "SA Price Before",
    "SA Price After",
    "Ticks",
    "Direction",
    "Move Duration",
]

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 35

prev_innings = None
row = 2

for idx, e in enumerate(events):
    mm = e["market_movement"]
    if not isinstance(mm, dict):
        continue

    innings = e["innings"]

    # Insert innings separator
    if innings != prev_innings:
        if prev_innings is not None:
            # blank separator row
            row += 1
        # Innings header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        c = ws.cell(row=row, column=1, value=f"  {innings.upper()}")
        c.font = Font(bold=True, size=11, color="1E40AF", name="Calibri")
        c.fill = INNINGS_FILL
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 28
        for ci in range(1, len(headers) + 1):
            ws.cell(row=row, column=ci).fill = INNINGS_FILL
        row += 1
        prev_innings = innings

    nz_before = mm["price_before"]
    nz_after = mm["price_after"]
    sa_before = round(1.0 - nz_before, 4)
    sa_after = round(1.0 - nz_after, 4)

    # Move duration
    from datetime import datetime
    try:
        t1 = datetime.strptime(mm["market_move_start"], "%H:%M:%S")
        t2 = datetime.strptime(mm["market_move_end"], "%H:%M:%S")
        dur = int((t2 - t1).total_seconds())
        dur_str = f"{dur}s"
    except:
        dur_str = ""

    score_time = e["ist"][11:19]

    values = [
        idx + 1,
        score_time,
        mm["market_move_start"],
        mm["latency_sec"],
        e["over"],
        e["event"],
        innings.replace(" batting", ""),
        e["score_before"],
        e["score_after"],
        e["side"],
        nz_before,
        nz_after,
        sa_before,
        sa_after,
        mm["ticks"],
        mm["direction"],
        dur_str,
    ]

    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

        # Row color by event type
        if e["event"] == "WICKET":
            c.fill = WICKET_FILL
        elif e["event"] == "FOUR":
            c.fill = FOUR_FILL
        elif e["event"] == "SIX":
            c.fill = SIX_FILL

    # Bold the ticks column
    ws.cell(row=row, column=15).font = Font(size=11, bold=True, name="Calibri")

    # Color the latency: red if > 120s, orange if > 90s
    lat = mm["latency_sec"]
    lat_cell = ws.cell(row=row, column=4)
    if lat > 120:
        lat_cell.font = Font(size=10, color="DC2626", bold=True, name="Calibri")
    elif lat > 90:
        lat_cell.font = Font(size=10, color="D97706", bold=True, name="Calibri")

    ws.row_dimensions[row].height = 22
    row += 1

# Summary row
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
total_events = len([e for e in events if isinstance(e["market_movement"], dict)])
wickets = len([e for e in events if e["event"] == "WICKET" and isinstance(e["market_movement"], dict)])
fours = len([e for e in events if e["event"] == "FOUR" and isinstance(e["market_movement"], dict)])
sixes = len([e for e in events if e["event"] == "SIX" and isinstance(e["market_movement"], dict)])
avg_ticks = sum(e["market_movement"]["ticks"] for e in events if isinstance(e["market_movement"], dict)) / total_events
avg_lat = sum(e["market_movement"]["latency_sec"] for e in events if isinstance(e["market_movement"], dict)) / total_events

c = ws.cell(row=row, column=1, value=f"SUMMARY: {total_events} events ({wickets}W, {fours}×4, {sixes}×6)")
c.font = Font(bold=True, size=11, name="Calibri")
row += 1
c = ws.cell(row=row, column=1, value=f"Avg ticks: {avg_ticks:.1f} | Avg latency: {avg_lat:.0f}s")
c.font = Font(size=10, color="6B7280", name="Calibri")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

# Breakdown by event type
row += 1
for evt_type in ["WICKET", "FOUR", "SIX"]:
    subset = [e for e in events if e["event"] == evt_type and isinstance(e["market_movement"], dict)]
    if not subset:
        continue
    avg_t = sum(e["market_movement"]["ticks"] for e in subset) / len(subset)
    avg_l = sum(e["market_movement"]["latency_sec"] for e in subset) / len(subset)
    max_t = max(e["market_movement"]["ticks"] for e in subset)
    c = ws.cell(row=row, column=1, value=f"  {evt_type}: {len(subset)} events | avg {avg_t:.1f} ticks (max {max_t}) | avg latency {avg_l:.0f}s")
    c.font = Font(size=10, name="Calibri")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

# Column widths
widths = {
    1: 5, 2: 14, 3: 16, 4: 11, 5: 7, 6: 8, 7: 8,
    8: 12, 9: 12, 10: 6, 11: 13, 12: 13, 13: 13, 14: 13,
    15: 6, 16: 12, 17: 12,
}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Freeze top row
ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"{total_events} events | {wickets}W {fours}×4 {sixes}×6 | avg {avg_ticks:.1f} ticks | avg latency {avg_lat:.0f}s")
