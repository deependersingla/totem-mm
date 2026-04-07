#!/usr/bin/env python3
"""
Event Impact Analysis: DC vs MI, IPL, April 4, 2026
Analyzes orderbook microstructure around 53 ball events (4/6/W)
"""
import openpyxl
import statistics
from collections import defaultdict

XLSX = '/Users/sobhagyaxd/DeepWork/totem-mm/captures/20260404_153643_event_book_cricipl-del-mum-2026-04-04.xlsx'
OUTPUT = '/Users/sobhagyaxd/DeepWork/totem-mm/polymarket-taker/docs/event_impact_analysis.md'

# Column indices (0-based, after IST and ms_from_event)
# Row 3 headers: IST, ms_from_event, DC_bid1_price, DC_bid1_size, ..., DC_ask5_size, MI_bid1_price, ...
COL_MS = 1
COL_DC_BID1_P = 2;  COL_DC_BID1_S = 3
COL_DC_BID2_P = 4;  COL_DC_BID2_S = 5
COL_DC_BID3_P = 6;  COL_DC_BID3_S = 7
COL_DC_BID4_P = 8;  COL_DC_BID4_S = 9
COL_DC_BID5_P = 10; COL_DC_BID5_S = 11
COL_DC_ASK1_P = 12; COL_DC_ASK1_S = 13
COL_DC_ASK2_P = 14; COL_DC_ASK2_S = 15
COL_DC_ASK3_P = 16; COL_DC_ASK3_S = 17
COL_DC_ASK4_P = 18; COL_DC_ASK4_S = 19
COL_DC_ASK5_P = 20; COL_DC_ASK5_S = 21
COL_MI_BID1_P = 22; COL_MI_BID1_S = 23
COL_MI_BID2_P = 24; COL_MI_BID2_S = 25
COL_MI_BID3_P = 26; COL_MI_BID3_S = 27
COL_MI_BID4_P = 28; COL_MI_BID4_S = 29
COL_MI_BID5_P = 30; COL_MI_BID5_S = 31
COL_MI_ASK1_P = 32; COL_MI_ASK1_S = 33
COL_MI_ASK2_P = 34; COL_MI_ASK2_S = 35
COL_MI_ASK3_P = 36; COL_MI_ASK3_S = 37
COL_MI_ASK4_P = 38; COL_MI_ASK4_S = 39
COL_MI_ASK5_P = 40; COL_MI_ASK5_S = 41


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def load_sheet_data(ws):
    """Load all rows from a sheet into a list of dicts."""
    rows = []
    meta = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Parse metadata
            meta['event_type'] = str(row[0]).split(':')[1].strip() if row[0] else '?'
            meta['score'] = str(row[1]).split(':')[1].strip() if row[1] else '?'
            meta['time'] = str(row[2]).split(': ')[1].strip() if row[2] else '?'
            meta['snapshots'] = str(row[3]).split(':')[1].strip() if row[3] else '?'
            continue
        if i < 3:
            continue
        ms = row[COL_MS]
        if ms is None:
            continue
        dc_b1p = safe_float(row[COL_DC_BID1_P])
        dc_b1s = safe_float(row[COL_DC_BID1_S])
        dc_a1p = safe_float(row[COL_DC_ASK1_P])
        dc_a1s = safe_float(row[COL_DC_ASK1_S])
        mi_b1p = safe_float(row[COL_MI_BID1_P])
        mi_b1s = safe_float(row[COL_MI_BID1_S])
        mi_a1p = safe_float(row[COL_MI_ASK1_P])
        mi_a1s = safe_float(row[COL_MI_ASK1_S])

        if dc_b1p is None or dc_a1p is None:
            continue

        mid = (dc_b1p + dc_a1p) / 2.0
        spread = dc_a1p - dc_b1p

        rows.append({
            'ms': int(ms),
            'dc_b1p': dc_b1p, 'dc_b1s': dc_b1s or 0,
            'dc_a1p': dc_a1p, 'dc_a1s': dc_a1s or 0,
            'mi_b1p': mi_b1p, 'mi_b1s': mi_b1s or 0,
            'mi_a1p': mi_a1p, 'mi_a1s': mi_a1s or 0,
            'dc_mid': mid,
            'dc_spread': spread,
        })
    return meta, rows


def get_snapshots_in_range(rows, ms_lo, ms_hi):
    return [r for r in rows if ms_lo <= r['ms'] <= ms_hi]


def closest_to_zero(rows):
    valid = [r for r in rows if r['dc_b1p'] is not None]
    if not valid:
        return None
    return min(valid, key=lambda r: abs(r['ms']))


def avg_field(snaps, field):
    vals = [s[field] for s in snaps if s.get(field) is not None]
    return statistics.mean(vals) if vals else None


def find_equilibrium_time(rows, final_mid, move_cents):
    """Find ms from event when price stabilizes within 0.5c of final price.
    Only meaningful for events with actual price moves >= 0.5c."""
    if final_mid is None:
        return None
    if move_cents is not None and abs(move_cents) < 0.5:
        return None  # No meaningful move to measure equilibrium for
    threshold = 0.005  # 0.5 cents
    # Look at snapshots after event, limit to 60s
    post = [r for r in rows if 0 <= r['ms'] <= 60000]
    # Find last time price was outside threshold, then equilibrium is after that
    last_outside = -1
    for r in post:
        if r['dc_mid'] is not None and abs(r['dc_mid'] - final_mid) > threshold:
            last_outside = r['ms']
    if last_outside == -1:
        return 0  # Was always at equilibrium
    # Find first snapshot after last_outside
    for r in post:
        if r['ms'] > last_outside:
            return r['ms']
    return None


def compute_move_pct_at_milestones(rows, total_move, mid_before):
    """At each milestone ms, what % of total move has happened?"""
    milestones = [0, 500, 1000, 1500, 2000, 2500, 3000]
    results = {}
    if total_move == 0 or mid_before is None:
        return {m: None for m in milestones}
    for target_ms in milestones:
        # Find snapshot closest to target_ms
        candidates = [r for r in rows if r['ms'] >= target_ms - 150 and r['ms'] <= target_ms + 300]
        if not candidates:
            results[target_ms] = None
            continue
        snap = min(candidates, key=lambda r: abs(r['ms'] - target_ms))
        if snap['dc_mid'] is not None:
            move_so_far = snap['dc_mid'] - mid_before
            pct = (move_so_far / total_move) * 100.0
            results[target_ms] = pct
        else:
            results[target_ms] = None
    return results


def check_arb(rows):
    """Check DC_bid1 + MI_ask1 and DC_ask1 + MI_bid1 for arb windows."""
    arb_windows = []
    for r in rows:
        if r['dc_b1p'] is not None and r['mi_a1p'] is not None:
            sum1 = round(r['dc_b1p'] + r['mi_a1p'], 4)
            if abs(sum1 - 1.0) > 0.001:
                arb_windows.append({
                    'ms': r['ms'],
                    'type': 'DC_bid+MI_ask',
                    'sum': sum1,
                    'deviation': round(sum1 - 1.0, 4)
                })
        if r['dc_a1p'] is not None and r['mi_b1p'] is not None:
            sum2 = round(r['dc_a1p'] + r['mi_b1p'], 4)
            if abs(sum2 - 1.0) > 0.001:
                arb_windows.append({
                    'ms': r['ms'],
                    'type': 'DC_ask+MI_bid',
                    'sum': sum2,
                    'deviation': round(sum2 - 1.0, 4)
                })
    return arb_windows


def detect_liquidity_vacuums(rows):
    """Detect when bid1 or ask1 size drops to near 0 during events."""
    vacuums = []
    # Focus on -1000 to +15000 window
    event_rows = [r for r in rows if -1000 <= r['ms'] <= 15000]
    in_vacuum_bid = False
    in_vacuum_ask = False
    vac_start_bid = None
    vac_start_ask = None

    for r in event_rows:
        # Bid vacuum: size < 10
        if r['dc_b1s'] is not None and r['dc_b1s'] < 10:
            if not in_vacuum_bid:
                in_vacuum_bid = True
                vac_start_bid = r['ms']
        else:
            if in_vacuum_bid:
                vacuums.append({
                    'side': 'bid',
                    'start_ms': vac_start_bid,
                    'end_ms': r['ms'],
                    'duration_ms': r['ms'] - vac_start_bid,
                    'refill_price': r['dc_b1p'],
                    'refill_size': r['dc_b1s']
                })
                in_vacuum_bid = False

        # Ask vacuum: size < 10
        if r['dc_a1s'] is not None and r['dc_a1s'] < 10:
            if not in_vacuum_ask:
                in_vacuum_ask = True
                vac_start_ask = r['ms']
        else:
            if in_vacuum_ask:
                vacuums.append({
                    'side': 'ask',
                    'start_ms': vac_start_ask,
                    'end_ms': r['ms'],
                    'duration_ms': r['ms'] - vac_start_ask,
                    'refill_price': r['dc_a1p'],
                    'refill_size': r['dc_a1s']
                })
                in_vacuum_ask = False

    # Handle vacuums that extend to end of window
    if in_vacuum_bid and vac_start_bid is not None:
        vacuums.append({
            'side': 'bid', 'start_ms': vac_start_bid,
            'end_ms': event_rows[-1]['ms'] if event_rows else None,
            'duration_ms': (event_rows[-1]['ms'] - vac_start_bid) if event_rows else None,
            'refill_price': None, 'refill_size': None
        })
    if in_vacuum_ask and vac_start_ask is not None:
        vacuums.append({
            'side': 'ask', 'start_ms': vac_start_ask,
            'end_ms': event_rows[-1]['ms'] if event_rows else None,
            'duration_ms': (event_rows[-1]['ms'] - vac_start_ask) if event_rows else None,
            'refill_price': None, 'refill_size': None
        })

    return vacuums


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    sheets = wb.sheetnames

    all_events = []
    all_arbs = []
    all_vacuums = []
    move_pcts_by_type = defaultdict(list)

    for sname in sheets:
        ws = wb[sname]
        meta, rows = load_sheet_data(ws)
        if not rows:
            continue

        event_type = meta['event_type']
        event_num = sname.split('_')[0]

        # Before window: -5000 to -1000
        before = get_snapshots_in_range(rows, -5000, -1000)
        # After window: +3000 to +10000
        after = get_snapshots_in_range(rows, 3000, 10000)
        # At event
        at_event = closest_to_zero(rows)

        mid_before = avg_field(before, 'dc_mid')
        mid_after = avg_field(after, 'dc_mid')
        mid_at_event = at_event['dc_mid'] if at_event else None

        spread_before = avg_field(before, 'dc_spread')
        spread_after = avg_field(after, 'dc_spread')

        depth_bid_before = avg_field(before, 'dc_b1s')
        depth_bid_after = avg_field(after, 'dc_b1s')
        depth_ask_before = avg_field(before, 'dc_a1s')
        depth_ask_after = avg_field(after, 'dc_a1s')

        move = None
        if mid_after is not None and mid_before is not None:
            move = mid_after - mid_before

        # Final price for equilibrium calc (average of +8000 to +10000)
        final_snaps = get_snapshots_in_range(rows, 8000, 10000)
        final_mid = avg_field(final_snaps, 'dc_mid')
        move_cents = round(move * 100, 2) if move is not None else None
        eq_time = find_equilibrium_time(rows, final_mid, move_cents)

        # Move percentage at milestones (only for moves >= 0.5c to avoid noise)
        if move is not None and abs(move) >= 0.005:
            mpct = compute_move_pct_at_milestones(rows, move, mid_before)
            move_pcts_by_type[event_type].append(mpct)

        # Arb check
        arbs = check_arb(rows)
        for a in arbs:
            a['event'] = sname
        all_arbs.extend(arbs)

        # Liquidity vacuums
        vacuums = detect_liquidity_vacuums(rows)
        for v in vacuums:
            v['event'] = sname
            v['event_type'] = event_type
        all_vacuums.extend(vacuums)

        ev = {
            'name': sname,
            'num': event_num,
            'type': event_type,
            'score': meta['score'],
            'time': meta['time'],
            'mid_before': mid_before,
            'mid_after': mid_after,
            'mid_at_event': mid_at_event,
            'move': move,
            'move_cents': round(move * 100, 2) if move is not None else None,
            'spread_before': spread_before,
            'spread_after': spread_after,
            'depth_bid_before': depth_bid_before,
            'depth_bid_after': depth_bid_after,
            'depth_ask_before': depth_ask_before,
            'depth_ask_after': depth_ask_after,
            'eq_time': eq_time,
        }
        all_events.append(ev)
        sb_str = f"{spread_before:.4f}" if spread_before is not None else "None"
        sa_str = f"{spread_after:.4f}" if spread_after is not None else "None"
        print(f"  {sname}: type={event_type}, move={ev['move_cents']}c, spread_b={sb_str} spread_a={sa_str}")

    wb.close()

    # ----- AGGREGATE BY TYPE -----
    agg = {}
    for etype in ['4', '6', 'W']:
        evts = [e for e in all_events if e['type'] == etype]
        moves = [e['move_cents'] for e in evts if e['move_cents'] is not None]
        spreads_b = [e['spread_before'] for e in evts if e['spread_before'] is not None]
        spreads_a = [e['spread_after'] for e in evts if e['spread_after'] is not None]
        dbid_b = [e['depth_bid_before'] for e in evts if e['depth_bid_before'] is not None]
        dbid_a = [e['depth_bid_after'] for e in evts if e['depth_bid_after'] is not None]
        dask_b = [e['depth_ask_before'] for e in evts if e['depth_ask_before'] is not None]
        dask_a = [e['depth_ask_after'] for e in evts if e['depth_ask_after'] is not None]
        eq_times = [e['eq_time'] for e in evts if e['eq_time'] is not None]

        agg[etype] = {
            'count': len(evts),
            'avg_move': statistics.mean(moves) if moves else 0,
            'max_move': max(moves) if moves else 0,
            'min_move': min(moves) if moves else 0,
            'median_move': statistics.median(moves) if moves else 0,
            'abs_avg_move': statistics.mean([abs(m) for m in moves]) if moves else 0,
            'avg_spread_before': statistics.mean(spreads_b) if spreads_b else 0,
            'avg_spread_after': statistics.mean(spreads_a) if spreads_a else 0,
            'avg_depth_bid_before': statistics.mean(dbid_b) if dbid_b else 0,
            'avg_depth_bid_after': statistics.mean(dbid_a) if dbid_a else 0,
            'avg_depth_ask_before': statistics.mean(dask_b) if dask_b else 0,
            'avg_depth_ask_after': statistics.mean(dask_a) if dask_a else 0,
            'avg_eq_time': statistics.mean(eq_times) if eq_times else 0,
            'median_eq_time': statistics.median(eq_times) if eq_times else 0,
        }

    # ----- SPEED ANALYSIS -----
    speed_analysis = {}
    for etype in ['4', '6', 'W']:
        milestones = [0, 500, 1000, 1500, 2000, 2500, 3000]
        pcts = move_pcts_by_type.get(etype, [])
        if not pcts:
            speed_analysis[etype] = {m: None for m in milestones}
            continue
        avg_pcts = {}
        for m in milestones:
            vals = [p[m] for p in pcts if p.get(m) is not None]
            avg_pcts[m] = statistics.mean(vals) if vals else None
        speed_analysis[etype] = avg_pcts

    # ----- ARB ANALYSIS -----
    # Count unique arb events and summarize
    arb_events = set()
    arb_by_type = defaultdict(list)
    for a in all_arbs:
        arb_events.add(a['event'])
        arb_by_type[a['type']].append(a)

    # ----- VACUUM ANALYSIS -----
    vac_by_type = defaultdict(list)
    for v in all_vacuums:
        vac_by_type[v['event_type']].append(v)

    # ----- WRITE REPORT -----
    lines = []
    L = lines.append

    L("# Event Impact Analysis: DC vs MI (IPL, April 4, 2026)")
    L("")
    L("Match: Delhi Capitals vs Mumbai Indians")
    L("Date: April 4, 2026")
    L(f"Total events analyzed: {len(all_events)} (4s: {agg['4']['count']}, 6s: {agg['6']['count']}, Ws: {agg['W']['count']})")
    L("")

    # ===== SECTION 1: PER-EVENT TABLE =====
    L("## 1. Per-Event Breakdown")
    L("")
    L("| # | Type | Score | Time | Mid Before | Mid After | Move (c) | Spread Before | Spread After | Bid1 Depth Before | Bid1 Depth After | Ask1 Depth Before | Ask1 Depth After | Eq Time (ms) |")
    L("|---|------|-------|------|------------|-----------|----------|---------------|--------------|-------------------|------------------|-------------------|------------------|--------------|")
    for e in all_events:
        mb = f"{e['mid_before']:.4f}" if e['mid_before'] else '-'
        ma = f"{e['mid_after']:.4f}" if e['mid_after'] else '-'
        mc = f"{e['move_cents']:+.2f}" if e['move_cents'] is not None else '-'
        sb = f"{e['spread_before']:.4f}" if e['spread_before'] else '-'
        sa = f"{e['spread_after']:.4f}" if e['spread_after'] else '-'
        dbb = f"{e['depth_bid_before']:.0f}" if e['depth_bid_before'] else '-'
        dba = f"{e['depth_bid_after']:.0f}" if e['depth_bid_after'] else '-'
        dab = f"{e['depth_ask_before']:.0f}" if e['depth_ask_before'] else '-'
        daa = f"{e['depth_ask_after']:.0f}" if e['depth_ask_after'] else '-'
        eq = f"{e['eq_time']}" if e['eq_time'] is not None else '-'
        L(f"| {e['num']} | {e['type']} | {e['score']} | {e['time']} | {mb} | {ma} | {mc} | {sb} | {sa} | {dbb} | {dba} | {dab} | {daa} | {eq} |")
    L("")

    # ===== SECTION 2: AGGREGATE BY TYPE =====
    L("## 2. Aggregate Statistics by Event Type")
    L("")
    L("| Metric | 4s | 6s | Ws |")
    L("|--------|----|----|-----|")
    L(f"| Count | {agg['4']['count']} | {agg['6']['count']} | {agg['W']['count']} |")
    L(f"| Avg Move (c) | {agg['4']['avg_move']:+.2f} | {agg['6']['avg_move']:+.2f} | {agg['W']['avg_move']:+.2f} |")
    L(f"| Avg Abs Move (c) | {agg['4']['abs_avg_move']:.2f} | {agg['6']['abs_avg_move']:.2f} | {agg['W']['abs_avg_move']:.2f} |")
    L(f"| Median Move (c) | {agg['4']['median_move']:+.2f} | {agg['6']['median_move']:+.2f} | {agg['W']['median_move']:+.2f} |")
    L(f"| Max Move (c) | {agg['4']['max_move']:+.2f} | {agg['6']['max_move']:+.2f} | {agg['W']['max_move']:+.2f} |")
    L(f"| Min Move (c) | {agg['4']['min_move']:+.2f} | {agg['6']['min_move']:+.2f} | {agg['W']['min_move']:+.2f} |")
    L(f"| Avg Spread Before | {agg['4']['avg_spread_before']:.4f} | {agg['6']['avg_spread_before']:.4f} | {agg['W']['avg_spread_before']:.4f} |")
    L(f"| Avg Spread After | {agg['4']['avg_spread_after']:.4f} | {agg['6']['avg_spread_after']:.4f} | {agg['W']['avg_spread_after']:.4f} |")
    L(f"| Avg Bid1 Depth Before | {agg['4']['avg_depth_bid_before']:.0f} | {agg['6']['avg_depth_bid_before']:.0f} | {agg['W']['avg_depth_bid_before']:.0f} |")
    L(f"| Avg Bid1 Depth After | {agg['4']['avg_depth_bid_after']:.0f} | {agg['6']['avg_depth_bid_after']:.0f} | {agg['W']['avg_depth_bid_after']:.0f} |")
    L(f"| Avg Ask1 Depth Before | {agg['4']['avg_depth_ask_before']:.0f} | {agg['6']['avg_depth_ask_before']:.0f} | {agg['W']['avg_depth_ask_before']:.0f} |")
    L(f"| Avg Ask1 Depth After | {agg['4']['avg_depth_ask_after']:.0f} | {agg['6']['avg_depth_ask_after']:.0f} | {agg['W']['avg_depth_ask_after']:.0f} |")
    L(f"| Avg Eq Time (ms) | {agg['4']['avg_eq_time']:.0f} | {agg['6']['avg_eq_time']:.0f} | {agg['W']['avg_eq_time']:.0f} |")
    L(f"| Median Eq Time (ms) | {agg['4']['median_eq_time']:.0f} | {agg['6']['median_eq_time']:.0f} | {agg['W']['median_eq_time']:.0f} |")
    L("")

    # ===== SECTION 3: SPEED OF BOOK CHANGES =====
    L("## 3. Speed of Price Adjustment (% of Total Move at Each Milestone)")
    L("")
    L("This shows what fraction of the eventual price move has been realized at each point after the event.")
    L("Only events with moves >= 0.5c are included (to avoid noise from tick-size discretization).")
    L("")
    for etype in ['4', '6', 'W']:
        n = len(move_pcts_by_type.get(etype, []))
        L(f"- **{etype}s**: {n} events with moves >= 0.5c")
    L("")
    L("| ms after event | 4s | 6s | Ws |")
    L("|----------------|----|----|-----|")
    for m in [0, 500, 1000, 1500, 2000, 2500, 3000]:
        vals = []
        for etype in ['4', '6', 'W']:
            v = speed_analysis[etype].get(m)
            vals.append(f"{v:.1f}%" if v is not None else '-')
        L(f"| +{m}ms | {vals[0]} | {vals[1]} | {vals[2]} |")
    L("")
    L("**Interpretation**: Values > 100% indicate overshoot (price moved past equilibrium then reverted).")
    L("If 80%+ of the move happens by +500ms, you need sub-500ms refresh to capture edge.")
    L("")

    # ===== SECTION 4: A+B=1 CHECK =====
    L("## 4. Pair Pricing Check (DC + MI = $1.00)")
    L("")
    # Re-analyze arb data
    L(f"Total arb deviations found (|sum - 1.00| > 0.1c): **{len(all_arbs)}**")
    L(f"Events with at least one deviation: **{len(arb_events)}** / {len(all_events)}")
    L("")

    if all_arbs:
        # Group by type
        for atype in ['DC_bid+MI_ask', 'DC_ask+MI_bid']:
            arbs_t = [a for a in all_arbs if a['type'] == atype]
            if arbs_t:
                devs = [a['deviation'] for a in arbs_t]
                L(f"**{atype}**:")
                L(f"- Count: {len(arbs_t)}")
                L(f"- Avg deviation: {statistics.mean(devs):+.4f}")
                L(f"- Max deviation: {max(devs, key=abs):+.4f}")
                L(f"- Range: [{min(devs):+.4f}, {max(devs):+.4f}]")
                L("")
                # Show examples
                L(f"Sample arb snapshots ({atype}):")
                L("")
                L("| Event | ms | Sum | Deviation |")
                L("|-------|----|-----|-----------|")
                shown = set()
                for a in sorted(arbs_t, key=lambda x: abs(x['deviation']), reverse=True)[:20]:
                    key = (a['event'], a['ms'])
                    if key not in shown:
                        shown.add(key)
                        L(f"| {a['event']} | {a['ms']} | {a['sum']:.4f} | {a['deviation']:+.4f} |")
                L("")
    else:
        L("No deviations found -- DC_bid1 + MI_ask1 = 1.00 and DC_ask1 + MI_bid1 = 1.00 always hold.")
        L("")

    # ===== SECTION 5: LIQUIDITY VACUUMS =====
    L("## 5. Liquidity Vacuums")
    L("")
    L(f"Total vacuum events detected (bid1 or ask1 size < 10 shares): **{len(all_vacuums)}**")
    L("")

    if all_vacuums:
        for etype in ['4', '6', 'W']:
            vacs = [v for v in all_vacuums if v['event_type'] == etype]
            if vacs:
                bid_vacs = [v for v in vacs if v['side'] == 'bid']
                ask_vacs = [v for v in vacs if v['side'] == 'ask']
                durations = [v['duration_ms'] for v in vacs if v['duration_ms'] is not None]
                L(f"### {etype}s")
                L(f"- Bid vacuums: {len(bid_vacs)}")
                L(f"- Ask vacuums: {len(ask_vacs)}")
                if durations:
                    L(f"- Avg duration: {statistics.mean(durations):.0f}ms")
                    L(f"- Median duration: {statistics.median(durations):.0f}ms")
                    L(f"- Max duration: {max(durations)}ms")
                    L(f"- Min duration: {min(durations)}ms")
                L("")

                # Show specific vacuum instances
                L(f"| Event | Side | Start (ms) | End (ms) | Duration (ms) | Refill Price | Refill Size |")
                L(f"|-------|------|------------|----------|---------------|--------------|-------------|")
                for v in sorted(vacs, key=lambda x: x['duration_ms'] or 0, reverse=True)[:15]:
                    rp = f"{v['refill_price']}" if v['refill_price'] is not None else '-'
                    rs = f"{v['refill_size']:.0f}" if v['refill_size'] is not None else '-'
                    dur = str(v['duration_ms']) if v['duration_ms'] is not None else '-'
                    L(f"| {v['event']} | {v['side']} | {v['start_ms']} | {v['end_ms']} | {dur} | {rp} | {rs} |")
                L("")
    else:
        L("No liquidity vacuums detected (all bid1/ask1 sizes stayed >= 10 shares).")
        L("")

    # ===== SECTION 6: KEY FINDINGS =====
    L("## 6. Key Findings & Trading Implications")
    L("")

    # Find biggest movers
    sorted_by_move = sorted([e for e in all_events if e['move_cents'] is not None],
                           key=lambda x: abs(x['move_cents']), reverse=True)

    L("### Biggest Price Movers")
    L("")
    L("| # | Type | Score | Move (c) | Direction |")
    L("|---|------|-------|----------|-----------|")
    for e in sorted_by_move[:10]:
        direction = "DC favored" if e['move_cents'] > 0 else "MI favored" if e['move_cents'] < 0 else "Flat"
        L(f"| {e['num']} | {e['type']} | {e['score']} | {e['move_cents']:+.2f} | {direction} |")
    L("")

    # Events with 0 move
    zero_moves = [e for e in all_events if e['move_cents'] is not None and abs(e['move_cents']) < 0.01]
    L(f"### Events with zero price impact: {len(zero_moves)} / {len(all_events)}")
    L("")

    # Spread behavior
    L("### Spread Dynamics")
    L("")
    spread_widened = [e for e in all_events
                     if e['spread_before'] is not None and e['spread_after'] is not None
                     and e['spread_after'] > e['spread_before'] + 0.001]
    spread_narrowed = [e for e in all_events
                      if e['spread_before'] is not None and e['spread_after'] is not None
                      and e['spread_after'] < e['spread_before'] - 0.001]
    L(f"- Events where spread widened after: {len(spread_widened)}")
    L(f"- Events where spread narrowed after: {len(spread_narrowed)}")
    L(f"- Events where spread unchanged: {len(all_events) - len(spread_widened) - len(spread_narrowed)}")
    L("")

    # Depth analysis
    L("### Depth Dynamics")
    L("")
    for etype in ['4', '6', 'W']:
        a = agg[etype]
        bid_change = ((a['avg_depth_bid_after'] - a['avg_depth_bid_before']) / a['avg_depth_bid_before'] * 100) if a['avg_depth_bid_before'] > 0 else 0
        ask_change = ((a['avg_depth_ask_after'] - a['avg_depth_ask_before']) / a['avg_depth_ask_before'] * 100) if a['avg_depth_ask_before'] > 0 else 0
        L(f"- **{etype}s**: Bid1 depth change: {bid_change:+.1f}%, Ask1 depth change: {ask_change:+.1f}%")
    L("")

    report = '\n'.join(lines)
    with open(OUTPUT, 'w') as f:
        f.write(report)
    print(f"\nReport written to {OUTPUT}")
    print(f"Total events: {len(all_events)}")
    print(f"Arb windows: {len(all_arbs)}")
    print(f"Liquidity vacuums: {len(all_vacuums)}")


if __name__ == '__main__':
    main()
