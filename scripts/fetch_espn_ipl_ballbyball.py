"""Fetch ball-by-ball data for all IPL 2026 matches from ESPN's playbyplay API
into a single Excel workbook, one tab per match.

Endpoint structure
------------------
- Scoreboard for a date (returns matches scheduled / completed that day):
    https://site.web.api.espn.com/apis/site/v2/sports/cricket/8048/scoreboard?dates=YYYYMMDD
- Ball-by-ball commentary (paginated, 25 items / page, chronological within page):
    https://site.web.api.espn.com/apis/site/v2/sports/cricket/8048/playbyplay?event={mid}&page={n}

Per-ball record we keep
-----------------------
- bbb_ts_ms       int   ESPN's ball-bowled timestamp (millisecond Unix)
- bbb_ist         str   the same, formatted in IST
- innings         int   1 or 2 (period)
- over_actual     float 0.1 / 0.2 / ... / 19.6 (ESPN's "actual" overs)
- ball_in_over    int   1..6
- over_complete   bool
- play_type       str   run / no run / four / six / wicket
- score_value     int   runs scored on this delivery
- is_wide         int   1 if wide
- is_noball       int   1 if no-ball
- is_legbye       int   1 if leg-byes
- is_bye          int   1 if byes
- home_score      str   e.g. "51/1"
- away_score      str   e.g. "0"
- bowler          str   short name
- batsman         str   short name
- dismissal       str   if any
- short_text      str   ESPN's one-line summary
- seq             int   ESPN sequence number

Idempotency
-----------
- Workbook lives at captures/espn_ipl2026_ballbyball.xlsx
- Each tab is named with the match slug (e.g. "che-kol-2026-04-14")
- On re-run we open the existing workbook, skip matches whose tab already
  exists, and only fetch + append new ones.

Usage
-----
    venv/bin/python scripts/fetch_espn_ipl_ballbyball.py
    venv/bin/python scripts/fetch_espn_ipl_ballbyball.py --start 2026-03-22 --end 2026-05-31
    venv/bin/python scripts/fetch_espn_ipl_ballbyball.py --refetch che-kol-2026-04-14
"""
from __future__ import annotations

import argparse
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

IST = timezone(timedelta(hours=5, minutes=30))
LEAGUE = 8048  # IPL 2026
SCOREBOARD = "https://site.web.api.espn.com/apis/site/v2/sports/cricket/{league}/scoreboard?dates={date}"
PBP = "https://site.web.api.espn.com/apis/site/v2/sports/cricket/{league}/playbyplay?event={mid}&page={page}"
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

OUT = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures/espn_ipl2026_ballbyball.xlsx")

COLUMNS = [
    "bbb_ts_ms", "bbb_ist", "innings", "over_actual", "ball_in_over",
    "over_complete", "play_type", "score_value", "batting_team",
    "is_wide", "is_noball", "is_legbye", "is_bye",
    "home_score", "away_score", "bowler", "batsman",
    "dismissal", "short_text", "seq",
]


@dataclass
class Match:
    espn_id: int
    date_iso: str            # 2026-04-14
    home_abbr: str           # CSK
    away_abbr: str           # KKR
    name: str                # "Chennai Super Kings v Kolkata Knight Riders"

    @property
    def slug(self) -> str:
        # Map ESPN abbreviations -> the slugs used by polymarket / our captures.
        # captures/.../cricipl-che-kol-2026-04-14
        m = {
            "CSK": "che", "KKR": "kol", "MI": "mum", "RCB": "roy",
            "RR": "raj", "DC": "del", "GT": "guj", "LSG": "luc",
            "PBKS": "pun", "SRH": "sun",
        }
        h = m.get(self.home_abbr, self.home_abbr.lower())
        a = m.get(self.away_abbr, self.away_abbr.lower())
        return f"{h}-{a}-{self.date_iso}"


def fetch_scoreboard(client: httpx.Client, day: date) -> list[Match]:
    url = SCOREBOARD.format(league=LEAGUE, date=day.strftime("%Y%m%d"))
    r = client.get(url, headers=HEADERS, timeout=15)
    r.raise_for_status()
    out: list[Match] = []
    for e in r.json().get("events", []):
        status = e.get("status", {}).get("type", {}).get("description", "")
        if status != "Result":
            continue  # skip scheduled / abandoned
        comp = e.get("competitions", [{}])[0]
        comps = comp.get("competitors", [])
        if len(comps) < 2:
            continue
        # ESPN orders by team list — first competitor is the home/listed-first team
        home = comps[0].get("team", {}).get("abbreviation", "")
        away = comps[1].get("team", {}).get("abbreviation", "")
        # event date is in UTC; for IPL the slug uses the IST calendar date
        utc = datetime.fromisoformat(e["date"].replace("Z", "+00:00"))
        ist_date = utc.astimezone(IST).date().isoformat()
        out.append(Match(
            espn_id=int(e["id"]),
            date_iso=ist_date,
            home_abbr=home,
            away_abbr=away,
            name=e.get("name", ""),
        ))
    return out


def _get_with_retry(client: httpx.Client, url: str, *, attempts: int = 6) -> httpx.Response:
    for i in range(attempts):
        try:
            r = client.get(url, headers=HEADERS, timeout=20)
            if r.status_code == 200:
                return r
            # 502/503/504 from ESPN are transient — back off and retry.
            if r.status_code in (502, 503, 504, 429) and i < attempts - 1:
                time.sleep(1.5 * (i + 1))
                continue
            r.raise_for_status()
            return r
        except (httpx.ReadTimeout, httpx.ConnectError, httpx.RemoteProtocolError):
            if i == attempts - 1:
                raise
            time.sleep(1.5 * (i + 1))
    raise RuntimeError(f"unreachable")


def fetch_ball_by_ball(client: httpx.Client, mid: int) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        url = PBP.format(league=LEAGUE, mid=mid, page=page)
        r = _get_with_retry(client, url)
        d = r.json().get("commentary", {})
        items = d.get("items", [])
        for it in items:
            over = it.get("over") or {}
            pt = it.get("playType") or {}
            bowler = (it.get("bowler") or {}).get("athlete") or {}
            batsman = (it.get("batsman") or {}).get("athlete") or {}
            dismissal = it.get("dismissal") or {}
            ts_ms = it.get("bbbTimestamp")
            if ts_ms is None:
                continue  # match-context comment, not a ball
            team_field = it.get("team") or {}
            rows.append({
                "bbb_ts_ms": ts_ms,
                "bbb_ist": datetime.fromtimestamp(ts_ms / 1000, IST)
                                  .strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                "innings": it.get("period"),
                "over_actual": over.get("actual"),
                "ball_in_over": over.get("ball"),
                "over_complete": over.get("complete"),
                "play_type": pt.get("description", ""),
                "score_value": it.get("scoreValue"),
                "batting_team": team_field.get("abbreviation", ""),
                "is_wide": int(bool(over.get("wide"))),
                "is_noball": int(bool(over.get("noBall"))),
                "is_legbye": int(bool(over.get("legByes"))),
                "is_bye": int(bool(over.get("byes"))),
                "home_score": it.get("homeScore", ""),
                "away_score": it.get("awayScore", ""),
                "bowler": bowler.get("shortName", ""),
                "batsman": batsman.get("shortName", ""),
                "dismissal": (dismissal.get("type") or "") if isinstance(dismissal, dict) else "",
                "short_text": (it.get("shortText") or "")[:300],
                "seq": it.get("sequence"),
            })
        page_count = d.get("pageCount", 1)
        if page >= page_count:
            break
        page += 1
        time.sleep(0.2)  # be kind to ESPN
    rows.sort(key=lambda r: (r["innings"] or 0, r["seq"] or 0))
    return rows


def open_or_create_workbook() -> Workbook:
    if OUT.exists():
        return load_workbook(OUT)
    wb = Workbook()
    # Drop the default sheet — we replace with summary later
    default = wb.active
    if default is not None:
        wb.remove(default)
    return wb


def write_match_sheet(wb: Workbook, match: Match, rows: list[dict]) -> None:
    sheet_name = match.slug[:31]  # Excel limit
    if sheet_name in wb.sheetnames:
        return
    ws = wb.create_sheet(title=sheet_name)
    # Metadata block (rows 1-4), then header row 6, then data
    ws.append(["espn_id", match.espn_id])
    ws.append(["match", match.name])
    ws.append(["date_ist", match.date_iso])
    ws.append(["fetched_at_utc", datetime.now(timezone.utc).isoformat()])
    ws.append([])
    ws.append(COLUMNS)
    for r in rows:
        ws.append([r.get(c) for c in COLUMNS])
    # Auto-size a few columns roughly
    widths = {"bbb_ist": 24, "play_type": 10, "short_text": 60,
              "bowler": 14, "batsman": 14, "dismissal": 14}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)


def upsert_summary(wb: Workbook, matches: list[Match], counts: dict[str, int]) -> None:
    if "_summary" in wb.sheetnames:
        wb.remove(wb["_summary"])
    ws = wb.create_sheet(title="_summary", index=0)
    ws.append(["slug", "espn_id", "date_ist", "name", "ball_count"])
    for m in matches:
        ws.append([m.slug, m.espn_id, m.date_iso, m.name, counts.get(m.slug, 0)])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--start", default="2026-03-22", help="YYYY-MM-DD (IPL season start)")
    parser.add_argument("--end", default=date.today().isoformat(), help="YYYY-MM-DD")
    parser.add_argument("--refetch", action="append", default=[],
                        help="slug to forcibly refetch (drop existing tab)")
    args = parser.parse_args()

    start = datetime.fromisoformat(args.start).date()
    end = datetime.fromisoformat(args.end).date()

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = open_or_create_workbook()
    existing_sheets = set(wb.sheetnames)

    for slug in args.refetch:
        if slug in wb.sheetnames:
            wb.remove(wb[slug])
            existing_sheets.discard(slug)

    print(f"Workbook: {OUT}")
    print(f"Existing match tabs: "
          f"{sorted(s for s in existing_sheets if s != '_summary')}")
    print(f"Scanning IPL 2026 scoreboard {start} → {end} ...")

    discovered: list[Match] = []
    seen_ids: set[int] = set()
    with httpx.Client() as client:
        cursor = start
        while cursor <= end:
            try:
                matches = fetch_scoreboard(client, cursor)
            except Exception as e:
                print(f"  {cursor}: ERR {e}")
                cursor += timedelta(days=1)
                continue
            for m in matches:
                if m.espn_id in seen_ids:
                    continue
                seen_ids.add(m.espn_id)
                discovered.append(m)
            time.sleep(0.1)
            cursor += timedelta(days=1)

        print(f"Discovered {len(discovered)} completed matches")
        ball_counts: dict[str, int] = {}
        for m in discovered:
            tab = m.slug[:31]
            if tab in existing_sheets:
                # Pull cached row count for the summary
                ws = wb[tab]
                ball_counts[m.slug] = ws.max_row - 6  # header at row 6
                print(f"  [skip] {m.slug:<26}  espn={m.espn_id}  (already cached)")
                continue
            print(f"  [fetch] {m.slug:<26}  espn={m.espn_id}  ", end="", flush=True)
            try:
                rows = fetch_ball_by_ball(client, m.espn_id)
            except Exception as e:
                print(f"ERR {e}")
                continue
            ball_counts[m.slug] = len(rows)
            write_match_sheet(wb, m, rows)
            print(f"{len(rows)} balls")

        upsert_summary(wb, discovered, ball_counts)
        wb.save(OUT)

    total = sum(ball_counts.values())
    print(f"\nSaved {OUT}")
    print(f"Tabs: {len([s for s in wb.sheetnames if s != '_summary'])}  "
          f"Total balls: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
