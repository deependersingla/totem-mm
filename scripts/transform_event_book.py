#!/usr/bin/env python3
"""Transform raw event_book_capture Excel into visual orderbook snapshot layout.

Reads the flat row-per-snapshot format and produces a new Excel where each sheet
has visual orderbook tables arranged in a grid:
  - 5 tables horizontally (one per 200ms snapshot = 1 second)
  - Next second starts a new row of 5 tables below

Each table looks like:
    ORDER BOOK  19:02:50.200  ms: -1200
    Delhi Capitals              Mumbai Indians
    0.36c                                0.64c
    BID    SIZE    ASK    SIZE  BID    SIZE    ASK    SIZE
    0.36   10000   0.37   80    0.63   80      0.64   10000
    0.35   10801   0.38   12297 0.62   12297   0.65   10801
    ...

Usage:
    # Transform a single sheet (first one):
    python transform_event_book.py <input.xlsx>

    # Transform a specific sheet:
    python transform_event_book.py <input.xlsx> --sheet E1_4_153740_762

    # Transform all sheets:
    python transform_event_book.py <input.xlsx> --all
"""

import argparse
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
TITLE_FONT = Font(bold=True, size=11)
TIME_FONT = Font(color="888888", size=9)
TEAM_FONT = Font(bold=True, size=10)
BID_FONT = Font(color="22C55E", size=9)   # green
ASK_FONT = Font(color="EF4444", size=9)   # red
PRICE_FONT = Font(size=9)
BID_PRICE_FONT = Font(color="22C55E", size=10, bold=True)
ASK_PRICE_FONT = Font(color="EF4444", size=10, bold=True)
EVENT_FONT = Font(bold=True, size=14, color="FFFFFF")
EVENT_4_FILL = PatternFill("solid", fgColor="166534")  # dark green
EVENT_6_FILL = PatternFill("solid", fgColor="1D4ED8")  # dark blue
EVENT_W_FILL = PatternFill("solid", fgColor="991B1B")  # dark red
SEPARATOR_FILL = PatternFill("solid", fgColor="F3F4F6")
PRE_EVENT_FILL = PatternFill("solid", fgColor="EFF6FF")   # light blue
POST_EVENT_FILL = PatternFill("solid", fgColor="F0FDF4")  # light green
EVENT_MOMENT_FILL = PatternFill("solid", fgColor="FEF3C7")  # yellow
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

SNAPSHOTS_PER_ROW = 5  # 5 x 200ms = 1 second
TABLE_WIDTH = 9         # columns per outcome pair table (BID, SIZE, ASK, SIZE x2 + gap)
OUTCOME_COLS = 4        # BID, SIZE, ASK, SIZE
GAP_COLS = 1            # gap between two outcome blocks
LEVELS = 5              # order book depth


def book_signature(books: dict) -> tuple:
    """Return a hashable signature of a book state for equality comparison."""
    parts = []
    for oname in sorted(books.keys()):
        book = books[oname]
        bids = tuple((p, s) for p, s in book.get("bids", []))
        asks = tuple((p, s) for p, s in book.get("asks", []))
        parts.append((oname, bids, asks))
    return tuple(parts)


def group_snapshots_by_second(snapshots: list[dict]) -> list[list[dict]]:
    """Group snapshots into rows by wall-clock second (based on IST HH:MM:SS)."""
    rows = []
    current_sec = None
    current_group = []
    for snap in snapshots:
        ist = snap.get("ist") or ""
        # IST format: "15:37:19.142" — extract HH:MM:SS
        sec_key = ist.split(".")[0] if "." in ist else ist
        if sec_key != current_sec:
            if current_group:
                rows.append(current_group)
            current_group = [snap]
            current_sec = sec_key
        else:
            current_group.append(snap)
    if current_group:
        rows.append(current_group)
    return rows


def merge_within_second(second_snaps: list[dict]) -> list[dict]:
    """Merge consecutive identical snapshots within a single second.
    Returns list of merged cells: [{books, timestamps, ms_offsets, snap_count}]
    """
    if not second_snaps:
        return []
    merged = []
    current_sig = None
    current_cell = None
    for snap in second_snaps:
        sig = book_signature(snap["books"])
        if sig == current_sig and current_cell is not None:
            # Merge into current cell
            current_cell["timestamps"].append(snap["ist"])
            current_cell["ms_offsets"].append(snap["ms_offset"])
            current_cell["snap_count"] += 1
        else:
            if current_cell is not None:
                merged.append(current_cell)
            current_cell = {
                "books": snap["books"],
                "timestamps": [snap["ist"]],
                "ms_offsets": [snap["ms_offset"]],
                "snap_count": 1,
            }
            current_sig = sig
    if current_cell is not None:
        merged.append(current_cell)
    return merged


def event_fill(event_type: str) -> PatternFill:
    if "6" in event_type:
        return EVENT_6_FILL
    if "W" in event_type or "w" in event_type:
        return EVENT_W_FILL
    return EVENT_4_FILL


# ── Parse raw sheet ─────────────────────────────────────────────────────────

def parse_sheet(ws):
    """Parse a raw event sheet into metadata + list of snapshot dicts."""
    rows = list(ws.iter_rows(values_only=True))

    # Row 1: metadata
    meta_row = rows[0]
    event_str = str(meta_row[0] or "")    # "Event: 4"
    score_str = str(meta_row[1] or "")    # "Score: 16/0 (1.3)"
    time_str = str(meta_row[2] or "")     # "Time: 15:37:40.762"
    snap_count = str(meta_row[3] or "")   # "Snapshots: 874"

    # Row 3: headers
    headers = [str(h or "") for h in rows[2]]

    # Find outcome names from headers
    # Format: "Delhi Capitals_bid1_price", "Mumbai Indians_bid1_price"
    outcomes = []
    for h in headers:
        if "_bid1_price" in h:
            oname = h.replace("_bid1_price", "")
            outcomes.append(oname)

    # Parse data rows (row index 3+)
    snapshots = []
    for row in rows[3:]:
        vals = list(row)
        ist = vals[0]
        ms_offset = vals[1]

        # Skip rows with no book data
        if all(v is None for v in vals[2:]):
            continue

        books = {}
        col = 2  # start after IST, ms_offset
        for oname in outcomes:
            bids = []
            for _ in range(LEVELS):
                p = vals[col] if col < len(vals) else None
                s = vals[col + 1] if col + 1 < len(vals) else None
                bids.append((p, s))
                col += 2
            asks = []
            for _ in range(LEVELS):
                p = vals[col] if col < len(vals) else None
                s = vals[col + 1] if col + 1 < len(vals) else None
                asks.append((p, s))
                col += 2
            books[oname] = {"bids": bids, "asks": asks}

        snapshots.append({
            "ist": ist,
            "ms_offset": ms_offset,
            "books": books,
        })

    meta = {
        "event": event_str,
        "score": score_str,
        "time": time_str,
        "snap_count": snap_count,
        "outcomes": outcomes,
    }
    return meta, snapshots


# ── Write visual sheet ──────────────────────────────────────────────────────

def write_visual_sheet(wb_out: Workbook, sheet_name: str, meta: dict,
                       snapshots: list[dict]):
    """Write a visually formatted sheet with orderbook tables in a grid."""
    ws = wb_out.create_sheet(title=sheet_name[:31])
    outcomes = meta["outcomes"]

    # ── Event header (row 1-2) ──────────────────────────────────────────
    event_type = meta["event"].replace("Event: ", "").strip()
    score = meta["score"].replace("Score: ", "").strip()
    event_time = meta["time"].replace("Time: ", "").strip()

    # Determine best bid for each outcome to show implied price
    # from first snapshot with data near event (ms_offset closest to 0)
    event_snap = None
    for s in snapshots:
        if s["ms_offset"] is not None and abs(s["ms_offset"]) < 1000:
            event_snap = s
            break
    if not event_snap and snapshots:
        event_snap = snapshots[len(snapshots) // 2]

    # Total width: SNAPSHOTS_PER_ROW tables, each table has columns for all outcomes
    single_table_cols = len(outcomes) * OUTCOME_COLS + (len(outcomes) - 1) * GAP_COLS
    total_cols = SNAPSHOTS_PER_ROW * (single_table_cols + GAP_COLS)

    # Row 1: Event banner
    fill = event_fill(event_type)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 50))
    cell = ws.cell(row=1, column=1,
                   value=f"  EVENT: {event_type}   |   {score}   |   {event_time}   |   {meta['snap_count']}")
    cell.font = EVENT_FONT
    cell.fill = fill
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2: Outcome implied prices at event time
    if event_snap:
        prices = []
        for oname in outcomes:
            book = event_snap["books"].get(oname, {"bids": [], "asks": []})
            best_bid = book["bids"][0][0] if book["bids"] and book["bids"][0][0] else "--"
            best_ask = book["asks"][0][0] if book["asks"] and book["asks"][0][0] else "--"
            if isinstance(best_bid, (int, float)):
                prices.append(f"{oname}: {best_bid:.2f}c bid / {best_ask:.2f}c ask" if isinstance(best_ask, (int, float)) else f"{oname}: {best_bid:.2f}c")
            else:
                prices.append(f"{oname}: --")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(total_cols, 50))
        ws.cell(row=2, column=1, value="  " + "   |   ".join(prices)).font = Font(size=10, color="666666")

    current_row = 4  # start grid from row 4

    # ── Grid of merged snapshot tables ──────────────────────────────────
    # Group by second, then merge consecutive identical snapshots within each second
    second_rows = group_snapshots_by_second(snapshots)

    for second_snaps in second_rows:
        merged_cells = merge_within_second(second_snaps)

        # Each table block takes: 1 (title) + 1 (team names) + 1 (header) + LEVELS (data) + 1 (blank)
        block_height = LEVELS + 4

        for cell_idx, mcell in enumerate(merged_cells):
            # Column offset for this table
            col_start = cell_idx * (single_table_cols + GAP_COLS) + 1

            snap = {"books": mcell["books"]}
            timestamps = mcell["timestamps"]
            ms_offsets = mcell["ms_offsets"]
            snap_count = mcell["snap_count"]

            # Timestamps display: first full, rest suffix-only after the HH:MM:SS prefix
            first_ts = timestamps[0] or ""
            if "." in first_ts:
                base = first_ts.split(".")[0]
                suffixes = []
                for t in timestamps:
                    if t and "." in t:
                        suffixes.append("." + t.split(".")[1])
                    else:
                        suffixes.append("")
                ts_str = base + " " + " ".join(suffixes)
            else:
                ts_str = " ".join(t or "" for t in timestamps)

            # ms offset: show range if merged, otherwise single value
            ms_vals = [m for m in ms_offsets if m is not None]
            if not ms_vals:
                ms_str = ""
            elif snap_count == 1:
                m = ms_vals[0]
                ms_str = f"+{m}ms" if m >= 0 else f"{m}ms"
            else:
                lo, hi = ms_vals[0], ms_vals[-1]
                ms_str = f"{lo}→{hi}ms ({snap_count}x)"

            # Determine time-phase fill based on first ms_offset
            ms = ms_vals[0] if ms_vals else None
            if ms is not None and abs(ms) < 500:
                phase_fill = EVENT_MOMENT_FILL
            elif ms is not None and ms < 0:
                phase_fill = PRE_EVENT_FILL
            else:
                phase_fill = POST_EVENT_FILL

            ist = ts_str
            r = current_row

            # Title row: ORDER BOOK  HH:MM:SS.mmm  ms: +/-XXX
            ws.merge_cells(start_row=r, start_column=col_start,
                           end_row=r, end_column=col_start + single_table_cols - 1)
            title_cell = ws.cell(row=r, column=col_start,
                                 value=f"ORDER BOOK   {ist}   {ms_str}")
            title_cell.font = TITLE_FONT
            title_cell.fill = phase_fill
            title_cell.border = THIN_BORDER

            r += 1

            # Team name row + best bid/ask price
            for oi, oname in enumerate(outcomes):
                oc = col_start + oi * (OUTCOME_COLS + (GAP_COLS if oi > 0 else 0))
                if oi > 0:
                    oc = col_start + OUTCOME_COLS + GAP_COLS

                book = snap["books"].get(oname, {"bids": [], "asks": []})
                best_bid = book["bids"][0][0] if book["bids"] and book["bids"][0][0] else None
                best_ask = book["asks"][0][0] if book["asks"] and book["asks"][0][0] else None

                # Team name
                ws.cell(row=r, column=oc, value=oname).font = TEAM_FONT

                # Best bid (green) and best ask (red) on same row
                if best_bid is not None:
                    ws.cell(row=r, column=oc + 1,
                            value=f"{best_bid:.3f}c").font = BID_PRICE_FONT
                if best_ask is not None:
                    ws.cell(row=r, column=oc + 3,
                            value=f"{best_ask:.3f}c").font = ASK_PRICE_FONT

            r += 1

            # Column headers: BID SIZE ASK SIZE | BID SIZE ASK SIZE
            for oi, oname in enumerate(outcomes):
                oc = col_start + oi * (OUTCOME_COLS + (GAP_COLS if oi > 0 else 0))
                if oi > 0:
                    oc = col_start + OUTCOME_COLS + GAP_COLS

                for ci, label in enumerate(["BID", "SIZE", "ASK", "SIZE"]):
                    cell = ws.cell(row=r, column=oc + ci, value=label)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="right")
                    cell.border = THIN_BORDER

            r += 1

            # Data rows (5 levels)
            for lvl in range(LEVELS):
                for oi, oname in enumerate(outcomes):
                    oc = col_start + oi * (OUTCOME_COLS + (GAP_COLS if oi > 0 else 0))
                    if oi > 0:
                        oc = col_start + OUTCOME_COLS + GAP_COLS

                    book = snap["books"].get(oname, {"bids": [], "asks": []})

                    # Bid
                    bp = book["bids"][lvl][0] if lvl < len(book["bids"]) else None
                    bs = book["bids"][lvl][1] if lvl < len(book["bids"]) else None
                    if bp is not None:
                        ws.cell(row=r, column=oc, value=bp).font = BID_FONT
                        ws.cell(row=r, column=oc, value=bp).number_format = '0.00'
                    else:
                        ws.cell(row=r, column=oc, value="--").font = Font(color="888888", size=9)
                    if bs is not None:
                        ws.cell(row=r, column=oc + 1, value=round(bs, 1)).font = PRICE_FONT
                        ws.cell(row=r, column=oc + 1).number_format = '#,##0.0'
                    else:
                        ws.cell(row=r, column=oc + 1, value="--").font = Font(color="888888", size=9)

                    # Ask
                    ap = book["asks"][lvl][0] if lvl < len(book["asks"]) else None
                    as_ = book["asks"][lvl][1] if lvl < len(book["asks"]) else None
                    if ap is not None:
                        ws.cell(row=r, column=oc + 2, value=ap).font = ASK_FONT
                        ws.cell(row=r, column=oc + 2).number_format = '0.00'
                    else:
                        ws.cell(row=r, column=oc + 2, value="--").font = Font(color="888888", size=9)
                    if as_ is not None:
                        ws.cell(row=r, column=oc + 3, value=round(as_, 1)).font = PRICE_FONT
                        ws.cell(row=r, column=oc + 3).number_format = '#,##0.0'
                    else:
                        ws.cell(row=r, column=oc + 3, value="--").font = Font(color="888888", size=9)

                    # Borders
                    for c in range(4):
                        ws.cell(row=r, column=oc + c).border = THIN_BORDER
                        ws.cell(row=r, column=oc + c).alignment = Alignment(horizontal="right")

                r += 1

        current_row += block_height

    # Column widths
    for col in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return ws.max_row


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Transform event_book_capture Excel to visual layout")
    parser.add_argument("input", help="Input Excel file from event_book_capture.py")
    parser.add_argument("--sheet", help="Specific sheet name to transform (default: first sheet)")
    parser.add_argument("--all", action="store_true", help="Transform all sheets")
    parser.add_argument("--output", help="Output file path (default: <input>_visual.xlsx)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"File not found: {args.input}")
        sys.exit(1)

    print(f"Loading {args.input}...")
    wb_in = load_workbook(args.input, read_only=True)
    print(f"Sheets: {len(wb_in.sheetnames)}")

    if args.all:
        sheets_to_process = wb_in.sheetnames
    elif args.sheet:
        if args.sheet not in wb_in.sheetnames:
            print(f"Sheet '{args.sheet}' not found. Available: {wb_in.sheetnames}")
            sys.exit(1)
        sheets_to_process = [args.sheet]
    else:
        sheets_to_process = [wb_in.sheetnames[0]]

    output_path = args.output or args.input.replace(".xlsx", "_visual.xlsx")
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sname in sheets_to_process:
        print(f"  Transforming '{sname}'...", end=" ", flush=True)
        ws_in = wb_in[sname]
        meta, snapshots = parse_sheet(ws_in)
        if not snapshots:
            print("(no data, skipped)")
            continue
        total_rows = write_visual_sheet(wb_out, sname, meta, snapshots)
        print(f"{len(snapshots)} snapshots → {total_rows} rows")

    print(f"\nSaving to {output_path}...")
    wb_out.save(output_path)
    print(f"Done. {len(sheets_to_process)} sheet(s) transformed.")


if __name__ == "__main__":
    main()
