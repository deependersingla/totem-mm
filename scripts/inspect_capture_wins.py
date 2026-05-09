"""Find the deliveries where our capture beat ESPN by the widest margin
(most negative delta = capture - espn). Shows the top-30 capture-wins across
the 11 clean matches, with full context.
"""
from __future__ import annotations

import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from compare_espn_vs_capture_timing import align_match  # type: ignore
from compare_clean_matches import CLEAN_SLUGS  # type: ignore

WORKBOOK = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures/espn_ipl2026_ballbyball.xlsx")
IST = timezone(timedelta(hours=5, minutes=30))


def fmt_ist(ts_ms: int) -> str:
    return datetime.fromtimestamp(ts_ms / 1000, IST).strftime("%H:%M:%S.%f")[:-3]


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=True)

    capture_wins: list[dict] = []
    for slug in CLEAN_SLUGS:
        if slug not in wb.sheetnames:
            continue
        result = align_match(slug, wb[slug])
        for m in result["matched"]:
            if m["delta_ms"] < -500:  # capture at least 0.5s ahead
                capture_wins.append({**m, "slug": slug})

    capture_wins.sort(key=lambda x: x["delta_ms"])  # most negative first

    print(f"Capture-wins (delta < -0.5s) across 11 clean matches: {len(capture_wins)}\n")

    print(f"{'rank':>4}  {'slug':<24} {'inn':>3}  {'over':>5}  "
          f"{'delta_s':>8}  {'play_type':<10} {'cap_sig':>7}  "
          f"{'espn_ist':>14}  {'cap_ist':>14}  {'espn_score':<10}  cap_score")
    print("-" * 130)
    for i, w in enumerate(capture_wins[:30], 1):
        delta = w["delta_ms"] / 1000.0
        print(f"{i:>4}  {w['slug']:<24} {str(w['innings']):>3}  "
              f"{str(w['overs']):>5}  {delta:>+8.2f}  "
              f"{(w['espn_play_type'] or ''):<10} "
              f"{str(w['cap_signal']):>7}  "
              f"{fmt_ist(w['espn_ts_ms']):>14}  "
              f"{fmt_ist(w['cap_ts_ms']):>14}  "
              f"{(w['espn_score'] or ''):<10}  {w['cap_score']}")

    # Quick distribution view
    print("\nDistribution of capture-win sizes (seconds, all clean matches):")
    sizes = sorted(-w["delta_ms"] / 1000 for w in capture_wins)  # positive = capture lead
    if sizes:
        print(f"  count : {len(sizes)}")
        print(f"  min   : {sizes[0]:.2f}s   (smallest capture lead)")
        print(f"  p50   : {sizes[len(sizes) // 2]:.2f}s")
        print(f"  p90   : {sizes[int(len(sizes) * 0.9)]:.2f}s")
        print(f"  max   : {sizes[-1]:.2f}s   (largest capture lead)")

    # Slice by event type
    print("\nCapture-wins by ESPN play_type:")
    from collections import Counter
    pt_counter = Counter((w["espn_play_type"] or "").lower() for w in capture_wins)
    for pt, n in pt_counter.most_common():
        sub = [-w["delta_ms"] / 1000 for w in capture_wins
               if (w["espn_play_type"] or "").lower() == pt]
        sub.sort()
        if sub:
            p50 = sub[len(sub) // 2]
            print(f"  {pt:<12} n={n:>4}  p50_lead={p50:.2f}s  max_lead={sub[-1]:.2f}s")


if __name__ == "__main__":
    main()
