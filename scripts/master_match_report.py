#!/usr/bin/env python3
"""master_match_report.py — single master match report.

Combines the local capture DB (cricket events + book snapshots) with
Goldsky subgraph trades (the authoritative source for wallet PnL).

Sheets:
  1_special_events   W/4/6 events with IST timestamps
  2_event_books      book snapshots -2m..+1m around each event (stacked mini-tables)
  3_real_trades      every OrderFilledEvent from Goldsky for the market's tokens
  4_wallets          wallets with >=100 trades AND >=10 trades on team A AND
                     >=10 trades on team B, with correct FIFO PnL, turnover,
                     capital rotation, bps, maker/taker split

Usage:
  venv/bin/python scripts/master_match_report.py <path_to_match_capture.db>
"""

import argparse, sqlite3, json, sys, time
from bisect import bisect_left, bisect_right
from collections import deque, defaultdict
from datetime import datetime, timezone, timedelta

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

IST = timezone(timedelta(hours=5, minutes=30))
GAMMA_API = "https://gamma-api.polymarket.com"
GOLDSKY_ENDPOINT = (
    "https://api.goldsky.com/api/public/"
    "project_cl6mb8i9h0003e201j6li0diw/"
    "subgraphs/orderbook-subgraph/0.0.1/gn"
)
CTF_EXCHANGE = "0x4bfb41d5b3570defd03c39a9a4d8de6bd8b8982e"
USDC_DECIMALS = 6
HEADERS = {"User-Agent": "Mozilla/5.0"}


# ── Time helpers ─────────────────────────────────────────────

def ms_to_ist(ms: int) -> str:
    if ms is None: return ""
    return datetime.fromtimestamp(ms / 1000, tz=IST).strftime("%H:%M:%S.%f")[:-3]


def s_to_ist(s: int) -> str:
    if s is None: return ""
    return datetime.fromtimestamp(s, tz=IST).strftime("%H:%M:%S")


# ── Gamma API ────────────────────────────────────────────────

def fetch_market(slug: str) -> dict:
    """Fetch market metadata; falls back to /events?slug= for cricket markets."""
    resp = requests.get(f"{GAMMA_API}/markets", params={"slug": slug}, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    markets = resp.json()
    if markets:
        return markets[0]
    resp = requests.get(f"{GAMMA_API}/events", params={"slug": slug}, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    events = resp.json()
    if events:
        for m in events[0].get("markets", []):
            if m.get("slug") == slug:
                return m
    raise RuntimeError(f"No market found for slug '{slug}'")


def parse_tokens(market: dict) -> tuple[list[str], list[str]]:
    tokens = market.get("clobTokenIds", "")
    outcomes = market.get("outcomes", "[]")
    if isinstance(tokens, str):
        try: tokens = json.loads(tokens)
        except json.JSONDecodeError: tokens = tokens.split(",") if tokens else []
    if isinstance(outcomes, str):
        try: outcomes = json.loads(outcomes)
        except json.JSONDecodeError: outcomes = []
    return tokens, outcomes


def parse_settlement(market: dict, outcomes: list[str]) -> tuple[str, dict[str, float]]:
    prices = market.get("outcomePrices")
    if isinstance(prices, str):
        try: prices = json.loads(prices)
        except json.JSONDecodeError: prices = None
    if not prices or len(prices) < 2:
        return "", {o: 0.5 for o in outcomes}
    p1, p2 = float(prices[0]), float(prices[1])
    winner = outcomes[0] if p1 >= p2 else outcomes[1]
    return winner, {outcomes[0]: p1, outcomes[1]: p2}


# ── Goldsky ──────────────────────────────────────────────────

def _query_cursor(field: str, token_ids: list[str], page_size: int = 1000) -> list[dict]:
    events = []
    last_id = ""
    side = "maker" if "maker" in field else "taker"
    page = 0
    while True:
        id_filter = f', id_gt: "{last_id}"' if last_id else ""
        query = f"""{{
          orderFilledEvents(
            first: {page_size},
            orderBy: id, orderDirection: asc,
            where: {{ {field}: {json.dumps(token_ids)}{id_filter} }}
          ) {{
            id maker taker makerAssetId takerAssetId
            makerAmountFilled takerAmountFilled
            fee timestamp transactionHash orderHash
          }}
        }}"""
        data = None
        for attempt in range(3):
            try:
                resp = requests.post(GOLDSKY_ENDPOINT, json={"query": query}, timeout=60)
                resp.raise_for_status()
                data = resp.json()
                if "errors" not in data: break
                data = None
            except Exception as e:
                print(f"  {side} page {page} attempt {attempt+1} error: {e}")
                data = None
                time.sleep(0.5)
        if data is None: break
        batch = data.get("data", {}).get("orderFilledEvents", [])
        if not batch: break
        events.extend(batch)
        last_id = batch[-1]["id"]
        page += 1
        print(f"  [{side}-side] page {page}: +{len(batch)} total={len(events)}")
        if len(batch) < page_size: break
        time.sleep(0.15)
    return events


def fetch_all_events(token_ids: list[str]) -> list[dict]:
    all_events: dict[str, dict] = {}
    for field in ["makerAssetId_in", "takerAssetId_in"]:
        for ev in _query_cursor(field, token_ids):
            all_events[ev["id"]] = ev
    print(f"  deduplicated: {len(all_events)} unique events")
    return list(all_events.values())


# ── Event processing ─────────────────────────────────────────

def process_events(events: list[dict], token_ids: list[str],
                   outcome_names: list[str]) -> pd.DataFrame:
    token_to_outcome = dict(zip(token_ids, outcome_names))
    rows = []
    for ev in events:
        maker = ev["maker"].lower()
        taker = ev["taker"].lower()
        ma_id = ev["makerAssetId"]
        ta_id = ev["takerAssetId"]
        ma_amt = int(ev["makerAmountFilled"])
        ta_amt = int(ev["takerAmountFilled"])
        fee = int(ev["fee"])
        ts = int(ev["timestamp"])

        if ma_id in token_ids:
            outcome = token_to_outcome.get(ma_id, ma_id[:16])
            token_amt_raw = ma_amt
            usdc_amt_raw = ta_amt
            maker_side = "SELL"
        elif ta_id in token_ids:
            outcome = token_to_outcome.get(ta_id, ta_id[:16])
            token_amt_raw = ta_amt
            usdc_amt_raw = ma_amt
            maker_side = "BUY"
        else:
            continue

        usdc = usdc_amt_raw / (10 ** USDC_DECIMALS)
        qty = token_amt_raw / (10 ** USDC_DECIMALS)
        fee_usdc = fee / (10 ** USDC_DECIMALS)
        price = usdc / qty if qty > 0 else 0.0

        if maker_side == "BUY":
            buyer, seller = maker, taker
        else:
            buyer, seller = taker, maker

        rows.append({
            "event_id": ev["id"],
            "timestamp_unix": ts,
            "outcome": outcome,
            "maker_side": maker_side,
            "price": round(price, 6),
            "token_amount": round(qty, 4),
            "usdc_amount": round(usdc, 4),
            "fee_usdc": round(fee_usdc, 4),
            "maker": maker,
            "taker": taker,
            "buyer": buyer,
            "seller": seller,
            "tx_hash": ev["transactionHash"],
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["timestamp_unix", "event_id"]).reset_index(drop=True)
    return df


def classify_all_roles(df: pd.DataFrame) -> dict[str, str]:
    """MAKER = resting fill. TAKER = wallet crossed the book (matchOrders)."""
    wallet_taker_txs: dict[str, set[str]] = defaultdict(set)
    for _, row in df.iterrows():
        t = row["taker"]
        if t != CTF_EXCHANGE:
            wallet_taker_txs[t].add(row["tx_hash"])

    roles: dict[str, str] = {}
    for _, row in df.iterrows():
        m = row["maker"]
        if m == CTF_EXCHANGE: continue
        eid = row["event_id"]
        if row["taker"] != CTF_EXCHANGE:
            roles[eid] = "MAKER"
        elif row["tx_hash"] in wallet_taker_txs.get(m, set()):
            roles[eid] = "TAKER"
        else:
            roles[eid] = "MAKER"
    return roles


# ── FIFO ─────────────────────────────────────────────────────

class FIFOQueue:
    __slots__ = ("lots", "total_qty")

    def __init__(self):
        self.lots: deque[list] = deque()
        self.total_qty = 0.0

    def buy(self, qty: float, price: float):
        self.lots.append([qty, price])
        self.total_qty += qty

    def sell(self, qty: float, sell_price: float) -> float:
        remaining, realized = qty, 0.0
        while remaining > 1e-8 and self.lots:
            lot = self.lots[0]
            take = min(remaining, lot[0])
            realized += (sell_price - lot[1]) * take
            lot[0] -= take
            remaining -= take
            self.total_qty -= take
            if lot[0] < 1e-8:
                self.lots.popleft()
        if remaining > 1e-4:
            realized += sell_price * remaining
        return realized

    def settle(self, sp: float) -> float:
        pnl = 0.0
        while self.lots:
            lot = self.lots.popleft()
            if lot[0] < 1e-8: continue
            pnl += (sp - lot[1]) * lot[0]
            self.total_qty -= lot[0]
        return pnl


def compute_wallet_summary(wallet: str, w_df: pd.DataFrame, roles: dict,
                           outcomes: list[str], settlement: dict[str, float],
                           team_a: str, team_b: str) -> dict:
    total = len(w_df)
    wl = wallet.lower()
    buys_mask = w_df["buyer"] == wl
    sells_mask = w_df["seller"] == wl

    fin_maker = sum(1 for eid in w_df["event_id"] if roles.get(eid) == "MAKER")
    fin_taker = total - fin_maker

    fifo = {o: FIFOQueue() for o in outcomes}
    per_outcome_pnl = {o: 0.0 for o in outcomes}
    mk_sell_pnl = 0.0
    tk_sell_pnl = 0.0

    for _, r in w_df.iterrows():
        o = r["outcome"]
        if o not in fifo: continue
        if r["buyer"] == wl:
            fifo[o].buy(r["token_amount"], r["price"])
        else:
            pnl = fifo[o].sell(r["token_amount"], r["price"])
            per_outcome_pnl[o] += pnl
            role = roles.get(r["event_id"], "")
            if role == "MAKER": mk_sell_pnl += pnl
            else: tk_sell_pnl += pnl

    settlement_usdc = 0.0
    mk_settle_pnl = 0.0
    tk_settle_pnl = 0.0
    for o in outcomes:
        sp = settlement.get(o, 0.0)
        # split remaining lots by role for maker/taker settlement attribution.
        # Easier: run two separate FIFOs (one for MAKER-only buys, one for TAKER-only buys).
        # To keep this simple we approximate: attribute settle PnL to MAKER proportionally
        # to each role's buy notional. A rough but serviceable split.
        s_pnl = fifo[o].settle(sp)
        per_outcome_pnl[o] += s_pnl
        net = w_df.loc[(w_df["outcome"] == o) & buys_mask, "token_amount"].sum() \
            - w_df.loc[(w_df["outcome"] == o) & sells_mask, "token_amount"].sum()
        if net > 0:
            settlement_usdc += net * sp

    total_pnl = sum(per_outcome_pnl.values())

    # Reattribute settle PnL between maker/taker by role-share of remaining inventory
    # (ratio of maker vs taker buy tokens still unsold at end).
    # For simplicity: maker_pnl = mk_sell_pnl + settle * (mk_buy_share)
    mk_buy_tokens = 0.0
    tk_buy_tokens = 0.0
    for _, r in w_df.iterrows():
        if r["buyer"] != wl: continue
        role = roles.get(r["event_id"], "")
        if role == "MAKER": mk_buy_tokens += r["token_amount"]
        else: tk_buy_tokens += r["token_amount"]
    total_buy_tokens = mk_buy_tokens + tk_buy_tokens
    settle_pnl_total = total_pnl - (mk_sell_pnl + tk_sell_pnl)
    if total_buy_tokens > 0:
        mk_settle_share = mk_buy_tokens / total_buy_tokens
    else:
        mk_settle_share = 0.5
    maker_pnl = mk_sell_pnl + settle_pnl_total * mk_settle_share
    taker_pnl = tk_sell_pnl + settle_pnl_total * (1 - mk_settle_share)

    total_bought = w_df.loc[buys_mask, "usdc_amount"].sum()
    total_sold = w_df.loc[sells_mask, "usdc_amount"].sum()
    turnover = total_bought + total_sold

    # Max exposure
    cum = 0.0
    max_exposure = 0.0
    for _, r in w_df.iterrows():
        cum += r["usdc_amount"] if r["buyer"] == wl else -r["usdc_amount"]
        if cum > max_exposure:
            max_exposure = cum

    # Direction
    a_net = w_df.loc[(w_df["outcome"] == team_a) & buys_mask, "token_amount"].sum() \
          - w_df.loc[(w_df["outcome"] == team_a) & sells_mask, "token_amount"].sum()
    b_net = w_df.loc[(w_df["outcome"] == team_b) & buys_mask, "token_amount"].sum() \
          - w_df.loc[(w_df["outcome"] == team_b) & sells_mask, "token_amount"].sum()
    delta = a_net - b_net
    if delta > 0.1:
        dir_bias = f"LONG_{team_a[:3].upper()}"
    elif delta < -0.1:
        dir_bias = f"LONG_{team_b[:3].upper()}"
    else:
        dir_bias = "NEUTRAL"

    return {
        "wallet": wallet,
        "n_trades": total,
        "n_maker": fin_maker,
        "n_taker": fin_taker,
        "maker_pct": round(fin_maker / total * 100, 1),
        f"trades_{team_a[:3].upper()}": int((w_df["outcome"] == team_a).sum()),
        f"trades_{team_b[:3].upper()}": int((w_df["outcome"] == team_b).sum()),
        "turnover": round(turnover, 2),
        "max_exposure": round(max_exposure, 2),
        "capital_rotation": round(turnover / max_exposure, 2) if max_exposure > 0 else 0,
        f"net_{team_a[:3].upper()}": round(a_net, 2),
        f"net_{team_b[:3].upper()}": round(b_net, 2),
        "settlement_usdc": round(settlement_usdc, 2),
        "maker_pnl": round(maker_pnl, 2),
        "taker_pnl": round(taker_pnl, 2),
        "total_pnl": round(total_pnl, 2),
        "pnl_bps": round(total_pnl / turnover * 10000, 1) if turnover > 0 else 0,
        "roi_pct": round(total_pnl / max_exposure * 100, 2) if max_exposure > 0 else 0,
        "direction_bias": dir_bias,
        "first_trade_ist": s_to_ist(int(w_df["timestamp_unix"].iloc[0])),
        "last_trade_ist": s_to_ist(int(w_df["timestamp_unix"].iloc[-1])),
    }


# ── Main ─────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("db", help="Path to match_capture_*.db")
    ap.add_argument("--out-dir", default="/Users/sobhagyaxd/DeepWork/totem-mm/captures")
    ap.add_argument("--min-trades", type=int, default=100)
    ap.add_argument("--min-side", type=int, default=10)
    ap.add_argument("--winner", help="Override winner name")
    args = ap.parse_args()

    con = sqlite3.connect(args.db)
    cur = con.cursor()
    meta = dict(cur.execute("SELECT key, value FROM match_meta").fetchall())
    slug = meta["slug"]
    tokens = json.loads(meta["token_ids"])
    names = json.loads(meta["outcome_names"])
    T1, T2 = tokens[0], tokens[1]
    N1, N2 = names[0], names[1]
    start_ms = int(datetime.fromisoformat(meta["start_time"]).timestamp() * 1000)
    print(f"Slug: {slug}")
    print(f"Match: {N1} vs {N2}")

    # 1. Market metadata + settlement
    print("\n[1/6] Fetching market metadata...")
    market = fetch_market(slug)
    if args.winner:
        winner = args.winner
        settlement = {N1: 1.0 if winner == N1 else 0.0, N2: 1.0 if winner == N2 else 0.0}
    else:
        winner, settlement = parse_settlement(market, [N1, N2])
    print(f"  Winner: {winner}")
    print(f"  Settlement: {N1}=${settlement[N1]}  {N2}=${settlement[N2]}")

    # 2. Goldsky events
    print("\n[2/6] Fetching all OrderFilled events from Goldsky...")
    events = fetch_all_events(tokens)
    if not events:
        print("No events returned. Aborting.")
        sys.exit(1)

    # 3. Process into DataFrame
    print("\n[3/6] Processing events into trades DataFrame...")
    df = process_events(events, tokens, [N1, N2])
    print(f"  trades: {len(df)}  range: {s_to_ist(int(df['timestamp_unix'].iloc[0]))} "
          f"→ {s_to_ist(int(df['timestamp_unix'].iloc[-1]))}")

    # 4. Role classification + wallet filter + PnL
    print("\n[4/6] Classifying roles + filtering wallets...")
    roles = classify_all_roles(df)
    mk_n = sum(1 for v in roles.values() if v == "MAKER")
    print(f"  MAKER events: {mk_n}  TAKER events: {len(roles) - mk_n}")

    all_wallets = df[df["maker"] != CTF_EXCHANGE]["maker"].unique()
    print(f"  distinct maker-field wallets: {len(all_wallets)}")

    qualified = []
    for w in all_wallets:
        w_df = df[df["maker"] == w]
        n_total = len(w_df)
        if n_total < args.min_trades: continue
        n_a = (w_df["outcome"] == N1).sum()
        n_b = (w_df["outcome"] == N2).sum()
        if n_a < args.min_side or n_b < args.min_side: continue
        qualified.append(w)
    print(f"  qualified: {len(qualified)} wallets "
          f"(>={args.min_trades} trades AND >={args.min_side} on each team)")

    wallet_rows = []
    for w in qualified:
        w_df = df[df["maker"] == w].sort_values(["timestamp_unix", "event_id"]).reset_index(drop=True)
        s = compute_wallet_summary(w, w_df, roles, [N1, N2], settlement, N1, N2)
        wallet_rows.append(s)
    wallet_df = pd.DataFrame(wallet_rows)
    if not wallet_df.empty:
        wallet_df = wallet_df.sort_values("turnover", ascending=False).reset_index(drop=True)

    # 5. Read capture DB for events + book snapshots
    print("\n[5/6] Reading capture DB for events + book snapshots...")
    cricket_events = cur.execute("""
        SELECT id, local_ts_ms, signal_type, runs, wickets, overs, score_str, innings
        FROM cricket_events WHERE signal_type IN ('W','4','6')
        ORDER BY local_ts_ms
    """).fetchall()
    print(f"  W/4/6 events: {len(cricket_events)}")

    snaps = {T1: [], T2: []}
    for row in cur.execute("""
        SELECT asset_id, local_ts_ms,
               bid3_p, bid3_s, bid2_p, bid2_s, bid1_p, bid1_s,
               ask1_p, ask1_s, ask2_p, ask2_s, ask3_p, ask3_s,
               mid_price, spread, total_bid_depth, total_ask_depth
        FROM book_snapshots WHERE asset_id IN (?,?) ORDER BY local_ts_ms
    """, (T1, T2)):
        snaps[row[0]].append(row[1:])
    print(f"  book snaps: T1={len(snaps[T1])}  T2={len(snaps[T2])}")

    def snap_at(asset, ts):
        arr = snaps[asset]
        lo, hi = 0, len(arr)
        while lo < hi:
            m = (lo+hi)//2
            if arr[m][0] <= ts: lo = m+1
            else: hi = m
        return arr[lo-1] if lo > 0 else None

    # 6. Build workbook
    print("\n[6/6] Building workbook...")
    wb = Workbook(); wb.remove(wb.active)
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    sub_fill = PatternFill("solid", fgColor="FFF2CC")

    N1S = (N1.split()[0] if " " in N1 else N1)[:3].upper()
    N2S = (N2.split()[0] if " " in N2 else N2)[:3].upper()

    # Sheet 1: special events
    ws1 = wb.create_sheet("1_special_events")
    hdr1 = ["event_id", "ist_time", "type", "runs", "wickets", "overs", "score", "innings", "t_from_start_min"]
    for i, h in enumerate(hdr1):
        c = ws1.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill
    for i, ev in enumerate(cricket_events, 2):
        mins = round((ev[1] - start_ms) / 60000, 1) if ev[1] else None
        vals = [ev[0], ms_to_ist(ev[1]), ev[2], ev[3], ev[4], ev[5], ev[6], ev[7], mins]
        for j, v in enumerate(vals):
            ws1.cell(row=i, column=j+1, value=v)
    ws1.freeze_panes = "A2"
    ws1.column_dimensions["B"].width = 16

    # Sheet 2: event books
    ws2 = wb.create_sheet("2_event_books")
    OFFSETS = [-120_000, -60_000, -30_000, -15_000, -10_000, -5_000, -3_000, -1_000,
               0, 1_000, 3_000, 5_000, 10_000, 15_000, 30_000, 60_000]
    snap_headers = ["offset", "token",
                    "bid3","b3sz","bid2","b2sz","bid1","b1sz",
                    "ask1","a1sz","ask2","a2sz","ask3","a3sz",
                    "mid","spread","bid_depth","ask_depth"]
    row_idx = 1
    for ev in cricket_events:
        ev_id, ev_ts, stype = ev[0], ev[1], ev[2]
        title = f"Event #{ev_id}  type={stype}  time={ms_to_ist(ev_ts)}  score={ev[6]}  overs={ev[5]}"
        c = ws2.cell(row=row_idx, column=1, value=title)
        c.font = Font(bold=True, size=11); c.fill = sub_fill
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(snap_headers))
        row_idx += 1
        for j, h in enumerate(snap_headers):
            c = ws2.cell(row=row_idx, column=j+1, value=h); c.font = bold; c.fill = hdr_fill
        row_idx += 1
        for off in OFFSETS:
            target = ev_ts + off
            for tok, lbl in [(T1, N1S), (T2, N2S)]:
                s = snap_at(tok, target)
                if s is None: continue
                vals = [f"{off//1000:+d}s", lbl,
                        s[1], s[2], s[3], s[4], s[5], s[6],
                        s[7], s[8], s[9], s[10], s[11], s[12],
                        round(s[13],4) if s[13] else None,
                        round(s[14],4) if s[14] else None,
                        round(s[15],1) if s[15] else None,
                        round(s[16],1) if s[16] else None]
                for j, v in enumerate(vals):
                    ws2.cell(row=row_idx, column=j+1, value=v)
                row_idx += 1
        row_idx += 1
    for L in "ABCDEFGHIJKLMNOPQR":
        ws2.column_dimensions[L].width = 9
    ws2.column_dimensions["A"].width = 10

    # Sheet 3: real trades (Goldsky)
    ws3 = wb.create_sheet("3_real_trades")
    hdr3 = ["ist_time", "outcome", "maker_side", "price", "tokens", "usdc",
            "fee_usdc", "role", "maker", "taker", "tx"]
    for i, h in enumerate(hdr3):
        c = ws3.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill
    for i, r in enumerate(df.itertuples(index=False), 2):
        role = roles.get(r.event_id, "")
        vals = [s_to_ist(r.timestamp_unix), r.outcome, r.maker_side,
                r.price, r.token_amount, r.usdc_amount, r.fee_usdc,
                role,
                (r.maker[:14]+"...") if r.maker else "",
                (r.taker[:14]+"...") if r.taker else "",
                (r.tx_hash[:14]+"...") if r.tx_hash else ""]
        for j, v in enumerate(vals):
            ws3.cell(row=i, column=j+1, value=v)
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["I"].width = 20
    ws3.column_dimensions["J"].width = 20
    print(f"  sheet 3: {ws3.max_row-1} rows")

    # Sheet 4: wallets
    ws4 = wb.create_sheet("4_wallets")
    if wallet_df.empty:
        ws4.cell(row=1, column=1, value="No wallets qualified").font = bold
    else:
        cols = list(wallet_df.columns)
        for i, h in enumerate(cols):
            c = ws4.cell(row=1, column=i+1, value=h); c.font = bold; c.fill = hdr_fill
        for ri, row in enumerate(wallet_df.itertuples(index=False), 2):
            for j, v in enumerate(row):
                if isinstance(v, str) and v.startswith("0x") and len(v) == 42:
                    v = v[:14] + "..."
                ws4.cell(row=ri, column=j+1, value=v)
        ws4.freeze_panes = "A2"
        ws4.column_dimensions["A"].width = 20
    print(f"  sheet 4: {max(0, ws4.max_row-1)} wallets")

    ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"{args.out_dir}/master_{slug}_{ts_now}.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

    # Console summary of top wallets
    if not wallet_df.empty:
        print(f"\n  TOP 10 WALLETS BY PNL:")
        top = wallet_df.sort_values("total_pnl", ascending=False).head(10)
        for _, r in top.iterrows():
            print(f"    {r['wallet'][:14]}...  n={r['n_trades']:>4}  "
                  f"mk%={r['maker_pct']:>5.1f}  turnover=${r['turnover']:>12,.0f}  "
                  f"pnl=${r['total_pnl']:>+10,.0f}  bps={r['pnl_bps']:>+6.0f}  "
                  f"rot={r['capital_rotation']:>5.2f}x  dir={r['direction_bias']}")


if __name__ == "__main__":
    main()
