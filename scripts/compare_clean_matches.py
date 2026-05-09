"""Run the timing comparison on the cleanest matches only and produce a
direct verdict on which feed is faster.

CLEAN matches = those with full ball-by-ball coverage in both feeds, no
opener gap in capture, no known feed blackout, and not a rescheduled-rain
artifact. From earlier analysis:

  EXCLUDE:
    raj-mum-2026-04-07  -> rescheduled, ~4671s offset artifact
    kol-raj-2026-04-19  -> 6.5-min cricket-events blackout in inn2
    raj-roy-2026-04-10  -> innings 2 truncated
    del-pun-04-25       -> capture started at 2.0 (missed 12 balls inn1)
    guj-mum-04-20       -> capture started at 4.0 (missed 24 balls inn1)
    mum-che-04-23       -> capture started at 2.0
    pun-sun-04-11       -> capture started at 1.4
    sun-che-04-18       -> capture started at 0.2

  KEEP (clean both innings):
    che-kol, kol-luc, pun-luc                (perfect 120 legal each side)
    guj-kol, kol-raj-inn1-only, mum-pun, raj-sun, roy-del,
    roy-guj, roy-luc, sun-del, luc-raj
"""
from __future__ import annotations

import statistics
from pathlib import Path

from openpyxl import load_workbook

import sys
sys.path.insert(0, str(Path(__file__).parent))
from compare_espn_vs_capture_timing import align_match  # type: ignore

WORKBOOK = Path("/Users/sobhagyaxd/DeepWork/totem-mm/captures/espn_ipl2026_ballbyball.xlsx")

CLEAN_SLUGS = [
    "che-kol-2026-04-14",
    "kol-luc-2026-04-09",
    "pun-luc-2026-04-19",
    "guj-kol-2026-04-17",
    "luc-raj-2026-04-22",
    "mum-pun-2026-04-16",
    "raj-sun-2026-04-25",
    "roy-del-2026-04-18",
    "roy-guj-2026-04-24",
    "roy-luc-2026-04-15",
    "sun-del-2026-04-21",
]


def percentile(s: list[float], p: float) -> float:
    if not s:
        return float("nan")
    k = (len(s) - 1) * p
    f = int(k)
    c = min(f + 1, len(s) - 1)
    if f == c:
        return s[f]
    return s[f] + (s[c] - s[f]) * (k - f)


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=True)

    print(f"Clean matches scanned: {len(CLEAN_SLUGS)}")
    print()
    print(f"{'match':<28} {'matched':>8} {'p10':>8} {'p50':>8} "
          f"{'mean':>8} {'p90':>8} {'min':>8} {'max':>8}")
    print(f"{'-' * 28} {'-' * 8} {'-' * 8} {'-' * 8} {'-' * 8} {'-' * 8} "
          f"{'-' * 8} {'-' * 8}")

    pooled: list[float] = []
    pooled_4: list[float] = []
    pooled_6: list[float] = []
    pooled_w: list[float] = []
    capture_faster = 0
    espn_faster = 0
    tie = 0

    for slug in CLEAN_SLUGS:
        if slug not in wb.sheetnames:
            print(f"  {slug}: missing tab")
            continue
        result = align_match(slug, wb[slug])
        if not result["matched"]:
            print(f"  {slug}: no overlap")
            continue
        deltas = [m["delta_ms"] / 1000 for m in result["matched"]]
        s = sorted(deltas)
        print(f"{slug:<28} {len(deltas):>8} "
              f"{percentile(s, 0.10):>+8.1f} {percentile(s, 0.50):>+8.1f} "
              f"{statistics.mean(s):>+8.1f} {percentile(s, 0.90):>+8.1f} "
              f"{s[0]:>+8.1f} {s[-1]:>+8.1f}")
        pooled.extend(deltas)
        for m in result["matched"]:
            d = m["delta_ms"] / 1000
            pt = (m["espn_play_type"] or "").lower()
            if pt == "four":
                pooled_4.append(d)
            elif pt == "six":
                pooled_6.append(d)
            elif pt in ("wicket", "out"):
                pooled_w.append(d)
            if d > 0.5:
                espn_faster += 1
            elif d < -0.5:
                capture_faster += 1
            else:
                tie += 1

    s = sorted(pooled)
    print()
    print(f"{'AGGREGATE (clean only)':<28} {len(s):>8} "
          f"{percentile(s, 0.10):>+8.1f} {percentile(s, 0.50):>+8.1f} "
          f"{statistics.mean(s):>+8.1f} {percentile(s, 0.90):>+8.1f} "
          f"{s[0]:>+8.1f} {s[-1]:>+8.1f}")


    print()
    print("Per event-type aggregate (clean matches only):")
    for label, vals in [("FOUR", pooled_4), ("SIX", pooled_6), ("WICKET", pooled_w)]:
        if not vals:
            continue
        ss = sorted(vals)
        print(f"  {label:<8} n={len(ss):>4}  p10={percentile(ss, 0.10):+6.1f}  "
              f"p50={percentile(ss, 0.50):+6.1f}  mean={statistics.mean(ss):+6.1f}  "
              f"p90={percentile(ss, 0.90):+6.1f}")

    total = capture_faster + espn_faster + tie
    print()
    print(f"Win-rate over {total} matched balls (±0.5s threshold):")
    print(f"  ESPN faster   : {espn_faster:>5}  ({100 * espn_faster / total:5.1f}%)")
    print(f"  Capture faster: {capture_faster:>5}  ({100 * capture_faster / total:5.1f}%)")
    print(f"  Tie (within 0.5s): {tie:>2}  ({100 * tie / total:5.1f}%)")

    p50 = percentile(s, 0.50)
    print()
    print("=" * 70)
    print("VERDICT")
    print("=" * 70)
    if p50 > 1:
        print(f"ESPN is faster. Median lead: {p50:.1f}s.")
    elif p50 < -1:
        print(f"Our capture is faster. Median lead: {-p50:.1f}s.")
    else:
        print(f"Tie. Median delta: {p50:.1f}s.")


if __name__ == "__main__":
    main()
