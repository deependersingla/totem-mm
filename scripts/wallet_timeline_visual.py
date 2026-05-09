#!/usr/bin/env python3
"""
Colored timeline per match for a wallet.

Rows = every wallet fill + every cricket event, time-sorted IST.
Colors:
  light GREEN  = BUY
  light RED    = SELL
  GOLD         = WICKET
  LIGHT GOLD   = boundary (4 / 6)
  (runs / dots / extras = uncolored)

Columns include real-time mark-to-market PnL:
    cash = running cash flow (+SELL, -BUY)
    inv_A, inv_B = running token inventory deltas (from fills only)
    mid_A, mid_B = market mid at that moment
    mtm_pnl = cash + inv_A*mid_A + inv_B*mid_B
    (MtM is invariant to pre-window inventory — a sell of pre-existing
     tokens at market mid nets to zero; only above/below-mid trades
     register as P&L.)
"""
import argparse
import json
import os
import sqlite3
from bisect import bisect_left
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

USDC = 1e6
IST = timezone(timedelta(hours=5, minutes=30))
CTF = "0x4bfb41d5b3570defd03c39a9a4d8de6bd8b8982e"


def ist(ts):
    return datetime.fromtimestamp(ts, tz=timezone.utc).astimezone(IST).strftime("%H:%M:%S")


def build_timeline(wallet, db_path, cache_path, out_xlsx):
    conn = sqlite3.connect(db_path)
    meta = dict(conn.execute("SELECT key, value FROM match_meta"))
    slug = meta["slug"]
    tids = json.loads(meta["token_ids"])
    names = json.loads(meta["outcome_names"])
    t2o = dict(zip(tids, names))
    tidA, tidB = tids[0], tids[1]
    nameA, nameB = names[0], names[1]

    cache = json.load(open(cache_path))
    w = wallet.lower()

    # Pre-cache book snapshots for both tokens for fast lookup
    book = {}
    for tok in tids:
        rows = list(conn.execute(
            "SELECT local_ts_ms, mid_price, bid1_p, ask1_p FROM book_snapshots "
            "WHERE asset_id=? ORDER BY local_ts_ms", (tok,)
        ))
        book[tok] = {
            "ts": [r[0] for r in rows],
            "mid": [r[1] for r in rows],
            "bid": [r[2] for r in rows],
            "ask": [r[3] for r in rows],
        }

    def lookup(tok, ts_ms):
        c = book[tok]
        if not c["ts"]:
            return None, None, None
        i = bisect_left(c["ts"], ts_ms) - 1
        if i < 0:
            return None, None, None
        return c["mid"][i], c["bid"][i], c["ask"][i]

    # Process fills
    fill_rows = []
    for ev in cache["fills"]:
        maker = ev["maker"].lower()
        taker = ev["taker"].lower()
        m_amt = int(ev["makerAmountFilled"])
        t_amt = int(ev["takerAmountFilled"])
        ts = int(ev["timestamp"])
        ma, ta = ev["makerAssetId"], ev["takerAssetId"]

        if ma in t2o:
            outcome = t2o[ma]
            tok_id = ma
            tokens = m_amt / USDC
            usdc_v = t_amt / USDC
            maker_side = "SELL"
        elif ta in t2o:
            outcome = t2o[ta]
            tok_id = ta
            tokens = t_amt / USDC
            usdc_v = m_amt / USDC
            maker_side = "BUY"
        else:
            continue
        price = usdc_v / tokens if tokens else 0

        if maker == w:
            role = "MAKER"
            side = maker_side
            cp = taker
        elif taker == w:
            role = "TAKER"
            side = "BUY" if maker_side == "SELL" else "SELL"
            cp = maker
        else:
            continue

        ts_ms = ts * 1000
        mid_s, bid_s, ask_s = lookup(tok_id, ts_ms)
        fill_rows.append({
            "ts": ts,
            "time_ist": ist(ts),
            "type": "FILL",
            "side": side,
            "role": role,
            "outcome": outcome,
            "tok_id": tok_id,
            "price": round(price, 4),
            "tokens": round(tokens, 2),
            "notional_$": round(usdc_v, 2),
            "mid": round(mid_s, 4) if mid_s is not None else None,
            "bid": round(bid_s, 4) if bid_s is not None else None,
            "ask": round(ask_s, 4) if ask_s is not None else None,
            "counterparty": cp[:10] if cp != CTF else "CTF_EXCH",
            "score": "",
            "signal": "",
        })

    # Cricket events
    cric_rows = []
    for r in conn.execute(
        "SELECT local_ts_ms/1000 AS ts, signal_type, score_str "
        "FROM cricket_events ORDER BY ts"
    ):
        ts, sig, score = r
        cric_rows.append({
            "ts": ts,
            "time_ist": ist(ts),
            "type": "CRICKET",
            "side": "",
            "role": "",
            "outcome": "",
            "tok_id": "",
            "price": None,
            "tokens": None,
            "notional_$": None,
            "mid": None,
            "bid": None,
            "ask": None,
            "counterparty": "",
            "score": score,
            "signal": sig,
        })

    # Merge + sort
    all_rows = sorted(fill_rows + cric_rows, key=lambda x: x["ts"])

    # Walk forward computing: cash flow, token inventory (A, B), MtM PnL
    cash = 0.0
    inv_A = 0.0
    inv_B = 0.0
    for r in all_rows:
        if r["type"] == "FILL":
            if r["side"] == "BUY":
                cash -= r["notional_$"]
                if r["tok_id"] == tidA:
                    inv_A += r["tokens"]
                else:
                    inv_B += r["tokens"]
            else:  # SELL
                cash += r["notional_$"]
                if r["tok_id"] == tidA:
                    inv_A -= r["tokens"]
                else:
                    inv_B -= r["tokens"]

        # Snapshot the current mid for BOTH tokens to compute MtM
        ts_ms = r["ts"] * 1000
        mid_A, _, _ = lookup(tidA, ts_ms)
        mid_B, _, _ = lookup(tidB, ts_ms)
        mtm = cash
        if mid_A is not None:
            mtm += inv_A * mid_A
        if mid_B is not None:
            mtm += inv_B * mid_B
        r["cash_$"] = round(cash, 2)
        r["inv_A"] = round(inv_A, 1)
        r["inv_B"] = round(inv_B, 1)
        r["mid_A"] = round(mid_A, 4) if mid_A is not None else None
        r["mid_B"] = round(mid_B, 4) if mid_B is not None else None
        r["mtm_pnl_$"] = round(mtm, 2)

    # === Write xlsx ===
    wb = Workbook()
    ws = wb.active
    ws.title = slug[:31]

    header = [
        "time_ist", "type", "side", "role", "outcome",
        "price", "tokens", "notional_$", "mid", "bid", "ask",
        "cash_$", "inv_A", "inv_B", "mid_A", "mid_B", "mtm_pnl_$",
        "counterparty", "score", "signal",
    ]

    # Summary rows
    n_fills = len(fill_rows)
    buys = sum(1 for f in fill_rows if f["side"] == "BUY")
    sells = sum(1 for f in fill_rows if f["side"] == "SELL")
    buy_not = sum(f["notional_$"] for f in fill_rows if f["side"] == "BUY")
    sell_not = sum(f["notional_$"] for f in fill_rows if f["side"] == "SELL")
    final_pnl = all_rows[-1]["mtm_pnl_$"] if all_rows else 0
    max_pnl = max((r["mtm_pnl_$"] for r in all_rows), default=0)
    min_pnl = min((r["mtm_pnl_$"] for r in all_rows), default=0)

    summary_fill = PatternFill("solid", fgColor="FFFFE0")
    bold = Font(bold=True)
    ws.cell(row=1, column=1, value=f"{slug} | A={nameA} | B={nameB} | wallet={wallet[:10]}").font = bold
    ws.cell(row=2, column=1,
            value=f"fills={n_fills}  (BUY {buys} ${buy_not:,.0f} / SELL {sells} ${sell_not:,.0f})  "
                  f"events={len(cric_rows)}  "
                  f"MtM P&L end=${final_pnl:,.0f}  max=${max_pnl:,.0f}  min=${min_pnl:,.0f}").font = bold
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).fill = summary_fill
        ws.cell(row=2, column=c).fill = summary_fill

    # Header row
    for i, h in enumerate(header, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="E0E0E0")
        cell.alignment = Alignment(horizontal="center")

    # Two colors for fills, two shades of gold for events
    buy_fill = PatternFill("solid", fgColor="C8E8C8")
    sell_fill = PatternFill("solid", fgColor="F4C8C8")
    wicket_fill = PatternFill("solid", fgColor="FFB800")
    boundary_fill = PatternFill("solid", fgColor="FFE680")

    for row_idx, r in enumerate(all_rows, start=5):
        for col_idx, h in enumerate(header, 1):
            ws.cell(row=row_idx, column=col_idx, value=r.get(h))

        if r["type"] == "FILL":
            fill = buy_fill if r["side"] == "BUY" else sell_fill
        else:
            if r["signal"] == "W":
                fill = wicket_fill
            elif r["signal"] in ("4", "6"):
                fill = boundary_fill
            else:
                fill = None

        if fill is not None:
            for col_idx in range(1, len(header) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    widths = {"time_ist": 9, "type": 8, "side": 6, "role": 7, "outcome": 22,
              "price": 7, "tokens": 9, "notional_$": 10, "mid": 7, "bid": 7, "ask": 7,
              "cash_$": 10, "inv_A": 9, "inv_B": 9, "mid_A": 7, "mid_B": 7, "mtm_pnl_$": 11,
              "counterparty": 11, "score": 14, "signal": 7}
    for i, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 10)

    ws.freeze_panes = "A5"
    wb.save(out_xlsx)

    return {
        "out": out_xlsx, "n_fills": n_fills, "n_events": len(cric_rows),
        "buy_not": buy_not, "sell_not": sell_not,
        "final_mtm_pnl": final_pnl, "max_mtm": max_pnl, "min_mtm": min_pnl,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--wallet", required=True)
    ap.add_argument("--db", required=True)
    ap.add_argument("--cache", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()
    r = build_timeline(args.wallet, args.db, args.cache, args.out)
    print(f"✓ {args.out} | fills={r['n_fills']} events={r['n_events']} | "
          f"MtM P&L end=${r['final_mtm_pnl']:,.0f}")


if __name__ == "__main__":
    main()
